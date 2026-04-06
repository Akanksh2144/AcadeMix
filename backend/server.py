from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).resolve().parent
load_dotenv(ROOT_DIR / '.env')

import os
import bcrypt
import jwt
import secrets
import httpx
import subprocess
import tempfile
from datetime import datetime, timezone, timedelta
from typing import Optional, List
# bson removed — MongoDB is archived
from fastapi import FastAPI, HTTPException, Request, Response, Depends, Query, UploadFile, File, Form
from fastapi.responses import HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
# certifi removed — was for MongoDB TLS
import io
import csv

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import or_, and_, delete, update, func, text, desc
from sqlalchemy.orm import selectinload
from database import get_db
import models

DEFAULT_GRADE_SCALE = [
    {"min_pct": 90, "grade": "O", "points": 10},
    {"min_pct": 80, "grade": "A+", "points": 9},
    {"min_pct": 70, "grade": "A", "points": 8},
    {"min_pct": 60, "grade": "B+", "points": 7},
    {"min_pct": 50, "grade": "B", "points": 6},
    {"min_pct": 45, "grade": "C", "points": 5},
    {"min_pct": 40, "grade": "D", "points": 4},
    {"min_pct": 0, "grade": "F", "points": 0}
]
# tenant.py archived to backend/legacy/ — functions inlined or removed below

import sentry_sdk
from sentry_sdk.integrations.fastapi import FastApiIntegration

import redis as pyredis
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

# ─── App & DB ───────────────────────────────────────────────────────────────
app = FastAPI(title="QuizPortal API")

sentry_dsn = os.environ.get("SENTRY_DSN", "")

def _scrub_pii(event, hint):
    """Aggressively strip PII from the event before it's sent to Sentry"""
    if 'request' in event and 'data' in event['request']:
        data = event['request']['data']
        if isinstance(data, dict):
            for pii_key in ['email', 'password', 'name', 'college_id']:
                if pii_key in data:
                    data[pii_key] = '[FILTERED]'
    return event

if sentry_dsn:
    sentry_sdk.init(
        dsn=sentry_dsn,
        enable_tracing=True,
        traces_sample_rate=1.0,
        send_default_pii=False,
        before_send=_scrub_pii,
        integrations=[FastApiIntegration()],
    )

# MongoDB Removed - Using SQLAlchemy PostgreSQL via Supabase
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGORITHM = "HS256"

frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
code_runner_url = os.environ.get("CODE_RUNNER_URL", "https://acadmix-code-runner.fly.dev")
cors_origins_env = os.environ.get("CORS_ORIGINS", "")
if not cors_origins_env:
    raise ValueError("CORS_ORIGINS must be explicitly set (no wildcard allowed)")

origins = [o.strip() for o in cors_origins_env.split(",") if o.strip()]

if "*" in origins:
    raise ValueError("CORS_ORIGINS cannot contain wildcard '*'. List origins explicitly.")

if os.getenv("ENVIRONMENT") == "production":
    for origin in origins:
        if not origin.startswith("https://"):
            raise ValueError(f"CORS origin '{origin}' must use HTTPS in production")

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "ngrok-skip-browser-warning"],
    max_age=3600,
)

redis_url = os.environ.get("REDIS_URL", "")
redis_client = pyredis.from_url(redis_url) if redis_url else None
if redis_url:
    limiter = Limiter(key_func=get_remote_address, storage_uri=redis_url)
else:
    limiter = Limiter(key_func=get_remote_address)

app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# ─── Helpers ────────────────────────────────────────────────────────────────
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

import uuid

class TokenBlacklistConfig:
    USE_BLACKLIST = os.getenv("USE_TOKEN_BLACKLIST", "false").lower() == "true"
    ACCESS_TOKEN_TTL_MINUTES = 1440  # 24 hours — frontend has no refresh flow yet
    REFRESH_TOKEN_TTL_DAYS = 7
    BLACKLIST_CHECK_REDIS = True

def create_access_token(user_id: str, role: str, tenant_id: str = "", permissions: dict = None) -> str:
    perms = permissions or {}
    return jwt.encode({
        "sub": user_id, 
        "role": role, 
        "tenant_id": tenant_id, 
        "permissions": perms, 
        "exp": datetime.now(timezone.utc) + timedelta(minutes=TokenBlacklistConfig.ACCESS_TOKEN_TTL_MINUTES), 
        "type": "access",
        "jti": str(uuid.uuid4())
    }, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    return jwt.encode({
        "sub": user_id, 
        "exp": datetime.now(timezone.utc) + timedelta(days=TokenBlacklistConfig.REFRESH_TOKEN_TTL_DAYS), 
        "type": "refresh",
        "jti": str(uuid.uuid4())
    }, JWT_SECRET, algorithm=JWT_ALGORITHM)

# serialize_doc removed — was MongoDB-specific dead code

def grade_to_points(grade: str) -> float:
    mapping = {"O": 10, "A+": 9, "A": 8, "B+": 7, "B": 6, "C": 5, "D": 4, "F": 0}
    return mapping.get(grade, 0)

async def get_current_user(request: Request, session: AsyncSession = Depends(get_db)) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
            
        if TokenBlacklistConfig.USE_BLACKLIST and redis_client:
            jti = payload.get("jti")
            if jti and redis_client.exists(f"revoked_access:{jti}"):
                raise HTTPException(status_code=401, detail="Token revoked")
        
        result = await session.execute(select(models.User).where(models.User.id == payload["sub"]))
        user = result.scalars().first()
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
            
        user_dict = {
            "id": user.id,
            "role": user.role,
            "email": user.email,
            "name": user.name,
            "tenant_id": user.college_id,
            "college_id": user.college_id,
            "permissions": payload.get("permissions", {})
        }
        if user.profile_data:
            user_dict.update(user.profile_data)
            
        # tenant_id is always set from college_id via the SQLAlchemy user row — fallback to empty string
        if "tenant_id" not in user_dict:
            user_dict["tenant_id"] = user_dict.get("college_id", "")
        
        import database
        database.tenant_context.set(user.college_id)
            
        return user_dict
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except (jwt.InvalidTokenError, Exception):
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(*roles):
    async def check(request: Request, session: AsyncSession = Depends(get_db)):
        user = await get_current_user(request, session)
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return check

def require_permission(module: str, action: str):
    async def check(request: Request, session: AsyncSession = Depends(get_db)):
        user = await get_current_user(request, session)
        if user["role"] in ["super_admin", "admin"]:
            return user
        
        perms = user.get("permissions", {})
        module_perms = perms.get(module, [])
        if action not in module_perms:
            role = user.get("role")
            if module == "quizzes" and role in ["teacher"]: return user
            if module == "marks" and role in ["teacher", "exam_cell", "hod"]: return user
            if module == "placements" and role in ["hod", "admin"]: return user
            raise HTTPException(status_code=403, detail=f"Insufficient permissions: requires {module}.{action}")
        return user
    return check

# ─── Pydantic Models ───────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    college_id: str
    password: str

class RegisterRequest(BaseModel):
    name: str = Field(..., max_length=150)
    college_id: str = Field(..., max_length=50)
    email: str = Field(..., max_length=255)
    password: str = Field(..., min_length=6, max_length=100)
    role: str = Field("student", max_length=30)
    college: str = Field("GNITC", max_length=50)
    department: str = Field("", max_length=100)
    batch: str = Field("", max_length=20)
    section: str = Field("", max_length=20)

class StudentProfileData(BaseModel):
    batch: Optional[str] = None
    section: Optional[str] = None
    department: Optional[str] = None
    first_graduate: Optional[bool] = None
    community: Optional[str] = None
    blood_group: Optional[str] = None
    hostel_required: Optional[bool] = None
    transport_required: Optional[bool] = None
    aadhaar_number: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    languages_known: Optional[list] = None

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None
    college_id: Optional[str] = None # roll number / faculty ID
    department: Optional[str] = None
    batch: Optional[str] = None
    section: Optional[str] = None
    password: Optional[str] = None
    profile_data: Optional[StudentProfileData] = None

class DepartmentCreate(BaseModel):
    name: str = Field(..., max_length=150)
    code: str = Field(..., max_length=20)

class DepartmentUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=150)
    code: Optional[str] = Field(None, max_length=20)

class SectionCreate(BaseModel):
    department_id: str
    name: str = Field(..., max_length=50)

class SectionUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=50)
    department_id: Optional[str] = None

class RoleCreate(BaseModel):
    name: str = Field(..., max_length=50)
    permissions: dict = {}

class RoleUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=50)
    permissions: Optional[dict] = None

class QuizCreate(BaseModel):
    title: str = Field(..., min_length=3, max_length=200)
    subject: str = Field(..., min_length=2, max_length=200)
    description: str = Field("", max_length=2000)
    total_marks: float = Field(0.0, ge=0.0, le=1000.0)
    duration_mins: int = Field(60, ge=1, le=480)
    negative_marking: bool = False
    timed: bool = True
    randomize_questions: bool = False
    randomize_options: bool = False
    show_answers_after: bool = True
    allow_reattempt: bool = False
    assigned_classes: list = []
    negative_marks: float = Field(0.0, ge=0.0, le=10.0)
    questions: list = []

class AnswerSubmit(BaseModel):
    question_index: int
    answer: object

class SemesterResultCreate(BaseModel):
    student_id: str
    semester: int
    subjects: list
    sgpa: float
    cgpa: float

class CodeExecuteRequest(BaseModel):
    code: str = Field(..., max_length=15000)
    language: str = Field("python", max_length=50)
    test_input: str = Field("", max_length=5000)

class FacultyAssignment(BaseModel):
    teacher_id: str
    subject_code: str
    subject_name: str
    department: str
    batch: str
    section: str
    semester: int = 1
    academic_year: str
    credits: Optional[int] = None
    hours_per_week: Optional[int] = None
    is_lab: bool = False

class SubjectAllocationUpdate(BaseModel):
    credits: Optional[int] = None
    hours_per_week: Optional[int] = None
    is_lab: Optional[bool] = None

class PermissionFlagsUpdate(BaseModel):
    flags: dict  # e.g. {"can_create_timetable": true, "is_subject_expert": false}

class CIATemplateCreate(BaseModel):
    name: str = Field(..., min_length=2, max_length=200)
    description: Optional[str] = None
    total_marks: int = Field(..., ge=1, le=200)
    components: list  # list of component dicts

class CIATemplateUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=200)
    description: Optional[str] = None
    total_marks: Optional[int] = Field(None, ge=1, le=200)
    components: Optional[list] = None

class SubjectCIAConfigCreate(BaseModel):
    subject_code: str
    subject_name: Optional[str] = None
    academic_year: str
    semester: int = Field(..., ge=1, le=8)
    template_id: str

class MarkEntryItem(BaseModel):
    student_id: str
    college_id: str
    student_name: str
    marks: Optional[float] = None

class MarkEntrySave(BaseModel):
    assignment_id: str
    exam_type: str  # mid1 or mid2
    component_id: Optional[str] = None
    semester: int
    max_marks: float = Field(30, gt=0, le=200)
    entries: List[MarkEntryItem] = Field(..., max_items=5000)
    revision_reason: Optional[str] = None

class MarkReview(BaseModel):
    action: str  # approve or reject
    remarks: str = ""

class EndtermEntry(BaseModel):
    subject_code: str
    subject_name: str
    department: str
    batch: str
    section: str
    semester: int
    max_marks: float = 100
    entries: list

class TimetableSlot(BaseModel):
    section: str
    day: str  # Mon, Tue, Wed, Thu, Fri, Sat
    period: int  # 1-6
    subject_code: str
    subject_name: str
    teacher_id: str
    teacher_name: str
    semester: int = 3

class AnnouncementCreate(BaseModel):
    title: str
    message: str
    priority: str = "info"  # info, warning, urgent
    visibility: str = "all"  # all, faculty, students

from pydantic import validator

class CollegeSettingsUpdate(BaseModel):
    settings: dict

    @validator("settings")
    def validate_settings(cls, v):
        # 1. Advanced Grade Scale Validator
        scale = v.get("grade_scale")
        if scale is not None:
            if not isinstance(scale, list) or len(scale) == 0:
                raise ValueError("grade_scale must be a non-empty list of dictionaries")
            prev_pct = 101.0
            has_zero = False
            for item in scale:
                pct = item.get("min_pct")
                if pct is None:
                    raise ValueError("Each grade scale item must have min_pct")
                if float(pct) >= prev_pct:
                    raise ValueError(f"grade_scale min_pct must be strictly monotonically decreasing. Violating element: {pct}")
                if float(pct) <= 0:
                    has_zero = True
                prev_pct = float(pct)
            if not has_zero:
                raise ValueError("grade_scale must end exactly at or below min_pct=0")
                
        # 2. Strict Attendance Bounds Validator
        if "attendance_min_pct" in v:
            val = v["attendance_min_pct"]
            if not isinstance(val, (int, float)):
                raise ValueError("attendance_min_pct must be a number")
            if not (1 <= float(val) <= 100):
                raise ValueError("attendance_min_pct must be between 1 and 100")
                
        # 3. Late Entry Validator
        if "late_entry_window_hours" in v:
            val = v["late_entry_window_hours"]
            if not isinstance(val, int) or val < 0:
                raise ValueError("late_entry_window_hours must be a positive integer")
                
        # 4. OD configuration
        if "od_counts_as_present" in v:
            val = v["od_counts_as_present"]
            if not isinstance(val, bool):
                raise ValueError("od_counts_as_present must be a boolean flag")
                
        return v
    
    # Keeping old method name mapped away just in case
    def validate_grade_scale(cls, v):
        scale = v.get("grade_scale")
        if scale is not None:
            if not isinstance(scale, list) or len(scale) == 0:
                raise ValueError("grade_scale must be a non-empty list of dictionaries")
            
            prev_pct = 101.0
            has_zero = False
            for item in scale:
                pct = item.get("min_pct")
                if pct is None:
                    raise ValueError("Each grade scale item must have min_pct")
                if float(pct) >= prev_pct:
                    raise ValueError(f"grade_scale min_pct must be strictly monotonically decreasing. Violating element: {pct}")
                if float(pct) <= 0:
                    has_zero = True
                prev_pct = float(pct)
                
            if not has_zero:
                raise ValueError("grade_scale must end exactly at or below min_pct=0")
        return v

class ExamScheduleCreate(BaseModel):
    department_id: str
    batch: str
    semester: int
    academic_year: str
    subject_code: str
    subject_name: str
    exam_date: str  # YYYY-MM-DD
    session: str    # FN or AN
    exam_time: str
    document_url: Optional[str] = None

class ExamScheduleUpdate(BaseModel):
    department_id: Optional[str] = None
    batch: Optional[str] = None
    semester: Optional[int] = None
    academic_year: Optional[str] = None
    subject_code: Optional[str] = None
    subject_name: Optional[str] = None
    exam_date: Optional[str] = None
    session: Optional[str] = None
    exam_time: Optional[str] = None
    document_url: Optional[str] = None

class ChallengeSubmit(BaseModel):
    challenge_id: str
    code: str
    language: str = "python"

class ViolationReport(BaseModel):
    violation_type: str = "tab_switch"  # tab_switch, fullscreen_exit, window_blur

class AcademicCalendarCreate(BaseModel):
    academic_year: str
    semester: int = Field(..., ge=1, le=8)
    start_date: str   # "YYYY-MM-DD"
    end_date: str
    working_days: Optional[list] = ["MON", "TUE", "WED", "THU", "FRI"]
    events: Optional[list] = []

class PeriodSlotCreate(BaseModel):
    department_id: str
    batch: str
    section: str
    semester: int = Field(..., ge=1, le=8)
    academic_year: str
    day: str          # MON, TUE, WED, THU, FRI
    period_no: int = Field(..., ge=1, le=10)
    start_time: str   # "09:00"
    end_time: str     # "09:50"
    subject_code: Optional[str] = None
    subject_name: Optional[str] = None
    faculty_id: Optional[str] = None
    slot_type: str = "regular"

class BulkSlotsUpsert(BaseModel):
    slots: List[PeriodSlotCreate]  # up to an entire week of slots at once

class AttendanceMarkItem(BaseModel):
    student_id: str
    status: str = Field(..., pattern="^(present|absent|od|medical|late)$")
    remarks: Optional[str] = None

class AttendanceMarkBatch(BaseModel):
    period_slot_id: str
    date: str          # "YYYY-MM-DD"
    entries: List[AttendanceMarkItem]

class LeaveApply(BaseModel):
    leave_type: str = Field(..., pattern="^(CL|EL|ML|OD|medical)$")
    from_date: str    # "YYYY-MM-DD"
    to_date: str      # "YYYY-MM-DD"
    reason: str = Field(..., max_length=500)
    document_url: Optional[str] = None

class LeaveReview(BaseModel):
    action: str = Field(..., pattern="^(approve|reject)$")
    remarks: Optional[str] = None

class LeaveCancelRequest(BaseModel):
    cancel_from: Optional[str] = None # "YYYY-MM-DD"
    cancel_to: Optional[str] = None   # "YYYY-MM-DD"

class ClassInChargeCreate(BaseModel):
    faculty_ids: List[str]
    department: str
    batch: str
    section: str
    semester: int = Field(..., ge=1, le=8)

class MentorAssignmentCreate(BaseModel):
    faculty_id: str
    student_ids: List[str]

class StudentProgressionCreate(BaseModel):
    student_id: str
    progression_type: str = Field(..., pattern="^(higher_studies|competitive_exam|co_curricular|employment)$")
    details: dict

class SubstituteAssign(BaseModel):
    faculty_id: str

class RegistrationWindowCreate(BaseModel):
    semester: int = Field(..., ge=1, le=8)
    academic_year: str
    open_at: str    # "YYYY-MM-DDTHH:MM"
    close_at: str   # "YYYY-MM-DDTHH:MM"

class CourseRegistrationSchema(BaseModel):
    subject_code: str
    semester: int = Field(..., ge=1, le=8)
    academic_year: str
    is_arrear: bool = False

class InstitutionProfileUpdate(BaseModel):
    recognitions: Optional[dict] = None
    infrastructure: Optional[dict] = None
    library: Optional[dict] = None
    mous: Optional[list] = None
    extension_activities: Optional[dict] = None
    research_publications: Optional[dict] = None

# ─── Phase 6: Teaching Work Schemas ──────────────────────────────────────────

VALID_METHODOLOGIES = {"Lecture", "Demo", "Lab", "Discussion", "Tutorial"}

class TeachingPlanCreate(BaseModel):
    period_slot_id: str
    date: str  # "YYYY-MM-DD"
    planned_topic: str = Field(..., max_length=500)

class ClassRecordCreate(BaseModel):
    period_slot_id: str
    date: str  # "YYYY-MM-DD"
    actual_topic: str = Field(..., max_length=500)
    methodology: str = Field(..., pattern="^(Lecture|Demo|Lab|Discussion|Tutorial)$")
    remarks: Optional[str] = Field(None, max_length=500)

class TeachingRecordUpdate(BaseModel):
    planned_topic: Optional[str] = Field(None, max_length=500)
    actual_topic: Optional[str] = Field(None, max_length=500)
    methodology: Optional[str] = None
    remarks: Optional[str] = Field(None, max_length=500)

class FacultyProfileUpdate(BaseModel):
    """Sectioned profile structure. Each section is a list of records with a status field.
    Structure: { "educational": [{"degree": "...", "status": "draft"}], "experience": [...], ... }
    """
    personal: Optional[dict] = None       # phone, dob, aadhaar, blood_group, gender, address
    educational: Optional[list] = None    # [{degree, university, year, percentage, status}]
    experience: Optional[list] = None     # [{position, institution, from_date, to_date, status}]
    research: Optional[list] = None       # [{title, journal, year, doi, status}]
    publications: Optional[list] = None   # [{title, publisher, year, isbn, status}]
    memberships: Optional[list] = None    # [{body, membership_id, from_date, status}]
    training: Optional[list] = None       # [{program, organizer, dates, certificate_url, status}]


async def log_audit(session: AsyncSession, user_id: str, resource: str, action: str, details: dict = None):
    log_entry = models.AuditLog(
        user_id=user_id,
        resource=resource,
        action=action,
        details=details or {}
    )
    session.add(log_entry)

# ─── Auth Routes ────────────────────────────────────────────────────────────
@app.post("/api/auth/login")
async def login(req: LoginRequest, response: Response, request: Request, session: AsyncSession = Depends(get_db)):
    college_id_normalized = req.college_id.strip().upper()
    key = f"login_failures:{college_id_normalized}"
    if redis_client:
        failures = redis_client.get(key)
        if failures and int(failures) >= 5:
            raise HTTPException(status_code=429, detail="Too many failed attempts. Try again in 5 minutes.")

    # Find user by college_id stored in profile_data or by email
    result = await session.execute(
        select(models.User).where(
            (models.User.profile_data["college_id"].astext == college_id_normalized) |
            (models.User.email.ilike(college_id_normalized))
        )
    )
    user = result.scalars().first()

    def _handle_failure():
        if redis_client:
            redis_client.incr(key)
            redis_client.expire(key, 300)
        raise HTTPException(status_code=401, detail="Invalid credentials")

    if not user:
        _handle_failure()
        
    if not verify_password(req.password, user.password_hash):
        _handle_failure()

    if redis_client:
        redis_client.delete(key)

    perms = {}
    if user.role not in ["student", "super_admin", "admin"]:
        role_result = await session.execute(select(models.Role).where(models.Role.name == user.role, models.Role.college_id == user.college_id))
        r = role_result.scalars().first()
        if r:
            perms = r.permissions

    tid = user.college_id or ""
    access = create_access_token(user.id, user.role, tid, perms)
    refresh = create_refresh_token(user.id)
    response.set_cookie("access_token", access, httponly=True, secure=True, samesite="lax", max_age=86400)
    response.set_cookie("refresh_token", refresh, httponly=True, secure=True, samesite="lax", max_age=604800)

    user_out = {
        "id": user.id,
        "name": user.name,
        "email": user.email,
        "role": user.role,
        "college_id": user.college_id,
        "tenant_id": user.college_id,
        "access_token": access,
    }
    if user.profile_data:
        user_out.update({k: v for k, v in user.profile_data.items() if k != "password_hash"})
    return user_out

@app.post("/api/auth/register")
async def register(req: RegisterRequest, response: Response):
    raise HTTPException(status_code=403, detail="Self-registration is disabled. Please contact your college administrator.")

@app.get("/api/auth/me")
async def get_me(user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    # Fetch UserPermission flags from DB (source of truth for admin-set gates)
    perm_r = await session.execute(
        select(models.UserPermission).where(models.UserPermission.user_id == user["id"])
    )
    perm_row = perm_r.scalars().first()
    permission_flags = perm_row.flags if perm_row else {}

    # Build data scope from FacultyAssignment (for faculty/hod)
    scope = {}
    if user["role"] in ("teacher", "faculty", "hod"):
        assigns_r = await session.execute(
            select(models.FacultyAssignment).where(
                models.FacultyAssignment.teacher_id == user["id"]
            )
        )
        assigns = assigns_r.scalars().all()
        if assigns:
            scope["subject_codes"] = list({a.subject_code for a in assigns})
            scope["batch_ids"] = list({a.batch for a in assigns})
            scope["department"] = assigns[0].department

    return {**user, "permissions": permission_flags, "scope": scope}

@app.post("/api/auth/logout")
async def logout(request: Request, response: Response):
    refresh_token = request.cookies.get("refresh_token")
    if refresh_token:
        try:
            payload = jwt.decode(refresh_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            jti = payload.get("jti")
            if redis_client and jti:
                redis_client.setex(f"revoked_refresh:{jti}", 604800, "revoked")
        except jwt.InvalidTokenError:
            pass
            
    response.delete_cookie("access_token")
    response.delete_cookie("refresh_token")
    return {"message": "Logged out"}

@app.post("/api/auth/refresh")
async def refresh_access_token(request: Request, response: Response):
    from database import AsyncSessionLocal
    refresh_token = request.cookies.get("refresh_token")
    if not refresh_token:
        raise HTTPException(status_code=401, detail="No refresh token")
        
    try:
        payload = jwt.decode(refresh_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
            
        jti = payload.get("jti")
        if redis_client and redis_client.exists(f"revoked_refresh:{jti}"):
            raise HTTPException(status_code=401, detail="Refresh token revoked")
            
        user_id = payload["sub"]
        
        async with AsyncSessionLocal() as session:
            result = await session.execute(select(models.User).where(models.User.id == user_id))
            user = result.scalars().first()
            if not user:
                raise HTTPException(status_code=401, detail="User not found")
                
        new_access = create_access_token(user_id, user.role, user.college_id)
        response.set_cookie("access_token", new_access, httponly=True, secure=True, samesite="lax", max_age=900)
        
        return {"access_token": new_access, "expires_in": 900}
        
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")
# ─── Phase 1: Permission Management ─────────────────────────────────────────

@app.get("/api/admin/users/{user_id}/permissions")
async def get_user_permissions(user_id: str, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    """Get the permission flags for a specific user."""
    perm_r = await session.execute(
        select(models.UserPermission).where(models.UserPermission.user_id == user_id)
    )
    row = perm_r.scalars().first()
    return {"user_id": user_id, "flags": row.flags if row else {}}

@app.put("/api/admin/users/{user_id}/permissions")
async def set_user_permissions(user_id: str, req: PermissionFlagsUpdate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    """Create or fully replace the permission flags for a user."""
    # Verify target user exists and belongs to same college
    target_r = await session.execute(
        select(models.User).where(models.User.id == user_id, models.User.college_id == user["college_id"])
    )
    if not target_r.scalars().first():
        raise HTTPException(status_code=404, detail="User not found")

    perm_r = await session.execute(
        select(models.UserPermission).where(models.UserPermission.user_id == user_id)
    )
    row = perm_r.scalars().first()
    if row:
        row.flags = req.flags
    else:
        row = models.UserPermission(user_id=user_id, college_id=user["college_id"], flags=req.flags)
        session.add(row)
    await log_audit(session, user["id"], "user_permissions", "update", {"target_user": user_id, "flags": req.flags})
    await session.commit()
    return {"message": "Permissions updated", "flags": req.flags}

@app.patch("/api/admin/users/{user_id}/permissions")
async def patch_user_permissions(user_id: str, req: PermissionFlagsUpdate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    """Merge new flags into existing ones (non-destructive toggle)."""
    target_r = await session.execute(
        select(models.User).where(models.User.id == user_id, models.User.college_id == user["college_id"])
    )
    if not target_r.scalars().first():
        raise HTTPException(status_code=404, detail="User not found")

    perm_r = await session.execute(
        select(models.UserPermission).where(models.UserPermission.user_id == user_id)
    )
    row = perm_r.scalars().first()
    if row:
        merged = {**(row.flags or {}), **req.flags}
        row.flags = merged
    else:
        row = models.UserPermission(user_id=user_id, college_id=user["college_id"], flags=req.flags)
        session.add(row)
    await session.commit()
    return {"message": "Permissions merged", "flags": row.flags}

@app.get("/api/admin/permissions/summary")
async def permissions_summary(user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    """List all users with their permission flags for the admin panel."""
    result = await session.execute(
        select(models.User, models.UserPermission)
        .outerjoin(models.UserPermission, models.User.id == models.UserPermission.user_id)
        .where(models.User.college_id == user["college_id"])
    )
    rows = result.all()
    return [
        {"id": u.id, "name": u.name, "role": u.role, "email": u.email,
         "flags": p.flags if p else {}}
        for u, p in rows
    ]

# ─── Phase 1: Subject Allocation (HOD) ───────────────────────────────────────

@app.post("/api/hod/subject-allocation")
async def create_subject_allocation(req: FacultyAssignment, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    """HOD assigns a faculty member to a subject + batch + section."""
    # Verify teacher exists in same college
    teacher_r = await session.execute(
        select(models.User).where(models.User.id == req.teacher_id, models.User.college_id == user["college_id"])
    )
    if not teacher_r.scalars().first():
        raise HTTPException(status_code=404, detail="Teacher not found in this college")

    course_r = await session.execute(
        select(models.Course).where(
            models.Course.subject_code == req.subject_code,
            models.Course.college_id == user["college_id"]
        )
    )
    if not course_r.scalars().first():
        raise HTTPException(status_code=400, detail=f"Subject code '{req.subject_code}' does not exist in the Course master catalog for this college. Please add it first.")

    new_assign = models.FacultyAssignment(
        college_id=user["college_id"],
        teacher_id=req.teacher_id,
        subject_code=req.subject_code,
        subject_name=req.subject_name,
        department=req.department,
        batch=req.batch,
        section=req.section,
        semester=req.semester,
        academic_year=req.academic_year,
        credits=req.credits,
        hours_per_week=req.hours_per_week,
        is_lab=req.is_lab,
    )
    session.add(new_assign)
    await log_audit(session, user["id"], "subject_allocation", "create",
                    {"subject": req.subject_code, "teacher": req.teacher_id})
    await session.commit()
    await session.refresh(new_assign)
    return {"id": new_assign.id, "message": "Allocation created"}

@app.get("/api/hod/subject-allocation")
async def list_subject_allocations(
    academic_year: Optional[str] = None,
    semester: Optional[int] = None,
    user: dict = Depends(require_role("hod", "admin")),
    session: AsyncSession = Depends(get_db)
):
    """HOD views the full subject allocation matrix for their college."""
    stmt = select(models.FacultyAssignment).where(
        models.FacultyAssignment.college_id == user["college_id"]
    )
    if academic_year:
        stmt = stmt.where(models.FacultyAssignment.academic_year == academic_year)
    if semester:
        stmt = stmt.where(models.FacultyAssignment.semester == semester)
    result = await session.execute(stmt)
    assigns = result.scalars().all()
    matrix = {}
    for a in assigns:
        if a.subject_code not in matrix:
            matrix[a.subject_code] = {
                "subject_code": a.subject_code,
                "subject_name": a.subject_name,
                "credits": a.credits,
                "assignments": []
            }
        matrix[a.subject_code]["assignments"].append({
             "id": a.id, "teacher_id": a.teacher_id, "department": a.department, 
             "batch": a.batch, "section": a.section, "semester": a.semester, 
             "academic_year": a.academic_year, "hours_per_week": a.hours_per_week, "is_lab": a.is_lab
        })
    return list(matrix.values())

@app.delete("/api/hod/subject-allocation/{assignment_id}")
async def delete_subject_allocation(assignment_id: str, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    """Remove a subject allocation."""
    result = await session.execute(
        select(models.FacultyAssignment).where(
            models.FacultyAssignment.id == assignment_id,
            models.FacultyAssignment.college_id == user["college_id"]
        )
    )
    row = result.scalars().first()
    if not row:
        raise HTTPException(status_code=404, detail="Allocation not found")
    row.is_deleted = True
    await session.commit()
    return {"message": "Allocation removed"}

@app.get("/api/faculty/my-subjects")
async def get_my_subjects(
    academic_year: Optional[str] = None,
    user: dict = Depends(require_role("teacher", "faculty", "hod")),
    session: AsyncSession = Depends(get_db)
):
    """Faculty sees their own allocated subjects for the semester."""
    stmt = select(models.FacultyAssignment).where(
        models.FacultyAssignment.teacher_id == user["id"]
    )
    if academic_year:
        stmt = stmt.where(models.FacultyAssignment.academic_year == academic_year)
    result = await session.execute(stmt.order_by(models.FacultyAssignment.semester))
    assigns = result.scalars().all()
    return [
        {"id": a.id, "subject_code": a.subject_code, "subject_name": a.subject_name,
         "batch": a.batch, "section": a.section, "semester": a.semester,
         "academic_year": a.academic_year, "credits": a.credits, "is_lab": a.is_lab}
        for a in assigns
    ]

# ─── Phase 1: CIA Template Engine (Admin / Nodal Officer) ────────────────────

@app.post("/api/admin/cia-templates")
async def create_cia_template(req: CIATemplateCreate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    """Create a new CIA assessment template."""
    # Validate total_marks matches sum of component max_marks
    comp_sum = sum(c.get("max_marks", 0) for c in req.components)
    if comp_sum != req.total_marks:
        raise HTTPException(status_code=400, detail=f"Component max_marks sum ({comp_sum}) must equal total_marks ({req.total_marks})")

    tmpl = models.CIATemplate(
        college_id=user["college_id"],
        name=req.name,
        description=req.description,
        total_marks=req.total_marks,
        components=req.components,
    )
    session.add(tmpl)
    await log_audit(session, user["id"], "cia_template", "create", {"name": req.name})
    await session.commit()
    await session.refresh(tmpl)
    return {"id": tmpl.id, "name": tmpl.name, "total_marks": tmpl.total_marks, "message": "Template created"}

@app.get("/api/admin/cia-templates")
async def list_cia_templates(user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.CIATemplate).where(models.CIATemplate.college_id == user["college_id"])
    )
    tmpls = result.scalars().all()
    return [{"id": t.id, "name": t.name, "description": t.description,
             "total_marks": t.total_marks, "components": t.components,
             "created_at": t.created_at.isoformat() if t.created_at else None}
            for t in tmpls]

@app.put("/api/admin/cia-templates/{template_id}")
async def update_cia_template(template_id: str, req: CIATemplateUpdate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.CIATemplate).where(
            models.CIATemplate.id == template_id,
            models.CIATemplate.college_id == user["college_id"]
        )
    )
    tmpl = result.scalars().first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")

    if req.name is not None: tmpl.name = req.name
    if req.description is not None: tmpl.description = req.description
    if req.total_marks is not None: tmpl.total_marks = req.total_marks
    if req.components is not None:
        comp_sum = sum(c.get("max_marks", 0) for c in req.components)
        if comp_sum != (req.total_marks or tmpl.total_marks):
            raise HTTPException(status_code=400, detail=f"Component max_marks sum ({comp_sum}) must equal total_marks")
        tmpl.components = req.components

    await session.commit()
    return {"message": "Template updated"}

@app.delete("/api/admin/cia-templates/{template_id}")
async def delete_cia_template(template_id: str, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.CIATemplate).where(
            models.CIATemplate.id == template_id,
            models.CIATemplate.college_id == user["college_id"]
        )
    )
    tmpl = result.scalars().first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")
    # Check if any SubjectCIAConfig references this template
    ref_r = await session.execute(
        select(models.SubjectCIAConfig).where(models.SubjectCIAConfig.template_id == template_id)
    )
    if ref_r.scalars().first():
        raise HTTPException(status_code=400, detail="Template is in use by a subject config. Remove config first.")
    tmpl.is_deleted = True
    await session.commit()
    return {"message": "Template deleted"}

@app.post("/api/admin/cia-config")
async def create_cia_config(req: SubjectCIAConfigCreate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    """Assign a CIA template to a subject for a given semester + academic year."""
    # Verify template exists
    tmpl_r = await session.execute(
        select(models.CIATemplate).where(
            models.CIATemplate.id == req.template_id,
            models.CIATemplate.college_id == user["college_id"]
        )
    )
    if not tmpl_r.scalars().first():
        raise HTTPException(status_code=404, detail="CIA template not found")

    # Prevent duplicate config for same subject + year + semester
    dup_r = await session.execute(
        select(models.SubjectCIAConfig).where(
            models.SubjectCIAConfig.college_id == user["college_id"],
            models.SubjectCIAConfig.subject_code == req.subject_code,
            models.SubjectCIAConfig.academic_year == req.academic_year,
            models.SubjectCIAConfig.semester == req.semester,
        )
    )
    if dup_r.scalars().first():
        raise HTTPException(status_code=400, detail="CIA config already exists for this subject/year/semester")

    cfg = models.SubjectCIAConfig(
        college_id=user["college_id"],
        subject_code=req.subject_code,
        subject_name=req.subject_name,
        academic_year=req.academic_year,
        semester=req.semester,
        template_id=req.template_id,
    )
    session.add(cfg)
    await session.commit()
    await session.refresh(cfg)
    return {"id": cfg.id, "message": "CIA config created"}

@app.get("/api/admin/cia-config")
async def list_cia_configs(
    academic_year: Optional[str] = None,
    semester: Optional[int] = None,
    user: dict = Depends(require_role("admin")),
    session: AsyncSession = Depends(get_db)
):
    stmt = select(models.SubjectCIAConfig).where(
        models.SubjectCIAConfig.college_id == user["college_id"]
    )
    if academic_year:
        stmt = stmt.where(models.SubjectCIAConfig.academic_year == academic_year)
    if semester:
        stmt = stmt.where(models.SubjectCIAConfig.semester == semester)
    result = await session.execute(stmt)
    cfgs = result.scalars().all()
    return [
        {"id": c.id, "subject_code": c.subject_code, "subject_name": c.subject_name,
         "academic_year": c.academic_year, "semester": c.semester,
         "template_id": c.template_id, "is_consolidation_enabled": c.is_consolidation_enabled}
        for c in cfgs
    ]

@app.put("/api/admin/cia-config/{config_id}/enable-consolidation")
async def toggle_cia_consolidation(
    config_id: str,
    enabled: bool = True,
    user: dict = Depends(require_role("admin")),
    session: AsyncSession = Depends(get_db)
):
    """Nodal Officer gate: enable/disable consolidated mark entry for a subject."""
    result = await session.execute(
        select(models.SubjectCIAConfig).where(
            models.SubjectCIAConfig.id == config_id,
            models.SubjectCIAConfig.college_id == user["college_id"]
        )
    )
    cfg = result.scalars().first()
    if not cfg:
        raise HTTPException(status_code=404, detail="CIA config not found")
    cfg.is_consolidation_enabled = enabled
    await log_audit(session, user["id"], "cia_config", "toggle_consolidation",
                    {"config_id": config_id, "enabled": enabled})
    await session.commit()
    return {"message": f"Consolidation {'enabled' if enabled else 'disabled'}", "subject": cfg.subject_code}

@app.get("/api/subjects/{subject_code}/cia-template")
async def get_subject_cia_template(
    subject_code: str,
    academic_year: str = "2024-25",
    semester: Optional[int] = None,
    user: dict = Depends(get_current_user),
    session: AsyncSession = Depends(get_db)
):
    """Get the active CIA template for a subject (used by faculty and students)."""
    stmt = select(models.SubjectCIAConfig, models.CIATemplate).join(
        models.CIATemplate, models.SubjectCIAConfig.template_id == models.CIATemplate.id
    ).where(
        models.SubjectCIAConfig.college_id == user["college_id"],
        models.SubjectCIAConfig.subject_code == subject_code,
        models.SubjectCIAConfig.academic_year == academic_year,
    )
    if semester:
        stmt = stmt.where(models.SubjectCIAConfig.semester == semester)
    result = await session.execute(stmt)
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="No CIA template configured for this subject")
    cfg, tmpl = row
    return {
        "subject_code": cfg.subject_code, "subject_name": cfg.subject_name,
        "academic_year": cfg.academic_year, "semester": cfg.semester,
        "is_consolidation_enabled": cfg.is_consolidation_enabled,
        "template": {"id": tmpl.id, "name": tmpl.name, "total_marks": tmpl.total_marks, "components": tmpl.components}
    }

# ─── Phase 6: Faculty CIA Dashboard ───────────────────────────────────────────

@app.get("/api/faculty/cia-dashboard")
async def get_faculty_cia_dashboard(
    user: dict = Depends(require_role("teacher", "faculty", "hod")),
    session: AsyncSession = Depends(get_db)
):
    """Get all assigned subjects with their CIA template components and per-component mark entry status.
    Returns a list of subjects, each with template components and mark entry progress."""
    # 1. Get faculty's subject assignments
    assigns_r = await session.execute(
        select(models.FacultyAssignment).where(
            models.FacultyAssignment.teacher_id == user["id"]
        )
    )
    assigns = assigns_r.scalars().all()
    if not assigns:
        return []

    # 2. Get CIA configs for those subjects
    subject_codes = list(set(a.subject_code for a in assigns))
    configs_r = await session.execute(
        select(models.SubjectCIAConfig, models.CIATemplate).join(
            models.CIATemplate, models.SubjectCIAConfig.template_id == models.CIATemplate.id
        ).where(
            models.SubjectCIAConfig.college_id == user["college_id"],
            models.SubjectCIAConfig.subject_code.in_(subject_codes)
        )
    )
    config_rows = configs_r.all()
    # Map subject_code -> (config, template)
    config_map = {}
    for cfg, tmpl in config_rows:
        config_map[cfg.subject_code] = (cfg, tmpl)

    # 3. Get all mark entries by this faculty for these subjects
    marks_r = await session.execute(
        select(models.MarkEntry).where(
            models.MarkEntry.faculty_id == user["id"],
            models.MarkEntry.course_id.in_(subject_codes)
        )
    )
    mark_entries = marks_r.scalars().all()
    # Map (subject_code, exam_type) -> mark_entry
    marks_map = {}
    for me in mark_entries:
        marks_map[(me.course_id, me.exam_type)] = me

    result = []
    for assign in assigns:
        subject_data = {
            "assignment_id": assign.id,
            "subject_code": assign.subject_code,
            "subject_name": assign.subject_name,
            "department": assign.department,
            "batch": assign.batch,
            "section": assign.section,
            "semester": assign.semester,
            "has_cia_template": assign.subject_code in config_map,
        }

        if assign.subject_code in config_map:
            cfg, tmpl = config_map[assign.subject_code]
            components_with_status = []
            for comp in (tmpl.components or []):
                comp_type = comp.get("type", "unknown")
                comp_name = comp.get("name", comp_type)
                # Check if there's a mark entry for this component type
                entry = marks_map.get((assign.subject_code, comp_type))
                entry_status = "not_started"
                entry_id = None
                student_count = 0
                if entry:
                    entry_status = (entry.extra_data or {}).get("status", "draft")
                    entry_id = entry.id
                    student_count = len((entry.extra_data or {}).get("entries", []))

                components_with_status.append({
                    "type": comp_type,
                    "name": comp_name,
                    "max_marks": comp.get("max_marks", 0),
                    "count": comp.get("count"),
                    "best_of": comp.get("best_of"),
                    "slabs": comp.get("slabs"),
                    "entry_status": entry_status,
                    "entry_id": entry_id,
                    "student_count": student_count,
                })

            subject_data["template"] = {
                "id": tmpl.id,
                "name": tmpl.name,
                "total_marks": tmpl.total_marks,
            }
            subject_data["components"] = components_with_status
            subject_data["is_consolidation_enabled"] = cfg.is_consolidation_enabled
        else:
            subject_data["template"] = None
            subject_data["components"] = []
            subject_data["is_consolidation_enabled"] = False

        result.append(subject_data)

    return result

# ─── Phase 2: Academic Calendar (Admin/Nodal Officer) ─────────────────────────

@app.post("/api/admin/academic-calendars")
async def create_academic_calendar(req: AcademicCalendarCreate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    try:
        start_dt = datetime.strptime(req.start_date, "%Y-%m-%d").date()
        end_dt = datetime.strptime(req.end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Dates must be in YYYY-MM-DD format")

    if end_dt <= start_dt:
        raise HTTPException(status_code=400, detail="end_date must be after start_date")

    # Prevent overlap for same semester
    dup_r = await session.execute(
        select(models.AcademicCalendar).where(
            models.AcademicCalendar.college_id == user["college_id"],
            models.AcademicCalendar.academic_year == req.academic_year,
            models.AcademicCalendar.semester == req.semester
        )
    )
    if dup_r.scalars().first():
        raise HTTPException(status_code=400, detail="Calendar already exists for this year and semester")

    cal = models.AcademicCalendar(
        college_id=user["college_id"],
        academic_year=req.academic_year,
        semester=req.semester,
        start_date=start_dt,
        end_date=end_dt,
        working_days=req.working_days,
        events=req.events
    )
    session.add(cal)
    await session.commit()
    await session.refresh(cal)
    return {"id": cal.id, "message": "Academic calendar created"}

@app.get("/api/academic-calendars")
async def list_academic_calendars(academic_year: Optional[str] = None, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    stmt = select(models.AcademicCalendar).where(models.AcademicCalendar.college_id == user["college_id"])
    if academic_year:
        stmt = stmt.where(models.AcademicCalendar.academic_year == academic_year)
    result = await session.execute(stmt)
    cals = result.scalars().all()
    return [{
        "id": c.id, "academic_year": c.academic_year, "semester": c.semester,
        "start_date": c.start_date.isoformat(), "end_date": c.end_date.isoformat(),
        "working_days": c.working_days, "events": c.events
    } for c in cals]

# ─── Phase 2: Timetable / Period Slots (HOD, Faculty, Student) ───────────────

@app.put("/api/hod/timetable/slots")
async def upsert_timetable_slots(req: BulkSlotsUpsert, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    """Bulk upsert period slots (usually representing a weekly template)."""
    if not req.slots:
        return {"message": "No slots provided"}

    updated_count = 0
    created_count = 0

    coords = set((s.department_id, s.batch, s.section, s.academic_year) for s in req.slots)
    conditions = [
        and_(
            models.PeriodSlot.department_id == c[0],
            models.PeriodSlot.batch == c[1],
            models.PeriodSlot.section == c[2],
            models.PeriodSlot.academic_year == c[3]
        ) for c in coords
    ]
    
    existing_map = {}
    if conditions:
        exist_r = await session.execute(
            select(models.PeriodSlot).where(
                models.PeriodSlot.college_id == user["college_id"],
                or_(*conditions)
            )
        )
        for s in exist_r.scalars().all():
            key = (s.department_id, s.batch, s.section, s.academic_year, s.day, s.period_no)
            existing_map[key] = s

    for slot_data in req.slots:
        key = (slot_data.department_id, slot_data.batch, slot_data.section, slot_data.academic_year, slot_data.day, slot_data.period_no)
        existing = existing_map.get(key)

        if existing:
            existing.semester = slot_data.semester
            existing.start_time = slot_data.start_time
            existing.end_time = slot_data.end_time
            existing.subject_code = slot_data.subject_code
            existing.subject_name = slot_data.subject_name
            existing.faculty_id = slot_data.faculty_id
            existing.slot_type = slot_data.slot_type
            updated_count += 1
        else:
            new_slot = models.PeriodSlot(
                college_id=user["college_id"],
                department_id=slot_data.department_id,
                batch=slot_data.batch,
                section=slot_data.section,
                semester=slot_data.semester,
                academic_year=slot_data.academic_year,
                day=slot_data.day,
                period_no=slot_data.period_no,
                start_time=slot_data.start_time,
                end_time=slot_data.end_time,
                subject_code=slot_data.subject_code,
                subject_name=slot_data.subject_name,
                faculty_id=slot_data.faculty_id,
                slot_type=slot_data.slot_type
            )
            session.add(new_slot)
            created_count += 1

    await session.commit()
    return {"message": f"Slots saved: {created_count} created, {updated_count} updated."}

@app.get("/api/hod/timetable")
async def get_department_timetable(
    department_id: str,
    batch: str,
    section: str,
    academic_year: str,
    user: dict = Depends(require_role("hod", "admin")),
    session: AsyncSession = Depends(get_db)
):
    """Get the weekly timetable grid for a specific batch/section."""
    result = await session.execute(
        select(models.PeriodSlot).where(
            models.PeriodSlot.college_id == user["college_id"],
            models.PeriodSlot.department_id == department_id,
            models.PeriodSlot.batch == batch,
            models.PeriodSlot.section == section,
            models.PeriodSlot.academic_year == academic_year
        )
    )
    slots = result.scalars().all()
    return [{
        "id": s.id, "day": s.day, "period_no": s.period_no, "start_time": s.start_time, "end_time": s.end_time,
        "subject_code": s.subject_code, "subject_name": s.subject_name,
        "faculty_id": s.faculty_id, "slot_type": s.slot_type
    } for s in slots]

async def get_current_academic_year(session: AsyncSession, college_id: str) -> str:
    today = datetime.now(timezone.utc).date()
    result = await session.execute(
        select(models.AcademicCalendar).where(
            models.AcademicCalendar.college_id == college_id,
            models.AcademicCalendar.start_date <= today,
            models.AcademicCalendar.end_date >= today
        )
    )
    calendar = result.scalars().first()
    if not calendar:
        raise HTTPException(status_code=400, detail="No active Academic Calendar configured for today.")
    return calendar.academic_year

@app.get("/api/faculty/timetable/today")
async def get_faculty_today_timetable(
    week: bool = False,
    user: dict = Depends(require_role("teacher", "faculty", "hod")),
    session: AsyncSession = Depends(get_db)
):
    """Get today's periods (default) or full weekly grid (?week=true) for the logged-in faculty."""
    today = datetime.now(timezone.utc)
    days = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    current_day = days[today.weekday()]

    current_academic_year = await get_current_academic_year(session, user.get("college_id", ""))

    stmt = select(models.PeriodSlot).where(
        models.PeriodSlot.faculty_id == user["id"],
        models.PeriodSlot.academic_year == current_academic_year
    )
    if not week:
        stmt = stmt.where(models.PeriodSlot.day == current_day)
    stmt = stmt.order_by(models.PeriodSlot.day, models.PeriodSlot.period_no)

    result = await session.execute(stmt)
    slots = result.scalars().all()
    return [{
        "id": s.id, "day": s.day, "period_no": s.period_no, "start_time": s.start_time, "end_time": s.end_time,
        "batch": s.batch, "section": s.section, "department_id": s.department_id,
        "subject_code": s.subject_code, "subject_name": s.subject_name, "slot_type": s.slot_type
    } for s in slots]

# ─── Phase 6: Daily Teaching Work ─────────────────────────────────────────────

@app.get("/api/faculty/teaching-records")
async def get_teaching_records(
    month: Optional[int] = None,
    year: Optional[int] = None,
    user: dict = Depends(require_role("teacher", "faculty", "hod")),
    session: AsyncSession = Depends(get_db)
):
    """Get teaching records for the faculty, optionally filtered by month/year."""
    params = {"faculty_id": user["id"]}
    where_clauses = ["tr.faculty_id = :faculty_id", "tr.is_deleted = false"]
    
    if month and year:
        where_clauses.append("EXTRACT(MONTH FROM tr.date) = :month")
        where_clauses.append("EXTRACT(YEAR FROM tr.date) = :year")
        params["month"] = month
        params["year"] = year

    where_sql = " AND ".join(where_clauses)
    stmt = text(f"""
        SELECT 
            tr.id, tr.date, tr.planned_topic, tr.actual_topic, 
            tr.methodology, tr.remarks, tr.is_class_record_submitted,
            tr.period_slot_id,
            ps.period_no, ps.start_time, ps.end_time, ps.day,
            ps.subject_code, ps.subject_name, ps.batch, ps.section
        FROM teaching_records tr
        JOIN period_slots ps ON tr.period_slot_id = ps.id
        WHERE {where_sql}
        ORDER BY tr.date DESC, ps.period_no ASC
    """)
    result = await session.execute(stmt, params)
    rows = result.all()
    return [{
        "id": r.id, "date": str(r.date),
        "planned_topic": r.planned_topic, "actual_topic": r.actual_topic,
        "methodology": r.methodology, "remarks": r.remarks,
        "is_class_record_submitted": r.is_class_record_submitted,
        "period_slot_id": r.period_slot_id,
        "period_no": r.period_no, "start_time": r.start_time, "end_time": r.end_time,
        "day": r.day, "subject_code": r.subject_code, "subject_name": r.subject_name,
        "batch": r.batch, "section": r.section
    } for r in rows]

@app.post("/api/faculty/teaching-plan")
async def save_teaching_plan(
    req: TeachingPlanCreate,
    user: dict = Depends(require_role("teacher", "faculty", "hod")),
    session: AsyncSession = Depends(get_db)
):
    """Save a teaching plan (planned topic). Enforces T+14 day window server-side."""
    try:
        target_date = datetime.strptime(req.date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")

    today = datetime.now(timezone.utc).date()
    if (target_date - today).days > 14:
        raise HTTPException(status_code=400, detail=f"Teaching plan can only be set up to 14 days in advance. Target date {req.date} is {(target_date - today).days} days away.")

    # Verify the period slot belongs to this faculty
    slot = await session.get(models.PeriodSlot, req.period_slot_id)
    if not slot or slot.faculty_id != user["id"]:
        raise HTTPException(status_code=404, detail="Period slot not found or does not belong to you.")

    # Check for existing record (upsert logic)
    existing = await session.execute(
        select(models.TeachingRecord).where(
            models.TeachingRecord.faculty_id == user["id"],
            models.TeachingRecord.period_slot_id == req.period_slot_id,
            models.TeachingRecord.date == target_date,
            models.TeachingRecord.is_deleted == False
        )
    )
    record = existing.scalars().first()
    if record:
        record.planned_topic = req.planned_topic
    else:
        record = models.TeachingRecord(
            college_id=user["college_id"],
            faculty_id=user["id"],
            period_slot_id=req.period_slot_id,
            date=target_date,
            planned_topic=req.planned_topic
        )
        session.add(record)
    
    await session.commit()
    return {"id": record.id, "message": "Teaching plan saved"}

@app.post("/api/faculty/class-record")
async def save_class_record(
    req: ClassRecordCreate,
    user: dict = Depends(require_role("teacher", "faculty", "hod")),
    session: AsyncSession = Depends(get_db)
):
    """Save a class record (actual topic + methodology). Enforces T to T-3 day window server-side."""
    try:
        target_date = datetime.strptime(req.date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")

    today = datetime.now(timezone.utc).date()
    if target_date > today:
        raise HTTPException(status_code=400, detail="Cannot submit class record for a future date.")
    if (today - target_date).days > 3:
        raise HTTPException(status_code=400, detail=f"Class record submission window closed. Date {req.date} is {(today - target_date).days} days ago (limit is 3 days).")

    # Verify slot ownership
    slot = await session.get(models.PeriodSlot, req.period_slot_id)
    if not slot or slot.faculty_id != user["id"]:
        raise HTTPException(status_code=404, detail="Period slot not found or does not belong to you.")

    # Upsert
    existing = await session.execute(
        select(models.TeachingRecord).where(
            models.TeachingRecord.faculty_id == user["id"],
            models.TeachingRecord.period_slot_id == req.period_slot_id,
            models.TeachingRecord.date == target_date,
            models.TeachingRecord.is_deleted == False
        )
    )
    record = existing.scalars().first()
    if record:
        record.actual_topic = req.actual_topic
        record.methodology = req.methodology
        record.remarks = req.remarks
        record.is_class_record_submitted = True
    else:
        record = models.TeachingRecord(
            college_id=user["college_id"],
            faculty_id=user["id"],
            period_slot_id=req.period_slot_id,
            date=target_date,
            actual_topic=req.actual_topic,
            methodology=req.methodology,
            remarks=req.remarks,
            is_class_record_submitted=True
        )
        session.add(record)

    await session.commit()
    return {"id": record.id, "message": "Class record submitted"}

@app.patch("/api/faculty/teaching-records/{record_id}")
async def update_teaching_record(
    record_id: str,
    req: TeachingRecordUpdate,
    user: dict = Depends(require_role("teacher", "faculty", "hod")),
    session: AsyncSession = Depends(get_db)
):
    """Edit an existing teaching record. Enforces same time windows as create."""
    record = await session.get(models.TeachingRecord, record_id)
    if not record or record.faculty_id != user["id"] or record.is_deleted:
        raise HTTPException(status_code=404, detail="Teaching record not found.")

    today = datetime.now(timezone.utc).date()

    # If updating planned_topic, enforce T+14 window
    if req.planned_topic is not None:
        if record.date < today:
            raise HTTPException(status_code=400, detail="Cannot edit teaching plan for a past date.")
        if (record.date - today).days > 14:
            raise HTTPException(status_code=400, detail="Teaching plan can only be edited up to 14 days in advance.")
        record.planned_topic = req.planned_topic

    # If updating actual_topic/methodology, enforce T to T-3 window
    if req.actual_topic is not None or req.methodology is not None:
        if record.date > today:
            raise HTTPException(status_code=400, detail="Cannot submit class record for a future date.")
        if (today - record.date).days > 3:
            raise HTTPException(status_code=400, detail=f"Class record edit window closed ({(today - record.date).days} days ago, limit is 3).")
        if req.actual_topic is not None:
            record.actual_topic = req.actual_topic
        if req.methodology is not None:
            if req.methodology not in VALID_METHODOLOGIES:
                raise HTTPException(status_code=400, detail=f"Invalid methodology. Must be one of: {', '.join(VALID_METHODOLOGIES)}")
            record.methodology = req.methodology
        record.is_class_record_submitted = True

    if req.remarks is not None:
        record.remarks = req.remarks

    await session.commit()
    return {"id": record.id, "message": "Teaching record updated"}

# ─── Phase 6: Faculty Profile ────────────────────────────────────────────────

@app.get("/api/faculty/profile")
async def get_faculty_profile(
    user: dict = Depends(require_role("teacher", "faculty", "hod")),
    session: AsyncSession = Depends(get_db)
):
    """Get the full faculty profile with sectioned DHTE fields."""
    u = await session.get(models.User, user["id"])
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    profile = u.profile_data or {}
    return {
        "id": u.id, "name": u.name, "email": u.email, "role": u.role,
        "college_id": u.college_id, "department": u.department,
        "designation": profile.get("designation", ""),
        "date_of_joining": profile.get("date_of_joining", ""),
        "personal": profile.get("personal", {}),
        "educational": profile.get("educational", []),
        "experience": profile.get("experience", []),
        "research": profile.get("research", []),
        "publications": profile.get("publications", []),
        "memberships": profile.get("memberships", []),
        "training": profile.get("training", []),
    }

@app.put("/api/faculty/profile")
async def update_faculty_profile(
    req: FacultyProfileUpdate,
    user: dict = Depends(require_role("teacher", "faculty", "hod")),
    session: AsyncSession = Depends(get_db)
):
    """Update faculty profile sections. Each record in a list section gets a status field."""
    u = await session.get(models.User, user["id"])
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    
    profile = dict(u.profile_data or {})
    updates = req.dict(exclude_unset=True)
    
    # For list sections, ensure every record has a status field
    list_sections = ["educational", "experience", "research", "publications", "memberships", "training"]
    for section_name in list_sections:
        if section_name in updates and updates[section_name] is not None:
            for record in updates[section_name]:
                if isinstance(record, dict) and "status" not in record:
                    record["status"] = "draft"
    
    profile.update(updates)
    u.profile_data = profile
    # Flag the column as modified for SQLAlchemy to detect JSONB changes
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(u, "profile_data")
    
    await session.commit()
    return {"message": "Profile updated successfully"}

@app.get("/api/student/timetable")
async def get_student_timetable(user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    """Get the weekly timetable for the logged-in student."""
    # Since courses aren't fully seeded, we use profile_data for batch/section
    department = user.get("department")
    batch = user.get("batch")
    section = user.get("section")
    
    if not all([department, batch, section]):
        return []

    # Requires looking up the department_id by code/name
    dept_r = await session.execute(
        select(models.Department).where(
            models.Department.college_id == user["college_id"],
            (models.Department.code == department) | (models.Department.name == department)
        )
    )
    dept = dept_r.scalars().first()
    if not dept:
        return []

    current_academic_year = await get_current_academic_year(session, user.get("college_id", ""))

    result = await session.execute(
        select(models.PeriodSlot).where(
            models.PeriodSlot.department_id == dept.id,
            models.PeriodSlot.batch == batch,
            models.PeriodSlot.section == section,
            models.PeriodSlot.academic_year == current_academic_year
        )
    )
    slots = result.scalars().all()
    
    # Needs faculty names. Could join, but fetching in Python is okay for small sets.
    faculty_ids = list(set([s.faculty_id for s in slots if s.faculty_id]))
    faculty_map = {}
    if faculty_ids:
        fac_r = await session.execute(select(models.User.id, models.User.name).where(models.User.id.in_(faculty_ids)))
        faculty_map = {f.id: f.name for f in fac_r.all()}

    return [{
        "id": s.id, "day": s.day, "period_no": s.period_no, "start_time": s.start_time, "end_time": s.end_time,
        "subject_code": s.subject_code, "subject_name": s.subject_name,
        "faculty_id": s.faculty_id, "faculty_name": faculty_map.get(s.faculty_id, "Unknown"), "slot_type": s.slot_type
    } for s in slots]

# ─── Phase 2: Attendance System ──────────────────────────────────────────────

@app.get("/api/faculty/attendance/today")
async def get_today_attendance_status(user: dict = Depends(require_role("teacher", "faculty", "hod")), session: AsyncSession = Depends(get_db)):
    """Get today's slots for faculty and indicate if attendance is marked."""
    today = datetime.now(timezone.utc)
    days = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    current_day = days[today.weekday()]
    current_date = today.date()
    try:
        current_academic_year = await get_current_academic_year(session, user.get("college_id", ""))
    except HTTPException:
        return []

    slots_r = await session.execute(
        select(models.PeriodSlot).where(
            models.PeriodSlot.faculty_id == user["id"],
            models.PeriodSlot.day == current_day,
            models.PeriodSlot.academic_year == current_academic_year
        ).order_by(models.PeriodSlot.period_no)
    )
    slots = slots_r.scalars().all()

    if not slots:
        return []

    slot_ids = [s.id for s in slots]
    
    # Find which slots have attendance records for today
    att_r = await session.execute(
        select(models.AttendanceRecord.period_slot_id, func.count(models.AttendanceRecord.id))
        .where(
            models.AttendanceRecord.period_slot_id.in_(slot_ids),
            models.AttendanceRecord.date == current_date
        )
        .group_by(models.AttendanceRecord.period_slot_id)
    )
    marked_counts = {row.period_slot_id: row.count for row in att_r.all()}

    return [{
        "slot": {
            "id": s.id, "period_no": s.period_no, "start_time": s.start_time, "end_time": s.end_time,
            "batch": s.batch, "section": s.section, "subject_code": s.subject_code, "subject_name": s.subject_name
        },
        "is_marked": s.id in marked_counts,
        "recorded_count": marked_counts.get(s.id, 0)
    } for s in slots]

@app.post("/api/faculty/attendance/mark")
async def mark_attendance_batch(req: AttendanceMarkBatch, user: dict = Depends(require_role("teacher", "faculty", "hod")), session: AsyncSession = Depends(get_db)):
    """Mark attendance for an entire class for a specific period slot."""
    try:
        mark_date = datetime.strptime(req.date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    slot_r = await session.execute(
        select(models.PeriodSlot).where(
            models.PeriodSlot.id == req.period_slot_id,
            models.PeriodSlot.college_id == user["college_id"]
        )
    )
    slot = slot_r.scalars().first()
    if not slot:
        raise HTTPException(status_code=404, detail="Period slot not found")

    if slot.faculty_id != user["id"]:
        raise HTTPException(status_code=403, detail="You are not assigned to this period slot")

    # 3-hour window enforcement
    now = datetime.now()
    try:
        period_end_time = datetime.strptime(slot.end_time, "%H:%M").time()
        # Combine mark_date and period_end_time to get a full datetime
        period_end_dt = datetime.combine(mark_date, period_end_time)
        delta_hours = (now - period_end_dt).total_seconds() / 3600
        is_late_entry = delta_hours > 3
    except Exception:
        is_late_entry = False

    # Check for existing records to prevent duplicates
    # Soft delete existing records for this slot+date
    await session.execute(
        update(models.AttendanceRecord).where(
            models.AttendanceRecord.period_slot_id == slot.id,
            models.AttendanceRecord.date == mark_date
        ).values(is_deleted=True, deleted_at=func.now())
    )

    records = [
        models.AttendanceRecord(
            college_id=user["college_id"],
            period_slot_id=slot.id,
            date=mark_date,
            faculty_id=user["id"],
            student_id=entry.student_id,
            subject_code=slot.subject_code,
            status=entry.status,
            is_late_entry=is_late_entry,
            remarks=entry.remarks
        )
        for entry in req.entries
    ]
    session.add_all(records)
    await log_audit(session, user["id"], "attendance", "mark_batch", 
                    {"slot_id": slot.id, "date": req.date, "is_late": is_late_entry, "count": len(records)})
    await session.commit()
    
    return {"message": f"Successfully marked attendance for {len(records)} students", "is_late_entry": is_late_entry}

@app.get("/api/student/attendance")
async def get_student_consolidated_attendance(user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    """Consolidated attendance across all subjects for a student."""
    # Using raw SQL equivalent for calculating present percentages per subject
    stmt = text("""
        SELECT 
            subject_code,
            COUNT(*) FILTER (WHERE status = 'present' OR status = 'od') AS present_count,
            COUNT(*) AS total_count
        FROM attendance_records
        WHERE student_id = :student_id AND is_deleted = false
        GROUP BY subject_code
    """)
    result = await session.execute(stmt, {"student_id": user["id"]})
    rows = result.all()
    
    response = []
    for row in rows:
        pct = round(row.present_count * 100.0 / row.total_count, 1) if row.total_count > 0 else 0
        response.append({
            "subject_code": row.subject_code,
            "present_count": row.present_count,
            "total_count": row.total_count,
            "percentage": pct
        })
    return response

@app.get("/api/student/attendance/detail")
async def get_student_attendance_detail(
    subject_code: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    user: dict = Depends(require_role("student")),
    session: AsyncSession = Depends(get_db)
):
    """Per-date, per-period attendance records for the student calendar view."""
    params = {"student_id": user["id"]}
    where_clauses = ["ar.student_id = :student_id", "ar.is_deleted = false"]
    
    if subject_code:
        where_clauses.append("ar.subject_code = :subject_code")
        params["subject_code"] = subject_code
    if month and year:
        where_clauses.append("EXTRACT(MONTH FROM ar.date) = :month")
        where_clauses.append("EXTRACT(YEAR FROM ar.date) = :year")
        params["month"] = month
        params["year"] = year

    where_sql = " AND ".join(where_clauses)
    stmt = text(f"""
        SELECT 
            ar.date, ar.subject_code, ar.status, ar.remarks,
            ps.period_no, ps.start_time, ps.end_time, ps.subject_name
        FROM attendance_records ar
        JOIN period_slots ps ON ar.period_slot_id = ps.id
        WHERE {where_sql}
        ORDER BY ar.date DESC, ps.period_no ASC
    """)
    result = await session.execute(stmt, params)
    rows = result.all()

    return [{
        "date": str(r.date),
        "subject_code": r.subject_code,
        "subject_name": r.subject_name,
        "period_no": r.period_no,
        "start_time": r.start_time,
        "end_time": r.end_time,
        "status": r.status,
        "remarks": r.remarks
    } for r in rows]

@app.get("/api/student/cia-marks")
async def get_student_cia_marks(
    semester: Optional[int] = None,
    academic_year: Optional[str] = None,
    user: dict = Depends(require_role("student")),
    session: AsyncSession = Depends(get_db)
):
    """Component-wise CIA marks breakdown for the student."""
    # Get mark entries for this student
    stmt = select(models.MarkEntry).where(
        models.MarkEntry.student_id == user["id"],
        models.MarkEntry.is_deleted == False
    )
    result = await session.execute(stmt)
    entries = result.scalars().all()

    # Get CIA configs for context
    cfg_stmt = select(models.SubjectCIAConfig).where(
        models.SubjectCIAConfig.college_id == user["college_id"],
        models.SubjectCIAConfig.is_deleted == False
    )
    if semester:
        cfg_stmt = cfg_stmt.where(models.SubjectCIAConfig.semester == semester)
    if academic_year:
        cfg_stmt = cfg_stmt.where(models.SubjectCIAConfig.academic_year == academic_year)
    cfg_result = await session.execute(cfg_stmt)
    configs = cfg_result.scalars().all()

    # Get templates for component details
    template_ids = list(set(c.template_id for c in configs))
    tpl_map = {}
    if template_ids:
        tpl_r = await session.execute(
            select(models.CIATemplate).where(models.CIATemplate.id.in_(template_ids))
        )
        tpl_map = {t.id: t for t in tpl_r.scalars().all()}

    # Build subject-wise response
    subject_map = {}
    for cfg in configs:
        tpl = tpl_map.get(cfg.template_id)
        components = (tpl.components if tpl else []) or []
        
        # Find student marks for this subject
        subj_entries = [e for e in entries if e.course_id == cfg.subject_code]
        component_marks = []
        total_obtained = 0
        total_max = 0
        
        for comp in components:
            entry = next((e for e in subj_entries if e.exam_type == comp.get("type", "")), None)
            obtained = entry.marks_obtained if entry else None
            max_m = comp.get("max_marks", 0)
            component_marks.append({
                "name": comp.get("name", comp.get("type", "")),
                "type": comp.get("type", ""),
                "max_marks": max_m,
                "marks_obtained": obtained
            })
            if obtained is not None:
                total_obtained += obtained
            total_max += max_m
        
        subject_map[cfg.subject_code] = {
            "subject_code": cfg.subject_code,
            "subject_name": cfg.subject_name or cfg.subject_code,
            "semester": cfg.semester,
            "academic_year": cfg.academic_year,
            "components": component_marks,
            "total_cia": round(total_obtained, 1),
            "total_max": total_max
        }

    return list(subject_map.values())

@app.get("/api/student/academic-calendar")
async def get_student_academic_calendar(
    user: dict = Depends(require_role("student")),
    session: AsyncSession = Depends(get_db)
):
    """Get the academic calendar with events for the student's college."""
    result = await session.execute(
        select(models.AcademicCalendar).where(
            models.AcademicCalendar.college_id == user["college_id"],
            models.AcademicCalendar.is_deleted == False
        ).order_by(models.AcademicCalendar.start_date.desc())
    )
    calendars = result.scalars().all()
    return [{
        "id": c.id,
        "academic_year": c.academic_year,
        "semester": c.semester,
        "start_date": str(c.start_date),
        "end_date": str(c.end_date),
        "working_days": c.working_days or [],
        "events": c.events or []
    } for c in calendars]

@app.get("/api/student/attendance/calendar")
async def get_student_attendance_calendar(
    month: Optional[int] = None,
    year: Optional[int] = None,
    user: dict = Depends(require_role("student")),
    session: AsyncSession = Depends(get_db)
):
    """Aggregated attendance calendar View. Groups records by date."""
    from sqlalchemy import extract
    stmt = select(models.AttendanceRecord).where(
        models.AttendanceRecord.student_id == user["id"],
        models.AttendanceRecord.is_deleted == False
    )
    if month and year:
        stmt = stmt.where(extract("month", models.AttendanceRecord.date) == month)
        stmt = stmt.where(extract("year", models.AttendanceRecord.date) == year)
        
    result = await session.execute(stmt)
    records = result.scalars().all()
    
    calendar = {}
    for r in records:
        d = str(r.date)
        if d not in calendar:
            calendar[d] = {"present": 0, "absent": 0, "od": 0, "details": []}
        
        status = r.status.lower() if r.status else ""
        if status in calendar[d]:
            calendar[d][status] += 1
            
        calendar[d]["details"].append({
            "subject_code": r.subject_code,
            "period": r.period_no,
            "status": status,
            "is_late": r.is_late_entry
        })
    return calendar

@app.get("/api/student/my-mentor")
async def get_my_mentor(user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    """Resolves the student's active mentor from MentorAssignment."""
    stmt = select(models.MentorAssignment).where(
        models.MentorAssignment.student_id == user["id"],
        models.MentorAssignment.is_active == True,
        models.MentorAssignment.college_id == user["college_id"]
    )
    result = await session.execute(stmt)
    mentor_assignment = result.scalars().first()
    
    if not mentor_assignment:
        return {"mentor": None}
        
    faculty = await session.get(models.User, mentor_assignment.faculty_id)
    if not faculty:
        return {"mentor": None}
        
    return {
        "mentor": {
            "id": faculty.id,
            "name": faculty.name,
            "email": faculty.email,
            "department": faculty.department
        },
        "assigned_at": str(mentor_assignment.created_at)
    }


@app.get("/api/student/subjects")
async def get_student_subjects(
    user: dict = Depends(require_role("student")),
    session: AsyncSession = Depends(get_db)
):
    """Get registered subjects with faculty details for the student."""
    # Get the student's course registrations
    regs_r = await session.execute(
        select(models.CourseRegistration).where(
            models.CourseRegistration.student_id == user["id"],
            models.CourseRegistration.is_deleted == False
        ).order_by(models.CourseRegistration.semester)
    )
    regs = regs_r.scalars().all()

    # Find faculty assignments for matched subjects
    subject_codes = list(set(r.subject_code for r in regs))
    fac_map = {}
    if subject_codes:
        fac_r = await session.execute(
            select(models.FacultyAssignment, models.User).join(
                models.User, models.FacultyAssignment.teacher_id == models.User.id
            ).where(
                models.FacultyAssignment.college_id == user["college_id"],
                models.FacultyAssignment.subject_code.in_(subject_codes),
                models.FacultyAssignment.is_deleted == False
            )
        )
        for fa, u in fac_r.all():
            fac_map[fa.subject_code] = {
                "faculty_name": u.name,
                "credits": fa.credits,
                "hours_per_week": fa.hours_per_week,
                "is_lab": fa.is_lab,
                "subject_name": fa.subject_name
            }

    return [{
        "subject_code": r.subject_code,
        "subject_name": fac_map.get(r.subject_code, {}).get("subject_name", r.subject_code),
        "semester": r.semester,
        "academic_year": r.academic_year,
        "status": r.status,
        "is_arrear": r.is_arrear,
        "faculty_name": fac_map.get(r.subject_code, {}).get("faculty_name", "—"),
        "credits": fac_map.get(r.subject_code, {}).get("credits"),
        "hours_per_week": fac_map.get(r.subject_code, {}).get("hours_per_week"),
        "is_lab": fac_map.get(r.subject_code, {}).get("is_lab", False)
    } for r in regs]

@app.get("/api/hod/attendance/defaulters")
async def get_attendance_defaulters(
    threshold: float = 75.0,
    academic_year: Optional[str] = None,
    user: dict = Depends(require_role("hod", "admin")),
    session: AsyncSession = Depends(get_db)
):
    """Get list of students below the attendance threshold."""
    stmt = text("""
        SELECT 
            student_id,
            u.name as student_name,
            u.profile_data->>'batch' as batch,
            subject_code,
            ROUND((COUNT(*) FILTER (WHERE status = 'present' OR status = 'od') * 100.0 / COUNT(*))::numeric, 1) as percentage
        FROM attendance_records ar
        JOIN users u ON ar.student_id = u.id
        WHERE ar.college_id = :college_id 
          AND ar.is_deleted = false 
          AND u.profile_data->>'department_id' = :department_id
        GROUP BY student_id, u.name, u.profile_data->>'batch', subject_code
        HAVING (COUNT(*) FILTER (WHERE status = 'present' OR status = 'od') * 100.0 / COUNT(*)) < :threshold
        ORDER BY batch, student_name, subject_code
    """)
    
    result = await session.execute(stmt, {
        "college_id": user["college_id"], 
        "department_id": user.get("profile_data", {}).get("department_id", ""),
        "threshold": threshold
    })
    rows = result.all()
    
    return [{
        "student_id": r.student_id,
        "name": r.student_name,
        "batch": r.batch,
        "subject_code": r.subject_code,
        "percentage": float(r.percentage) if r.percentage is not None else 0.0
    } for r in rows]

# ─── Phase 3: Leave Management & Free Periods ────────────────────────────────

@app.post("/api/leave/apply")
async def apply_leave(req: LeaveApply, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    try:
        from_dt = datetime.strptime(req.from_date, "%Y-%m-%d")
        to_dt = datetime.strptime(req.to_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")

    leave = models.LeaveRequest(
        college_id=user["college_id"],
        applicant_id=user["id"],
        applicant_role=user["role"],
        leave_type=req.leave_type,
        from_date=from_dt,
        to_date=to_dt,
        reason=req.reason,
        document_url=req.document_url
    )
    session.add(leave)
    await session.commit()
    await session.refresh(leave)
    return {"id": leave.id, "message": "Leave request submitted"}

@app.patch("/api/leave/{leave_id}/cancel")
async def cancel_leave(leave_id: str, req: LeaveCancelRequest, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    leave_r = await session.execute(
        select(models.LeaveRequest).where(
            models.LeaveRequest.id == leave_id,
            models.LeaveRequest.applicant_id == user["id"]
        )
    )
    leave = leave_r.scalars().first()
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
        
    if leave.status == "pending":
        leave.status = "cancelled"
        await log_audit(session, user["id"], "leave_request", "cancel_pending", {"leave_id": leave.id})
        await session.commit()
        return {"message": "Pending leave cancelled successfully"}
        
    if leave.status != "approved":
        raise HTTPException(status_code=400, detail="Only approved or pending leaves can be cancelled")
        
    # Validation for partial cancellation dates
    if req.cancel_from or req.cancel_to:
        if not req.cancel_from or not req.cancel_to:
            raise HTTPException(status_code=400, detail="Both cancel_from and cancel_to are required for partial cancellation")
        try:
            cancel_from_dt = datetime.strptime(req.cancel_from, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            cancel_to_dt = datetime.strptime(req.cancel_to, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")
            
        if cancel_from_dt < leave.from_date or cancel_to_dt > leave.to_date or cancel_from_dt > cancel_to_dt:
            raise HTTPException(status_code=400, detail="Cancellation dates must be valid and fall within the originally approved leave period")
            
        # Check if the entire request is in the past (already completed)
        if cancel_to_dt.date() < datetime.now(timezone.utc).date() and leave.to_date.date() < datetime.now(timezone.utc).date():
             raise HTTPException(status_code=400, detail="Cannot cancel a leave that has already been fully completed")

        leave.status = "partially_cancelled"
        leave.cancellation_meta = {"cancel_from": req.cancel_from, "cancel_to": req.cancel_to}
    else:
        # Full cancellation of an approved leave
        if leave.to_date.date() < datetime.now(timezone.utc).date():
             raise HTTPException(status_code=400, detail="Cannot cancel a leave that has already been fully completed")
        leave.status = "cancellation_requested"
        leave.cancellation_meta = None

    await log_audit(session, user["id"], "leave_request", "request_cancellation", {"leave_id": leave.id, "partial": bool(req.cancel_from)})
    await session.commit()
    return {"message": "Cancellation request submitted for HOD approval"}

@app.get("/api/leave/my")
async def get_my_leaves(user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.LeaveRequest).where(
            models.LeaveRequest.applicant_id == user["id"],
            models.LeaveRequest.college_id == user["college_id"]
        ).order_by(models.LeaveRequest.created_at.desc())
    )
    leaves = result.scalars().all()
    return [{
        "id": l.id, "leave_type": l.leave_type,
        "from_date": l.from_date.isoformat(), "to_date": l.to_date.isoformat(),
        "reason": l.reason, "status": l.status,
        "reviewed_at": l.reviewed_at.isoformat() if l.reviewed_at else None,
        "review_remarks": l.review_remarks
    } for l in leaves]

@app.get("/api/hod/leave/pending")
async def get_pending_faculty_leaves(user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.LeaveRequest, models.User).join(
            models.User, models.LeaveRequest.applicant_id == models.User.id
        ).where(
            models.LeaveRequest.college_id == user["college_id"],
            models.LeaveRequest.status.in_(["pending", "cancellation_requested", "partially_cancelled"]),
            models.LeaveRequest.applicant_role.in_(["teacher", "faculty"])
        )
    )
    rows = result.all()
    return [{
        "id": l.id, "applicant_id": l.applicant_id, "applicant_name": u.name, "applicant_email": u.email,
        "leave_type": l.leave_type, "from_date": l.from_date.isoformat(), "to_date": l.to_date.isoformat(),
        "reason": l.reason, "document_url": l.document_url, "created_at": l.created_at.isoformat()
    } for l, u in rows]

@app.get("/api/hod/leave/student-pending")
async def get_pending_student_leaves(user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.LeaveRequest, models.User).join(
            models.User, models.LeaveRequest.applicant_id == models.User.id
        ).where(
            models.LeaveRequest.college_id == user["college_id"],
            models.LeaveRequest.status.in_(["pending", "cancellation_requested", "partially_cancelled"]),
            models.LeaveRequest.applicant_role == "student"
        )
    )
    rows = result.all()
    return [{
        "id": l.id, "applicant_id": l.applicant_id, "applicant_name": u.name, 
        "batch": u.profile_data.get("batch") if u.profile_data else None,
        "leave_type": l.leave_type, "from_date": l.from_date.isoformat(), "to_date": l.to_date.isoformat(),
        "reason": l.reason, "created_at": l.created_at.isoformat()
    } for l, u in rows]

@app.put("/api/hod/leave/{leave_id}/review")
async def review_leave(leave_id: str, req: LeaveReview, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    leave_r = await session.execute(
        select(models.LeaveRequest).where(
            models.LeaveRequest.id == leave_id,
            models.LeaveRequest.college_id == user["college_id"]
        )
    )
    leave = leave_r.scalars().first()
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
    if leave.status != "pending":
        raise HTTPException(status_code=400, detail="Leave is already reviewed")

    leave.status = req.action
    leave.reviewed_by = user["id"]
    leave.reviewed_at = datetime.now(timezone.utc)
    leave.review_remarks = req.remarks

    # If approved and faculty, auto-release their slots
    affected_slot_ids = []
    if req.action == "approve" and leave.applicant_role in ["teacher", "faculty"]:
        # Find all period slots for this faculty where the day maps to a date inside the leave window
        # To do this accurately, we should really map dates to days of the week in the range.
        
        # Simple implementation: we'll just release all slots that week if duration > 5 days
        # Otherwise, properly map the specific days of the week affected.
        leave_days = []
        current_date_iter = leave.from_date
        while current_date_iter <= leave.to_date:
            days = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
            day_str = days[current_date_iter.weekday()]
            if day_str not in leave_days:
                leave_days.append(day_str)
            current_date_iter += timedelta(days=1)
        
        if leave_days:
            try:
                current_academic_year = await get_current_academic_year(session, user.get("college_id", ""))
            except HTTPException:
                current_academic_year = None
                
            query = select(models.PeriodSlot).where(
                models.PeriodSlot.faculty_id == leave.applicant_id,
                models.PeriodSlot.day.in_(leave_days)
            )
            if current_academic_year:
                query = query.where(models.PeriodSlot.academic_year == current_academic_year)
                
            slots_r = await session.execute(query)
            slots = slots_r.scalars().all()
            for slot in slots:
                slot.original_faculty_id = slot.faculty_id
                slot.faculty_id = None
                slot.slot_type = "released"
                affected_slot_ids.append(slot.id)
            leave.affected_slots = affected_slot_ids

    await log_audit(session, user["id"], "leave_request", req.action, {"leave_id": leave.id})
    await session.commit()
    return {"message": f"Leave {req.action}d", "affected_slots": len(affected_slot_ids)}

@app.patch("/api/hod/leave/{leave_id}/review-cancellation")
async def review_leave_cancellation(leave_id: str, req: LeaveReview, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    leave_r = await session.execute(
        select(models.LeaveRequest).where(
            models.LeaveRequest.id == leave_id,
            models.LeaveRequest.college_id == user["college_id"]
        )
    )
    leave = leave_r.scalars().first()
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
        
    if leave.status not in ["cancellation_requested", "partially_cancelled"]:
        raise HTTPException(status_code=400, detail="Leave does not have a pending cancellation request")

    if req.action == "reject":
        # Revert status to approved
        leave.status = "approved"
        leave.review_remarks = f"Cancellation Rejected: {req.remarks or ''}"
        leave.cancellation_meta = None
        await log_audit(session, user["id"], "leave_request", "reject_cancellation", {"leave_id": leave.id})
        await session.commit()
        return {"message": "Cancellation rejected. Leave remains approved."}
        
    # Action is approve
    old_status = leave.status
    leave.status = "cancelled" if old_status == "cancellation_requested" else "approved" # if partial, rest is still approved
    
    # Define the range to cancel
    cancel_from = leave.from_date
    cancel_to = leave.to_date
    if old_status == "partially_cancelled" and leave.cancellation_meta:
        cancel_from = datetime.strptime(leave.cancellation_meta.get("cancel_from"), "%Y-%m-%d").replace(tzinfo=timezone.utc)
        cancel_to = datetime.strptime(leave.cancellation_meta.get("cancel_to"), "%Y-%m-%d").replace(tzinfo=timezone.utc)

    # 1. Faculty Free Period / PeriodSlot Reversal
    if leave.applicant_role in ["teacher", "faculty"] and leave.affected_slots:
        # We need to reclaim slots that fall within the cancelled range.
        slots_to_reclaim = []
        slots_r = await session.execute(
            select(models.PeriodSlot).where(
                models.PeriodSlot.id.in_(leave.affected_slots),
                models.PeriodSlot.college_id == user["college_id"]
            )
        )
        for slot in slots_r.scalars().all():
            if slot.slot_type == "released":
                slot.faculty_id = slot.original_faculty_id
                slot.slot_type = "regular"
                slots_to_reclaim.append(slot.id)
                # Remove from affected
                if slot.id in leave.affected_slots:
                    leave.affected_slots.remove(slot.id)
                    
    # 2. Student Attendance Record (system_leave) Deletion
    if leave.applicant_role == "student":
        # Delete AttendanceRecord where source='system_leave' between cancel_from and cancel_to
        await session.execute(
            delete(models.AttendanceRecord).where(
                models.AttendanceRecord.student_id == leave.applicant_id,
                models.AttendanceRecord.college_id == user["college_id"],
                models.AttendanceRecord.source == "system_leave",
                models.AttendanceRecord.date >= cancel_from.date(),
                models.AttendanceRecord.date <= cancel_to.date()
            )
        )

    leave.review_remarks = f"Cancellation Approved: {req.remarks or ''}"
    await log_audit(session, user["id"], "leave_request", "approve_cancellation", {"leave_id": leave.id})
    await session.commit()
    return {"message": "Cancellation approved successfully"}

@app.get("/api/hod/free-periods")
async def get_free_period_pool(user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    """View the released period pool."""
    dept_id = user.get("scope", {}).get("department")
    # For robust production, department scope should be strictly enforced.
    
    stmt = select(models.PeriodSlot, models.User).outerjoin(
        models.User, models.PeriodSlot.original_faculty_id == models.User.id
    ).where(
        models.PeriodSlot.college_id == user["college_id"],
        models.PeriodSlot.slot_type == "released"
    )
    
    result = await session.execute(stmt)
    rows = result.all()
    return [{
        "id": s.id, "day": s.day, "period_no": s.period_no, "start_time": s.start_time, "end_time": s.end_time,
        "batch": s.batch, "section": s.section, "department_id": s.department_id,
        "subject_code": s.subject_code, "subject_name": s.subject_name,
        "original_faculty_name": u.name if u else "Unknown",
        "slot_type": s.slot_type
    } for s, u in rows]

@app.put("/api/hod/free-periods/{slot_id}/assign")
async def assign_substitute_faculty(slot_id: str, req: SubstituteAssign, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    slot_r = await session.execute(
        select(models.PeriodSlot).where(
            models.PeriodSlot.id == slot_id,
            models.PeriodSlot.college_id == user["college_id"]
        )
    )
    slot = slot_r.scalars().first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    if slot.slot_type != "released":
        raise HTTPException(status_code=400, detail="Slot is not in the released pool")

    # Verify target faculty exists
    fac_r = await session.execute(
        select(models.User).where(
            models.User.id == req.faculty_id,
            models.User.college_id == user["college_id"]
        )
    )
    if not fac_r.scalars().first():
        raise HTTPException(status_code=404, detail="Substitute faculty not found")

    slot.faculty_id = req.faculty_id
    slot.slot_type = "substitute"
    
    await log_audit(session, user["id"], "free_period", "assign_substitute", {"slot_id": slot.id, "substitute_id": req.faculty_id})
    await session.commit()
    return {"message": "Substitute assigned successfully"}

# ─── Phase 4: Course Registration & Hall Tickets ────────────────────────────

@app.post("/api/examcell/registration-window")
async def create_registration_window(req: RegistrationWindowCreate, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    try:
        open_dt = datetime.strptime(req.open_at, "%Y-%m-%dT%H:%M")
        close_dt = datetime.strptime(req.close_at, "%Y-%m-%dT%H:%M")
    except ValueError:
        raise HTTPException(status_code=400, detail="Dates must be in YYYY-MM-DDTHH:MM format")

    if close_dt <= open_dt:
        raise HTTPException(status_code=400, detail="close_at must be after open_at")

    # Prevent overlapping windows for the same semester
    dup_r = await session.execute(
        select(models.RegistrationWindow).where(
            models.RegistrationWindow.college_id == user["college_id"],
            models.RegistrationWindow.academic_year == req.academic_year,
            models.RegistrationWindow.semester == req.semester
        )
    )
    if dup_r.scalars().first():
        raise HTTPException(status_code=400, detail="Window already exists for this year and semester")

    window = models.RegistrationWindow(
        college_id=user["college_id"],
        semester=req.semester,
        academic_year=req.academic_year,
        open_at=open_dt,
        close_at=close_dt,
        is_active=False,
        created_by=user["id"]
    )
    session.add(window)
    await session.commit()
    await session.refresh(window)
    return {"id": window.id, "message": "Registration window created"}

@app.get("/api/examcell/registration-window")
async def get_registration_windows(user: dict = Depends(require_role("exam_cell", "admin", "student")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.RegistrationWindow).where(models.RegistrationWindow.college_id == user["college_id"])
    
    # Students only see active windows
    if user["role"] == "student":
        stmt = stmt.where(models.RegistrationWindow.is_active == True)
        
    result = await session.execute(stmt.order_by(models.RegistrationWindow.created_at.desc()))
    windows = result.scalars().all()
    
    return [{
        "id": w.id, "semester": w.semester, "academic_year": w.academic_year,
        "open_at": w.open_at.isoformat(), "close_at": w.close_at.isoformat(),
        "is_active": w.is_active
    } for w in windows]

@app.put("/api/examcell/registration-window/{window_id}/toggle")
async def toggle_registration_window(window_id: str, active: bool = True, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    window_r = await session.execute(
        select(models.RegistrationWindow).where(
            models.RegistrationWindow.id == window_id,
            models.RegistrationWindow.college_id == user["college_id"]
        )
    )
    window = window_r.scalars().first()
    if not window:
        raise HTTPException(status_code=404, detail="Registration window not found")
        
    window.is_active = active
    await log_audit(session, user["id"], "registration_window", "toggle", {"window_id": window.id, "active": active})
    await session.commit()
    return {"message": f"Window {'opened' if active else 'closed'}"}

@app.post("/api/student/register-courses")
async def register_courses(req: List[CourseRegistrationSchema], user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    if not req:
        raise HTTPException(status_code=400, detail="No courses provided")
        
    # Check if a registration window is open for the target semester(s)
    semesters = list(set([r.semester for r in req]))
    academic_years = list(set([r.academic_year for r in req]))
    
    now = datetime.now()
    window_r = await session.execute(
        select(models.RegistrationWindow).where(
            models.RegistrationWindow.college_id == user["college_id"],
            models.RegistrationWindow.semester.in_(semesters),
            models.RegistrationWindow.academic_year.in_(academic_years),
            models.RegistrationWindow.is_active == True,
            models.RegistrationWindow.open_at <= now,
            models.RegistrationWindow.close_at >= now
        )
    )
    active_windows = window_r.scalars().all()
    active_pairs = [(w.semester, w.academic_year) for w in active_windows]
    
    # Check for duplicates using a single IN clause
    subject_codes = [c.subject_code for c in req]
    dup_r = await session.execute(
        select(models.CourseRegistration.subject_code, models.CourseRegistration.academic_year)
        .where(
            models.CourseRegistration.student_id == user["id"],
            models.CourseRegistration.subject_code.in_(subject_codes)
        )
    )
    existing_pairs = set([(row.subject_code, row.academic_year) for row in dup_r.all()])

    inserted = 0
    for course_data in req:
        # Validate window is open
        if (course_data.semester, course_data.academic_year) not in active_pairs:
            raise HTTPException(status_code=400, detail=f"No active registration window for semester {course_data.semester} ({course_data.academic_year})")
            
        if (course_data.subject_code, course_data.academic_year) not in existing_pairs:
            reg = models.CourseRegistration(
                student_id=user["id"],
                college_id=user["college_id"],
                subject_code=course_data.subject_code,
                semester=course_data.semester,
                academic_year=course_data.academic_year,
                is_arrear=course_data.is_arrear,
                status="registered"
            )
            session.add(reg)
            inserted += 1
            
    await log_audit(session, user["id"], "course_registration", "submit", {"count": inserted})
    await session.commit()
    return {"message": f"Successfully registered for {inserted} courses"}

@app.get("/api/student/my-registrations")
async def get_my_registrations(user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.CourseRegistration).where(
            models.CourseRegistration.student_id == user["id"]
        ).order_by(models.CourseRegistration.academic_year.desc(), models.CourseRegistration.semester.desc())
    )
    regs = result.scalars().all()
    return [{
        "id": r.id, "subject_code": r.subject_code, "semester": r.semester,
        "academic_year": r.academic_year, "is_arrear": r.is_arrear,
        "status": r.status, "registered_at": r.registered_at.isoformat()
    } for r in regs]

@app.get("/api/examcell/registrations")
async def view_course_registrations(status: Optional[str] = None, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.CourseRegistration, models.User).join(
        models.User, models.CourseRegistration.student_id == models.User.id
    ).where(models.CourseRegistration.college_id == user["college_id"])
    
    if status:
        stmt = stmt.where(models.CourseRegistration.status == status)
        
    result = await session.execute(stmt)
    rows = result.all()
    
    return [{
        "id": r.id, "student_id": r.student_id, "student_name": u.name, "student_email": u.email,
        "subject_code": r.subject_code, "semester": r.semester, "academic_year": r.academic_year,
        "is_arrear": r.is_arrear, "status": r.status, "registered_at": r.registered_at.isoformat()
    } for r, u in rows]

@app.put("/api/examcell/registrations/{reg_id}/approve")
async def approve_registration(reg_id: str, action: str = Query(..., pattern="^(approve|reject)$"), user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    reg_r = await session.execute(
        select(models.CourseRegistration).where(
            models.CourseRegistration.id == reg_id,
            models.CourseRegistration.college_id == user["college_id"]
        )
    )
    reg = reg_r.scalars().first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
        
    if reg.status != "registered":
        raise HTTPException(status_code=400, detail="Registration is already processed")
        
    reg.status = "approved" if action == "approve" else "rejected"
    reg.reviewed_by = user["id"]
    reg.reviewed_at = datetime.now()
    
    await session.commit()
    return {"message": f"Registration {reg.status}"}

@app.put("/api/examcell/registrations/bulk-approve")
async def bulk_approve_registrations(semester: int, academic_year: str, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    stmt = update(models.CourseRegistration).where(
        models.CourseRegistration.college_id == user["college_id"],
        models.CourseRegistration.semester == semester,
        models.CourseRegistration.academic_year == academic_year,
        models.CourseRegistration.status == "registered"
    ).values(
        status="approved",
        reviewed_by=user["id"],
        reviewed_at=datetime.now()
    )
    
    result = await session.execute(stmt)
    await log_audit(session, user["id"], "course_registration", "bulk_approve", {"semester": semester, "affected": result.rowcount})
    await session.commit()
    
    return {"message": f"Bulk approved {result.rowcount} pending registrations"}

@app.post("/api/examcell/hall-tickets/generate")
async def generate_hall_tickets(semester: int, academic_year: str, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    # Verify that all registrations for this semester are approved
    pending_r = await session.execute(
        select(func.count(models.CourseRegistration.id)).where(
            models.CourseRegistration.college_id == user["college_id"],
            models.CourseRegistration.semester == semester,
            models.CourseRegistration.academic_year == academic_year,
            models.CourseRegistration.status == "registered"
        )
    )
    pending_count = pending_r.scalar()
    if pending_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot generate hall tickets. There are {pending_count} pending registrations.")
        
    # In a fully fleshed out system, this would trigger a background worker (e.g. Celery) 
    # to generate PDFs and send emails. For AcadMix, we just log the generation event.
    await log_audit(session, user["id"], "hall_ticket", "generate", {"semester": semester, "academic_year": academic_year})
    
    return {"message": "Hall tickets generation triggered successfully. Students can now download them."}

# ─── Exam Cell & Student: Schedules and Hall Tickets ─────────────────────────

@app.get("/api/examcell/settings")
async def get_college_settings(user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.College).where(models.College.id == user["college_id"]))
    college = result.scalars().first()
    return {"settings": college.settings if college and college.settings else {}}

@app.put("/api/examcell/settings")
async def update_college_settings(req: CollegeSettingsUpdate, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.College).where(models.College.id == user["college_id"]))
    college = result.scalars().first()
    if not college:
        raise HTTPException(status_code=404, detail="College not found")
        
    college.settings = req.settings
    await log_audit(session, user["id"], "examcell_settings", "update", {})
    await session.commit()
    return {"message": "Settings updated successfully", "settings": college.settings}

@app.post("/api/examcell/schedule")
async def create_exam_schedule(req: ExamScheduleCreate, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    try:
        dt = datetime.strptime(req.exam_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="exam_date must be YYYY-MM-DD")
        
    sched = models.ExamSchedule(
        college_id=user["college_id"],
        department_id=req.department_id,
        batch=req.batch,
        semester=req.semester,
        academic_year=req.academic_year,
        subject_code=req.subject_code,
        subject_name=req.subject_name,
        exam_date=dt,
        session=req.session,
        exam_time=req.exam_time,
        document_url=req.document_url,
        created_by=user["id"]
    )
    session.add(sched)
    await session.commit()
    await session.refresh(sched)
    return sched

@app.get("/api/examcell/schedule")
async def get_exam_schedules(department_id: Optional[str] = None, batch: Optional[str] = None, semester: Optional[int] = None, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.ExamSchedule).where(models.ExamSchedule.college_id == user["college_id"])
    if department_id:
        stmt = stmt.where(models.ExamSchedule.department_id == department_id)
    if batch:
        stmt = stmt.where(models.ExamSchedule.batch == batch)
    if semester:
        stmt = stmt.where(models.ExamSchedule.semester == semester)
        
    result = await session.execute(stmt.order_by(models.ExamSchedule.exam_date.desc()))
    items = result.scalars().all()
    out = []
    for s in items:
        out.append({
            "id": s.id,
            "department_id": s.department_id,
            "batch": s.batch,
            "semester": s.semester,
            "academic_year": s.academic_year,
            "subject_code": s.subject_code,
            "subject_name": s.subject_name,
            "exam_date": str(s.exam_date) if s.exam_date else None,
            "session": s.session,
            "exam_time": s.exam_time,
            "document_url": s.document_url,
            "is_published": s.is_published
        })
    return out

@app.put("/api/examcell/schedule/{id}/publish")
async def toggle_exam_schedule_publish(id: str, published: bool = Query(..., description="Set to true to publish, false to unpublish"), user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.ExamSchedule).where(models.ExamSchedule.id == id, models.ExamSchedule.college_id == user["college_id"]))
    sched = result.scalars().first()
    if not sched:
        raise HTTPException(status_code=404, detail="Exam schedule not found")
        
    sched.is_published = published
    await session.commit()
    return {"message": f"Schedule {'published' if published else 'unpublished'}"}

@app.delete("/api/examcell/schedule/{id}")
async def delete_exam_schedule(id: str, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.ExamSchedule).where(models.ExamSchedule.id == id, models.ExamSchedule.college_id == user["college_id"]))
    sched = result.scalars().first()
    if not sched:
        raise HTTPException(status_code=404, detail="Exam schedule not found")
    
    sched.is_deleted = True
    sched.deleted_at = func.now()
    await log_audit(session, user["id"], "examcell_schedule", "delete", {"schedule_id": id})
    await session.commit()
    return {"message": "Exam schedule deleted"}

@app.get("/api/student/exam-schedule")
async def get_student_exam_schedule(user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    profile = user.get("profile_data") or {}
    dept_str = user.get("department") or profile.get("department", "")
    batch_str = user.get("batch") or profile.get("batch", "")
    
    dept_r = await session.execute(
        select(models.Department).where(
            models.Department.college_id == user["college_id"],
            or_(models.Department.code == dept_str, models.Department.name == dept_str)
        )
    )
    dept = dept_r.scalars().first()
    if not dept:
        return []

    stmt = select(models.ExamSchedule).where(
        models.ExamSchedule.college_id == user["college_id"],
        models.ExamSchedule.department_id == dept.id,
        models.ExamSchedule.batch == batch_str,
        models.ExamSchedule.is_published == True
    ).order_by(models.ExamSchedule.exam_date.asc())
    
    result = await session.execute(stmt)
    return result.scalars().all()

@app.get("/api/examcell/dashboard-stats")
async def get_examcell_dashboard_stats(user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    """
    Dashboard Stats definitions:
    - Schedules: Published versus unpublished entries in the ExamSchedule model for the tenant.
    - Hall Tickets (generated): A hall ticket is considered "generated" when a student has an 
      approved CourseRegistration AND a published ExamSchedule exists for that specific semester/year.
    """
    sched_r = await session.execute(
        select(models.ExamSchedule.is_published, func.count(models.ExamSchedule.id))
        .where(models.ExamSchedule.college_id == user["college_id"])
        .group_by(models.ExamSchedule.is_published)
    )
    sched_counts = {str(k): v for k, v in sched_r.all()}
    
    regs_r = await session.execute(
        select(func.count(models.CourseRegistration.id)).where(
            models.CourseRegistration.college_id == user["college_id"],
            models.CourseRegistration.status == "approved"
        )
    )
    total_approved = regs_r.scalar() or 0
    
    # Single SQL aggregation to avoid multiple roundtrips and Python loops for resolution
    gen_stmt = select(func.count(models.CourseRegistration.id.distinct())).join(
        models.ExamSchedule,
        and_(
            models.CourseRegistration.college_id == models.ExamSchedule.college_id,
            models.CourseRegistration.semester == models.ExamSchedule.semester,
            models.CourseRegistration.academic_year == models.ExamSchedule.academic_year,
            models.ExamSchedule.is_published == True
        )
    ).where(
        models.CourseRegistration.college_id == user["college_id"],
        models.CourseRegistration.status == "approved"
    )
    
    gen_r = await session.execute(gen_stmt)
    total_generated = gen_r.scalar() or 0
        
    return {
        "schedules": {
            "published": int(sched_counts.get("True", 0)),
            "unpublished": int(sched_counts.get("False", 0))
        },
        "hall_tickets": {
            "total_approved": total_approved,
            "generated": total_generated
        }
    }

@app.get("/api/student/hall-ticket", response_class=HTMLResponse)
async def get_hall_ticket(semester: int, academic_year: str, user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    # Phase C: Fee Gate
    fee_r = await session.execute(
        select(models.FeePayment, models.FeeTemplate)
        .join(models.FeeTemplate, models.FeePayment.fee_template_id == models.FeeTemplate.id)
        .where(
            models.FeePayment.student_id == user["id"],
            models.FeeTemplate.academic_year == academic_year,
            models.FeeTemplate.semester == semester
        )
    )
    fee_records = fee_r.all()
    for payment, template in fee_records:
        if template.fee_type == "exam" and payment.status != "paid":
            raise HTTPException(status_code=402, detail="Pending Dues: Your exam fee must be paid before downloading the hall ticket.")

    regs_r = await session.execute(
        select(models.CourseRegistration).where(
            models.CourseRegistration.student_id == user["id"],
            models.CourseRegistration.semester == semester,
            models.CourseRegistration.academic_year == academic_year,
            models.CourseRegistration.status == "approved"
        )
    )
    regs = regs_r.scalars().all()
    
    if not regs:
        raise HTTPException(status_code=404, detail="No approved registrations found for this semester. Hall ticket unavailable.")
        
    profile = user.get("profile_data") or {}
    dept_str = user.get("department") or profile.get("department", "")
    batch_str = user.get("batch") or profile.get("batch", "")
    
    dept_r = await session.execute(
        select(models.Department).where(
            models.Department.college_id == user["college_id"],
            or_(models.Department.code == dept_str, models.Department.name == dept_str)
        )
    )
    dept = dept_r.scalars().first()
    dept_id = dept.id if dept else ""
    
    scheds = []
    if dept_id:
        sched_r = await session.execute(
            select(models.ExamSchedule).where(
                models.ExamSchedule.college_id == user["college_id"],
                models.ExamSchedule.department_id == dept_id,
                models.ExamSchedule.batch == batch_str,
                models.ExamSchedule.semester == semester,
                models.ExamSchedule.academic_year == academic_year,
                models.ExamSchedule.is_published == True
            )
        )
        scheds = sched_r.scalars().all()
    
    sched_map = { s.subject_code: s for s in scheds }
    sched_map = { s.subject_code: s for s in scheds }
    
    photo_url = profile.get("photo_url")
    if photo_url:
        photo_html = f'<img src="{photo_url}" alt="Student Photo" style="width:100px; height:120px; object-fit:cover; border: 1px solid #ccc;">'
    else:
        initials = "".join([n[0] for n in user["name"].split() if n])[:2].upper()
        photo_html = f'<div style="width:100px; height:120px; border: 1px solid #ccc; display:flex; align-items:center; justify-content:center; background:#eee; font-size:32px; font-weight:bold; color:#888;">{initials}</div>'

    rows = ""
    for r in regs:
        s = sched_map.get(r.subject_code)
        
        status_text = "Pending" if not s else "Published"
        dt = s.exam_date.strftime("%d-%b-%Y") if s and s.exam_date else "Pending Schedule"
        session_text = f"{s.session} ({s.exam_time})" if s else "Pending Schedule"
        hall = s.document_url if s and s.document_url else "Check Notice Board"
        
        # Determine color for the status field
        color = "red" if not s else "green"
        
        rows += f"""
        <tr>
            <td style="padding:8px; border:1px solid #ddd;">{r.subject_code}</td>
            <td style="padding:8px; border:1px solid #ddd;">{s.subject_name if s else "Subject TBA"}</td>
            <td style="padding:8px; border:1px solid #ddd; text-align:center;">{"Arrear" if r.is_arrear else "Regular"}</td>
            <td style="padding:8px; border:1px solid #ddd; text-align:center;">{dt}</td>
            <td style="padding:8px; border:1px solid #ddd; text-align:center;">{session_text}</td>
            <td style="padding:8px; border:1px solid #ddd; text-align:center;">{hall}</td>
            <td style="padding:8px; border:1px solid #ddd; text-align:center; color:{color}; font-weight:bold;">{status_text}</td>
        </tr>
        """
        
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Hall Ticket - {user['name']}</title>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 800px; margin: 0 auto; padding: 20px; }}
            .header {{ text-align: center; border-bottom: 2px solid #333; padding-bottom: 15px; margin-bottom: 20px; }}
            .header h1 {{ margin: 0; font-size: 24px; text-transform: uppercase; }}
            .header p {{ margin: 5px 0 0 0; font-size: 16px; color: #666; }}
            .student-info {{ display: flex; justify-content: space-between; margin-bottom: 30px; }}
            .info-grid {{ display: grid; grid-template-columns: 120px 1fr; gap: 10px; }}
            .info-label {{ font-weight: bold; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; }}
            th {{ background: #f4f4f4; padding: 10px; border: 1px solid #ddd; text-align: center; }}
            .footer {{ display: flex; justify-content: space-between; margin-top: 50px; text-align: center; }}
            .signature-box {{ width: 200px; padding-top: 50px; border-top: 1px solid #333; }}
            @media print {{
                .no-print {{ display: none; }}
                body {{ padding: 0; }}
            }}
            .print-btn {{ display: block; margin: 20px auto; padding: 10px 20px; background: #007BFF; color: white; border: none; cursor: pointer; border-radius: 4px; font-size: 16px; }}
            .print-btn:hover {{ background: #0056b3; }}
        </style>
    </head>
    <body>
        <div class="no-print" style="text-align:center;">
            <button class="print-btn" onclick="window.print()">Print Hall Ticket</button>
        </div>
        
        <div class="header">
            <h1>Hall Ticket</h1>
            <p>End Semester Examination - {academic_year}</p>
        </div>
        
        <div class="student-info">
            <div class="info-grid">
                <div class="info-label">Name:</div><div>{user['name']}</div>
                <div class="info-label">Reg No:</div><div>{profile.get("college_id", user["id"])}</div>
                <div class="info-label">Department:</div><div>{dept_str}</div>
                <div class="info-label">Semester:</div><div>{semester}</div>
            </div>
            <div class="photo">
                {photo_html}
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Subject Code</th>
                    <th>Subject Name</th>
                    <th>Type</th>
                    <th>Date</th>
                    <th>Session</th>
                    <th>Hall / Room Allocation</th>
                    <th>Schedule Status</th>
                </tr>
            </thead>
            <tbody>
                {rows}
            </tbody>
        </table>
        
        <div class="footer">
            <div class="signature-box">Signature of Student</div>
            <div class="signature-box">Controller of Examinations</div>
        </div>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

# ─── Phase 5: NAAC Institutional Data ───────────────────────────────────────


@app.get("/api/principal/institution-profile")
async def get_institution_profile(user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.InstitutionProfile).where(
            models.InstitutionProfile.college_id == user["college_id"]
        )
    )
    profile = result.scalars().first()
    
    if not profile:
        return {}
        
    return {
        "id": profile.id,
        "recognitions": profile.recognitions,
        "infrastructure": profile.infrastructure,
        "library": profile.library,
        "mous": profile.mous,
        "extension_activities": profile.extension_activities,
        "research_publications": profile.research_publications,
        "updated_at": profile.updated_at.isoformat() if profile.updated_at else None
    }

@app.put("/api/principal/institution-profile")
async def update_institution_profile(req: InstitutionProfileUpdate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.InstitutionProfile).where(
            models.InstitutionProfile.college_id == user["college_id"]
        )
    )
    profile = result.scalars().first()
    
    if not profile:
        profile = models.InstitutionProfile(
            college_id=user["college_id"],
            updated_by=user["id"]
        )
        session.add(profile)
    else:
        profile.updated_by = user["id"]
        
    # Update fields that were provided
    if req.recognitions is not None:
        profile.recognitions = req.recognitions
    if req.infrastructure is not None:
        profile.infrastructure = req.infrastructure
    if req.library is not None:
        profile.library = req.library
    if req.mous is not None:
        profile.mous = req.mous
    if req.extension_activities is not None:
        profile.extension_activities = req.extension_activities
    if req.research_publications is not None:
        profile.research_publications = req.research_publications
        
    await log_audit(session, user["id"], "institution_profile", "update", {})
    await session.commit()
    
    return {"message": "Institution profile updated successfully"}

# ─── User Routes ────────────────────────────────────────────────────────────
@app.get("/api/users")
async def list_users(
    role: Optional[str] = None, 
    limit: int = Query(50, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    user: dict = Depends(require_role("admin", "teacher", "hod", "exam_cell")), 
    session: AsyncSession = Depends(get_db)):
    stmt = select(models.User).where(models.User.college_id == user["college_id"])
    if role:
        stmt = stmt.where(models.User.role == role)
    result = await session.execute(stmt.offset(offset).limit(limit))
    users = result.scalars().all()
    return [{
        "id": u.id, "name": u.name, "email": u.email, "role": u.role,
        "college_id": u.college_id, **(u.profile_data or {})
    } for u in users]

@app.get("/api/users/{user_id}")
async def get_user(user_id: str, user: dict = Depends(require_role("admin", "teacher", "hod", "exam_cell")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.User).where(models.User.id == user_id))
    u = result.scalars().first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    return {"id": u.id, "name": u.name, "email": u.email, "role": u.role, "college_id": u.college_id, **(u.profile_data or {})}

@app.post("/api/users")
async def create_user(req: RegisterRequest, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    # Check duplicate college_id inside profile_data
    result = await session.execute(
        select(models.User).where(
            models.User.profile_data["college_id"].astext == req.college_id.upper()
        )
    )
    existing = result.scalars().first()
    if existing:
        raise HTTPException(status_code=400, detail="College ID already exists")
    new_user = models.User(
        name=req.name,
        email=req.email.lower(),
        password_hash=hash_password(req.password),
        role=req.role,
        college_id=user["college_id"],
        profile_data={
            "college_id": req.college_id.upper(),
            "college": req.college,
            "department": req.department,
            "batch": req.batch,
            "section": req.section,
        }
    )
    session.add(new_user)
    await session.commit()
    await session.refresh(new_user)
    return {"id": new_user.id, "name": new_user.name, "email": new_user.email, "role": new_user.role, **(new_user.profile_data or {})}

@app.delete("/api/users/{user_id}")
async def delete_user(user_id: str, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.User).where(models.User.id == user_id))
    u = result.scalars().first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    await session.delete(u)
    await session.commit()
    return {"message": "User deleted"}

@app.put("/api/users/{user_id}")
async def update_user(user_id: str, req: UserUpdate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.User).where(models.User.id == user_id))
    u = result.scalars().first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
        
    if req.name is not None:
        u.name = req.name
    if req.email is not None:
        u.email = req.email.lower()
    if req.role is not None:
        u.role = req.role
    if req.password is not None and req.password.strip():
        u.password_hash = hash_password(req.password)
        
    profile = dict(u.profile_data or {})
    if req.college_id is not None:
        profile["college_id"] = req.college_id.upper()
    if req.department is not None:
        profile["department"] = req.department
    if req.batch is not None:
        profile["batch"] = req.batch
    if req.section is not None:
        profile["section"] = req.section
        
    u.profile_data = profile
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(u, "profile_data")
    
    await session.commit()
    await session.refresh(u)
    return {"id": u.id, "name": u.name, "email": u.email, "role": u.role, "college_id": u.college_id, **(u.profile_data or {})}

@app.get("/api/departments")
async def list_departments(user: dict = Depends(require_role("admin", "teacher", "hod", "exam_cell")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.Department).where(models.Department.college_id == user["college_id"])
    result = await session.execute(stmt)
    return result.scalars().all()

@app.post("/api/departments")
async def create_department(req: DepartmentCreate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    new_dept = models.Department(
        college_id=user["college_id"],
        name=req.name,
        code=req.code.upper()
    )
    session.add(new_dept)
    await session.commit()
    await session.refresh(new_dept)
    return new_dept

@app.put("/api/departments/{dept_id}")
async def update_department(dept_id: str, req: DepartmentUpdate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.Department).where(models.Department.id == dept_id))
    dept = result.scalars().first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    if req.name is not None:
        dept.name = req.name
    if req.code is not None:
        dept.code = req.code.upper()
    await session.commit()
    await session.refresh(dept)
    return dept

@app.delete("/api/departments/{dept_id}")
async def delete_department(dept_id: str, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.Department).where(models.Department.id == dept_id))
    dept = result.scalars().first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    await session.delete(dept)
    await session.commit()
    return {"message": "Department deleted"}

@app.get("/api/sections")
async def list_sections(user: dict = Depends(require_role("admin", "teacher", "hod", "exam_cell", "student")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.Section).where(models.Section.college_id == user["college_id"])
    result = await session.execute(stmt)
    return result.scalars().all()

@app.post("/api/sections")
async def create_section(req: SectionCreate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    # Verify department belongs to college
    dept_res = await session.execute(select(models.Department).where(models.Department.id == req.department_id, models.Department.college_id == user["college_id"]))
    if not dept_res.scalars().first():
        raise HTTPException(status_code=400, detail="Invalid department")
    new_sec = models.Section(college_id=user["college_id"], department_id=req.department_id, name=req.name.upper())
    session.add(new_sec)
    await session.commit()
    await session.refresh(new_sec)
    return new_sec

@app.put("/api/sections/{sec_id}")
async def update_section(sec_id: str, req: SectionUpdate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.Section).where(models.Section.id == sec_id, models.Section.college_id == user["college_id"]))
    sec = result.scalars().first()
    if not sec:
        raise HTTPException(status_code=404, detail="Section not found")
    if req.name is not None:
        sec.name = req.name.upper()
    if req.department_id is not None:
        sec.department_id = req.department_id
    await session.commit()
    await session.refresh(sec)
    return sec

@app.delete("/api/sections/{sec_id}")
async def delete_section(sec_id: str, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.Section).where(models.Section.id == sec_id, models.Section.college_id == user["college_id"]))
    sec = result.scalars().first()
    if not sec:
        raise HTTPException(status_code=404, detail="Section not found")
    await session.delete(sec)
    await session.commit()
    return {"message": "Section deleted"}

@app.get("/api/roles")
async def list_roles(user: dict = Depends(require_role("admin", "teacher", "hod")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.Role).where(models.Role.college_id == user["college_id"])
    result = await session.execute(stmt)
    return result.scalars().all()

@app.post("/api/roles")
async def create_role(req: RoleCreate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    new_role = models.Role(college_id=user["college_id"], name=req.name, permissions=req.permissions)
    session.add(new_role)
    await session.commit()
    await session.refresh(new_role)
    return new_role

@app.put("/api/roles/{role_id}")
async def update_role(role_id: str, req: RoleUpdate, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.Role).where(models.Role.id == role_id, models.Role.college_id == user["college_id"]))
    role = result.scalars().first()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    if req.name is not None:
        role.name = req.name
    if req.permissions is not None:
        role.permissions = req.permissions
    await session.commit()
    await session.refresh(role)
    return role

@app.delete("/api/roles/{role_id}")
async def delete_role(role_id: str, user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.Role).where(models.Role.id == role_id, models.Role.college_id == user["college_id"]))
    role = result.scalars().first()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    await session.delete(role)
    await session.commit()
    return {"message": "Role deleted"}

# ─── Quiz Routes ────────────────────────────────────────────────────────────
@app.get("/api/quizzes")
async def list_quizzes(status: Optional[str] = None, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    stmt = select(models.Quiz).where(models.Quiz.college_id == user["college_id"])
    if status:
        stmt = stmt.where(models.Quiz.type == status)
    if user["role"] == "teacher":
        stmt = stmt.where(models.Quiz.faculty_id == user["id"])
    result = await session.execute(stmt.order_by(models.Quiz.created_at.desc()))
    quizzes = result.scalars().all()
    out = []
    for q in quizzes:
        # Fetch questions
        qr = await session.execute(select(models.Question).where(models.Question.quiz_id == q.id))
        questions = qr.scalars().all()
        q_list = []
        for question in questions:
            qd = {"id": question.id, "type": question.type, "marks": question.marks, **(question.content or {})}
            if user["role"] == "student":
                qd.pop("correct_answer", None)
                qd.pop("correct_answers", None)
                qd.pop("keywords", None)
            q_list.append(qd)
        out.append({
            "id": q.id, "title": q.title, "subject": q.type,
            "duration_mins": q.duration_minutes, "created_at": q.created_at.isoformat() if q.created_at else None,
            "faculty_id": q.faculty_id, "college_id": q.college_id,
            "questions": q_list,
        })
    return out

@app.post("/api/quizzes")
async def create_quiz(req: QuizCreate, user: dict = Depends(require_permission("quizzes", "create")), session: AsyncSession = Depends(get_db)):
    new_quiz = models.Quiz(
        college_id=user["college_id"],
        faculty_id=user["id"],
        title=req.title,
        duration_minutes=req.duration_mins,
        type=req.subject,
    )
    session.add(new_quiz)
    await session.flush()  # get generated ID before adding questions

    for q in req.questions:
        question = models.Question(
            quiz_id=new_quiz.id,
            type=q.get("type", "mcq"),
            marks=q.get("marks", 1),
            points=q.get("points", 1),
            content=q,  # store full question dict as JSONB
        )
        session.add(question)

    await log_audit(session, user["id"], "quiz", "create", {"title": req.title})
    await session.commit()
    await session.refresh(new_quiz)
    return {"id": new_quiz.id, "title": new_quiz.title, "subject": new_quiz.type, "duration_mins": new_quiz.duration_minutes}

@app.get("/api/quizzes/{quiz_id}")
async def get_quiz(quiz_id: str, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.Quiz).where(models.Quiz.id == quiz_id))
    q = result.scalars().first()
    if not q:
        raise HTTPException(status_code=404, detail="Quiz not found")
    qr = await session.execute(select(models.Question).where(models.Question.quiz_id == q.id).order_by(models.Question.id))
    questions = qr.scalars().all()
    q_list = []
    for question in questions:
        qd = {"id": question.id, "type": question.type, "marks": question.marks, **(question.content or {})}
        if user["role"] == "student":
            qd.pop("correct_answer", None)
            qd.pop("correct_answers", None)
            qd.pop("keywords", None)
        q_list.append(qd)
    return {
        "id": q.id, "title": q.title, "subject": q.type,
        "duration_mins": q.duration_minutes, "college_id": q.college_id,
        "faculty_id": q.faculty_id, "questions": q_list,
        "created_at": q.created_at.isoformat() if q.created_at else None,
    }

@app.get("/api/quizzes/live/{quiz_id}")
async def live_quiz_monitor(quiz_id: str, user: dict = Depends(require_role("teacher", "admin", "hod", "exam_cell")), session: AsyncSession = Depends(get_db)):
    quiz_r = await session.execute(select(models.Quiz).where(models.Quiz.id == quiz_id))
    q = quiz_r.scalars().first()
    if not q:
        raise HTTPException(status_code=404, detail="Quiz not found")
    attempts_r = await session.execute(
        select(models.QuizAttempt).where(models.QuizAttempt.quiz_id == quiz_id)
    )
    attempts = attempts_r.scalars().all()
    student_ids = [a.student_id for a in attempts]
    
    # Pre-fetch answers for efficiency
    answers_by_attempt = {}
    if attempts:
        ans_r = await session.execute(select(models.QuizAnswer).where(models.QuizAnswer.attempt_id.in_([a.id for a in attempts])))
        for ans in ans_r.scalars().all():
            answers_by_attempt.setdefault(ans.attempt_id, []).append(ans)
            
    # Pre-fetch total questions
    q_questions_r = await session.execute(select(models.Question).where(models.Question.quiz_id == quiz_id))
    total_questions = len(q_questions_r.scalars().all())

    students_r = await session.execute(
        select(models.User).where(models.User.id.in_(student_ids))
    )
    students_dict = {s.id: s for s in students_r.scalars().all()}
    live_data = []
    for a in attempts:
        if not a.start_time:
            continue
        student = students_dict.get(a.student_id)
        if a.status == "in_progress":
            td = datetime.now(timezone.utc) - a.start_time.replace(tzinfo=timezone.utc)
            time_elapsed = int(td.total_seconds() / 60)
            submit_time_str = None
        else:
            end = a.end_time or a.start_time
            td = end.replace(tzinfo=timezone.utc) - a.start_time.replace(tzinfo=timezone.utc)
            time_elapsed = int(td.total_seconds() / 60)
            submit_time_str = end.strftime("%I:%M %p")
            
        answers = answers_by_attempt.get(a.id, [])
        progress = len(answers)
        
        live_data.append({
            "id": a.id,
            "name": student.name if student else "Unknown",
            "rollNo": (student.profile_data or {}).get("college_id", a.student_id) if student else a.student_id,
            "status": "active" if a.status == "in_progress" else "submitted",
            "progress": progress, "totalQuestions": total_questions,
            "violations": 0, "timeElapsed": max(0, time_elapsed),
            "startTime": a.start_time.strftime("%I:%M %p"),
            "submitTime": submit_time_str
        })
    return live_data

@app.patch("/api/quizzes/{quiz_id}")
async def update_quiz(quiz_id: str, updates: dict, user: dict = Depends(require_role("teacher", "admin")), session: AsyncSession = Depends(get_db)):
    result_r = await session.execute(select(models.Quiz).where(models.Quiz.id == quiz_id))
    q = result_r.scalars().first()
    if not q:
        raise HTTPException(status_code=404, detail="Quiz not found")
    if "title" in updates: q.title = updates["title"]
    if "subject" in updates: q.type = updates["subject"]
    if "status" in updates: q.status = updates["status"]
    if "duration_mins" in updates: q.duration_minutes = updates["duration_mins"]
    if "total_marks" in updates: q.total_marks = updates["total_marks"]
    if "questions" in updates:
        q.questions = updates["questions"]
        q.total_marks = sum(qu.get("marks", 0) for qu in updates["questions"])
    await session.commit()
    return {"id": q.id, "title": q.title, "status": q.status, "subject": q.type}

@app.delete("/api/quizzes/{quiz_id}")
async def delete_quiz(quiz_id: str, user: dict = Depends(require_permission("quizzes", "delete")), session: AsyncSession = Depends(get_db)):
    result_r = await session.execute(select(models.Quiz).where(models.Quiz.id == quiz_id))
    quiz = result_r.scalars().first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    await log_audit(session, user["id"], "quiz", "delete", {"quiz_id": quiz.id, "title": quiz.title})
    await session.delete(quiz)
    await session.commit()
    return {"message": "Quiz deleted"}

@app.post("/api/quizzes/{quiz_id}/publish")
async def publish_quiz(quiz_id: str, user: dict = Depends(require_permission("quizzes", "publish")), session: AsyncSession = Depends(get_db)):
    result_r = await session.execute(select(models.Quiz).where(models.Quiz.id == quiz_id))
    quiz = result_r.scalars().first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    quiz.status = "active"
    await log_audit(session, user["id"], "quiz", "publish", {"quiz_id": quiz.id, "title": quiz.title})
    await session.commit()
    return {"message": "Quiz published"}

@app.post("/api/quizzes/{quiz_id}/extend-time")
async def extend_quiz_time(quiz_id: str, body: dict = {}, user: dict = Depends(require_role("teacher", "admin", "hod")), session: AsyncSession = Depends(get_db)):
    mins = int(body.get("mins", 10)) if body else 10
    if mins < 1 or mins > 300:
        raise HTTPException(status_code=400, detail="Minutes must be between 1 and 300")
    result_r = await session.execute(select(models.Quiz).where(models.Quiz.id == quiz_id))
    quiz = result_r.scalars().first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    quiz.duration_minutes = (quiz.duration_minutes or 60) + mins
    await session.commit()
    return {"message": f"Extended by {mins} mins. New duration: {quiz.duration_minutes} mins", "duration_mins": quiz.duration_minutes}

@app.post("/api/quizzes/{quiz_id}/end")
async def end_quiz_now(quiz_id: str, user: dict = Depends(require_role("teacher", "admin", "hod")), session: AsyncSession = Depends(get_db)):
    quiz_r = await session.execute(select(models.Quiz).where(models.Quiz.id == quiz_id))
    quiz = quiz_r.scalars().first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    # Force-submit all in-progress attempts
    atts_r = await session.execute(
        select(models.QuizAttempt).where(
            models.QuizAttempt.quiz_id == quiz_id,
            models.QuizAttempt.status == "in_progress"
        )
    )
    force_submitted = 0
    for attempt in atts_r.scalars().all():
        attempt.status = "submitted"
        attempt.end_time = datetime.now(timezone.utc)
        attempt.final_score = 0
        force_submitted += 1
    quiz.status = "ended"
    await session.commit()
    return {"message": f"Quiz ended. {force_submitted} attempts force-submitted.", "force_submitted": force_submitted}

@app.post("/api/quizzes/{quiz_id}/start")
@limiter.limit("5/minute")
async def start_attempt(quiz_id: str, request: Request, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.Quiz).where(models.Quiz.id == quiz_id))
    quiz = result.scalars().first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")

    # Check for existing submitted attempt (no reattempt)
    existing_r = await session.execute(
        select(models.QuizAttempt).where(
            models.QuizAttempt.quiz_id == quiz_id,
            models.QuizAttempt.student_id == user["id"],
            models.QuizAttempt.status == "submitted"
        )
    )
    if existing_r.scalars().first():
        raise HTTPException(status_code=400, detail="Already attempted this quiz")

    # Return existing in-progress attempt
    prog_r = await session.execute(
        select(models.QuizAttempt).where(
            models.QuizAttempt.quiz_id == quiz_id,
            models.QuizAttempt.student_id == user["id"],
            models.QuizAttempt.status == "in_progress"
        )
    )
    in_progress = prog_r.scalars().first()
    if in_progress:
        return {"id": in_progress.id, "quiz_id": in_progress.quiz_id, "student_id": in_progress.student_id,
                "status": in_progress.status, "start_time": in_progress.start_time.isoformat() if in_progress.start_time else None}

    attempt = models.QuizAttempt(
        quiz_id=quiz_id,
        student_id=user["id"],
        status="in_progress",
        start_time=datetime.now(timezone.utc),
        final_score=0,
    )
    session.add(attempt)
    await session.commit()
    await session.refresh(attempt)
    return {"id": attempt.id, "quiz_id": attempt.quiz_id, "student_id": attempt.student_id,
            "status": attempt.status, "start_time": attempt.start_time.isoformat()}

@app.post("/api/attempts/{attempt_id}/answer")
@limiter.limit("60/minute")
async def submit_answer(attempt_id: str, req: AnswerSubmit, request: Request, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.QuizAttempt).where(
            models.QuizAttempt.id == attempt_id,
            models.QuizAttempt.student_id == user["id"]
        )
    )
    attempt = result.scalars().first()
    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")
    if attempt.status != "in_progress":
        raise HTTPException(status_code=400, detail="Attempt already submitted")

    # Map index to actual Question UUID
    qr = await session.execute(
        select(models.Question.id)
        .where(models.Question.quiz_id == attempt.quiz_id)
        .order_by(models.Question.id)
    )
    q_ids = qr.scalars().all()
    if req.question_index < 0 or req.question_index >= len(q_ids):
        raise HTTPException(status_code=400, detail="Invalid question index")
    actual_question_id = q_ids[req.question_index]

    # Save answer as a QuizAnswer row
    existing_ans = await session.execute(
        select(models.QuizAnswer).where(
            models.QuizAnswer.attempt_id == attempt_id,
            models.QuizAnswer.question_id == actual_question_id
        )
    )
    ans_row = existing_ans.scalars().first()
    if ans_row:
        ans_row.code_submitted = str(req.answer) if req.answer is not None else None
    else:
        ans_row = models.QuizAnswer(
            attempt_id=attempt_id,
            question_id=actual_question_id,
            code_submitted=str(req.answer) if req.answer is not None else None,
        )
        session.add(ans_row)
    # Track first interaction time
    if not attempt.start_time:
        attempt.start_time = datetime.now(timezone.utc)
    await session.commit()
    return {"message": "Answer saved", "question_index": req.question_index}

@app.post("/api/attempts/{attempt_id}/violation")
async def log_violation(attempt_id: str, req: ViolationReport = ViolationReport(), user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.QuizAttempt).where(models.QuizAttempt.id == attempt_id))
    attempt = result.scalars().first()
    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")
    violation = models.ProctoringViolation(
        attempt_id=attempt_id,
        violation_type=req.violation_type,
        suspicion_score=1.0,
    )
    session.add(violation)
    await log_audit(session, user["id"], "proctoring_violation", "create", {"attempt_id": attempt_id, "type": req.violation_type})
    await session.commit()
    # Count total violations for this attempt
    count_r = await session.execute(
        select(models.ProctoringViolation).where(models.ProctoringViolation.attempt_id == attempt_id)
    )
    total = len(count_r.scalars().all())
    return {"message": "Violation logged", "total_violations": total}

@app.post("/api/attempts/{attempt_id}/submit")
@limiter.limit("3/minute")
async def submit_attempt(attempt_id: str, request: Request, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.QuizAttempt).where(
            models.QuizAttempt.id == attempt_id,
            models.QuizAttempt.student_id == user["id"]
        )
    )
    attempt = result.scalars().first()
    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")
    if attempt.status == "submitted":
        raise HTTPException(status_code=400, detail="Already submitted")

    quiz_r = await session.execute(select(models.Quiz).where(models.Quiz.id == attempt.quiz_id))
    quiz = quiz_r.scalars().first()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")

    questions_r = await session.execute(select(models.Question).where(models.Question.quiz_id == quiz.id).order_by(models.Question.id))
    questions = questions_r.scalars().all()

    answers_r = await session.execute(select(models.QuizAnswer).where(models.QuizAnswer.attempt_id == attempt_id))
    answers_map = {a.question_id: a for a in answers_r.scalars().all()}

    score = 0.0
    results = []
    total_marks = sum(q.marks for q in questions)

    for i, q in enumerate(questions):
        content = q.content or {}
        student_answer = None
        ans_row = answers_map.get(q.id)
        if ans_row:
            student_answer = ans_row.code_submitted

        is_correct = False
        marks_awarded = 0.0
        q_type = q.type

        if q_type in ("mcq", "mcq-single", "boolean"):
            correct_ans = content.get("correctAnswer") or content.get("correct_answer")
            if student_answer is not None and student_answer == str(correct_ans):
                is_correct = True
                marks_awarded = q.marks
        elif q_type in ("multiple", "mcq-multiple"):
            import json
            correct_set = set(content.get("correctAnswers") or content.get("correct_answers", []))
            try:
                sel = json.loads(student_answer) if student_answer else []
            except Exception:
                sel = []
            if correct_set == set(sel):
                is_correct = True
                marks_awarded = q.marks
        elif q_type == "short":
            expected = str(content.get("expectedAnswer") or content.get("expected_answer", "")).strip()
            if student_answer and expected and student_answer.strip().lower() == expected.lower():
                is_correct = True
                marks_awarded = q.marks
            elif student_answer:
                marks_awarded = round(q.marks * 0.5)
        elif q_type == "coding":
            if student_answer and str(student_answer).strip():
                marks_awarded = round(q.marks * 0.5)

        score += marks_awarded

        # Update the answer row with grading
        if ans_row:
            ans_row.is_correct = is_correct
            ans_row.marks_awarded = marks_awarded

        results.append({
            "question_index": i, "type": q_type,
            "student_answer": student_answer, "is_correct": is_correct,
            "marks_awarded": marks_awarded, "max_marks": q.marks
        })

    percentage = round((score / total_marks) * 100, 1) if total_marks > 0 else 0
    attempt.status = "submitted"
    attempt.final_score = percentage
    attempt.end_time = datetime.now(timezone.utc)
    await session.commit()

    return {
        "id": attempt.id, "quiz_id": attempt.quiz_id, "student_id": attempt.student_id,
        "status": attempt.status, "final_score": attempt.final_score,
        "percentage": percentage, "results": results,
        "submitted_at": attempt.end_time.isoformat()
    }

@app.get("/api/attempts/{attempt_id}/result")
async def get_attempt_result(attempt_id: str, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.QuizAttempt).where(models.QuizAttempt.id == attempt_id))
    attempt = result.scalars().first()
    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")
    if user["role"] == "student" and attempt.student_id != user["id"]:
        raise HTTPException(status_code=403, detail="Not your attempt")
    return {
        "id": attempt.id, "quiz_id": attempt.quiz_id, "student_id": attempt.student_id,
        "status": attempt.status, "final_score": attempt.final_score,
        "start_time": attempt.start_time.isoformat() if attempt.start_time else None,
        "end_time": attempt.end_time.isoformat() if attempt.end_time else None,
    }

@app.get("/api/attempts")
async def list_attempts(quiz_id: Optional[str] = None, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    stmt = select(models.QuizAttempt).where(models.QuizAttempt.status == "submitted")
    if user["role"] == "student":
        stmt = stmt.where(models.QuizAttempt.student_id == user["id"])
    if quiz_id:
        stmt = stmt.where(models.QuizAttempt.quiz_id == quiz_id)
    result = await session.execute(stmt.order_by(models.QuizAttempt.end_time.desc()))
    attempts = result.scalars().all()
    return [{
        "id": a.id, "quiz_id": a.quiz_id, "student_id": a.student_id,
        "status": a.status, "final_score": a.final_score,
        "submitted_at": a.end_time.isoformat() if a.end_time else None
    } for a in attempts]

# ─── Student Search & Profile (HOD / Admin) ───────────────────────────────
@app.get("/api/students/search")
async def search_students(
    q: str = "", 
    department: Optional[str] = None, 
    college: Optional[str] = None, 
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    user: dict = Depends(require_role("hod", "admin", "exam_cell", "teacher")), 
    session: AsyncSession = Depends(get_db)):
    stmt = select(models.User).where(models.User.role == "student")
    if q:
        stmt = stmt.where(
            models.User.name.ilike(f"%{q}%") |
            models.User.profile_data["college_id"].astext.ilike(f"%{q}%")
        )
    result = await session.execute(stmt.order_by(models.User.name).offset(offset).limit(limit))
    students = result.scalars().all()
    return [{"id": s.id, "name": s.name, "email": s.email, "role": s.role, **(s.profile_data or {})} for s in students]

@app.get("/api/students/{student_id}/profile")
async def student_profile(student_id: str, user: dict = Depends(require_role("hod", "admin", "exam_cell", "teacher")), session: AsyncSession = Depends(get_db)):
    student_r = await session.execute(select(models.User).where(models.User.id == student_id))
    student = student_r.scalars().first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    semesters_r = await session.execute(
        select(models.SemesterGrade)
        .where(models.SemesterGrade.student_id == student_id)
        .order_by(models.SemesterGrade.semester.asc())
    )
    from collections import defaultdict
    sem_map = defaultdict(list)
    for row in semesters_r.scalars().all():
        sem_map[row.semester].append({"course_id": row.course_id, "grade": row.grade, "credits": row.credits_earned})
    semesters = [{"semester": sem, "subjects": subjs} for sem, subjs in sorted(sem_map.items())]
    attempts_r = await session.execute(
        select(models.QuizAttempt)
        .where(models.QuizAttempt.student_id == student_id, models.QuizAttempt.status == "submitted")
        .order_by(models.QuizAttempt.end_time.desc())
    )
    attempts = attempts_r.scalars().all()
    marks_r = await session.execute(
        select(models.MarkEntry).where(
            models.MarkEntry.student_id == student_id
        )
    )
    marks_rows = marks_r.scalars().all()
    mid_marks = [{
        "course_id": r.course_id, "exam_type": r.exam_type,
        "marks_obtained": r.marks_obtained, "max_marks": r.max_marks
    } for r in marks_rows]
    return {
        "student": {"id": student.id, "name": student.name, "email": student.email, **(student.profile_data or {})},
        "semesters": semesters,
        "quiz_attempts": [{"quiz_id": a.quiz_id, "score": a.final_score, "percentage": a.final_score,
                           "submitted_at": a.end_time.isoformat() if a.end_time else ""} for a in attempts[:10]],
        "mid_marks": mid_marks
    }

# ─── Semester Results ────────────────────────────────────────────────────────────
@app.get("/api/results/semester/{student_id}")
async def get_semester_results(student_id: str, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    if user["role"] == "student" and user["id"] != student_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    result = await session.execute(
        select(models.SemesterGrade)
        .where(models.SemesterGrade.student_id == student_id)
        .order_by(models.SemesterGrade.semester.asc())
    )
    rows = result.scalars().all()
    from collections import defaultdict
    sem_map = defaultdict(list)
    for row in rows:
        sem_map[row.semester].append({"course_id": row.course_id, "grade": row.grade, "credits": row.credits_earned})
    return [
        {"student_id": student_id, "semester": sem, "subjects": subjects}
        for sem, subjects in sorted(sem_map.items())
    ]

@app.post("/api/results/semester")
async def create_semester_result(req: SemesterResultCreate, user: dict = Depends(require_role("teacher", "admin")), session: AsyncSession = Depends(get_db)):
    for subj in req.subjects:
        row = models.SemesterGrade(
            student_id=req.student_id,
            semester=req.semester,
            course_id=subj.get("code", subj.get("name", "UNKNOWN")),
            grade=subj.get("grade", "O"),
            credits_earned=int(subj.get("credits", 3)),
        )
        session.add(row)
    await session.commit()
    return {"message": "Semester result saved", "semester": req.semester, "student_id": req.student_id}

# ─── Analytics & Leaderboard ──────────────────────────────────────────────
@app.get("/api/analytics/student/{student_id}")
async def student_analytics(student_id: str, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    if user["role"] == "student" and user["id"] != student_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    attempts_r = await session.execute(
        select(models.QuizAttempt)
        .where(models.QuizAttempt.student_id == student_id, models.QuizAttempt.status == "submitted")
        .order_by(models.QuizAttempt.end_time.asc())
    )
    attempts = attempts_r.scalars().all()
    semesters_r = await session.execute(
        select(models.SemesterGrade)
        .where(models.SemesterGrade.student_id == student_id)
        .order_by(models.SemesterGrade.semester.asc())
    )
    sem_rows = semesters_r.scalars().all()
    total_quizzes = len(attempts)
    avg_score = round(sum(a.final_score or 0 for a in attempts) / total_quizzes, 1) if total_quizzes > 0 else 0
    best_score = max((a.final_score or 0 for a in attempts), default=0)
    quiz_trend = [{
        "date": a.end_time.isoformat() if a.end_time else "",
        "score": a.final_score or 0, "quiz": a.quiz_id
    } for a in attempts[-10:]]
    from collections import defaultdict
    sem_map = defaultdict(list)
    for row in sem_rows:
        sem_map[row.semester].append({"course_id": row.course_id, "grade": row.grade, "credits": row.credits_earned})
    semesters = [{"semester": sem, "subjects": subjs} for sem, subjs in sorted(sem_map.items())]
    return {
        "total_quizzes": total_quizzes, "avg_score": avg_score, "best_score": best_score,
        "latest_cgpa": 0, "quiz_trend": quiz_trend, "subject_averages": {},
        "semesters": semesters
    }

@app.get("/api/analytics/teacher/class-results")
async def class_results_analytics(user: dict = Depends(require_role("teacher", "hod", "exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    # Fetch faculty assignments from the dedicated table
    stmt = select(models.FacultyAssignment)
    if user["role"] == "teacher":
        stmt = stmt.where(models.FacultyAssignment.teacher_id == user["id"])
    result = await session.execute(stmt)
    assignments = result.scalars().all()

    assigned_classes = []
    class_details: dict = {}
    for a in assignments:
        class_key = f"{a.subject_code}_{a.batch}_{a.section}"
        if any(c.get("class_key") == class_key for c in assigned_classes):
            continue
        assigned_classes.append({
            "id": a.id, "class_key": class_key,
            "section": f"{a.department} {a.batch} {a.section}",
            "department": a.department, "subject": a.subject_name,
            "batch": str(a.batch), "totalStudents": 0
        })
        class_details[class_key] = {"totalStudents": 0, "department": a.department,
                                    "batch": str(a.batch), "section": str(a.section)}

    # Quiz results from QuizAttempt
    attempts_r = await session.execute(
        select(models.QuizAttempt).where(models.QuizAttempt.status == "submitted")
    )
    all_attempts = attempts_r.scalars().all()
    quiz_results: dict = {}
    for at in all_attempts:
        quiz_results.setdefault(at.quiz_id, {"completed": 0, "total_score": 0.0, "passed": 0})
        quiz_results[at.quiz_id]["completed"] += 1
        quiz_results[at.quiz_id]["total_score"] += at.final_score or 0
        if (at.final_score or 0) >= 40:
            quiz_results[at.quiz_id]["passed"] += 1

    final_quiz_results = {}
    for qid, stat in quiz_results.items():
        avg = round(stat["total_score"] / stat["completed"], 1) if stat["completed"] > 0 else 0
        final_quiz_results[qid] = [{
            "id": qid, "completed": stat["completed"], "avgScore": avg,
            "passRate": round((stat["passed"] / stat["completed"]) * 100) if stat["completed"] > 0 else 0,
            "topPerformers": []
        }]

    return {"assignedClasses": assigned_classes, "quizResults": final_quiz_results, "midMarks": {}}

@app.get("/api/analytics/teacher/quiz-results/{quiz_id}")
async def get_quiz_detailed_analytics(quiz_id: str, department: str = "", batch: str = "", section: str = "", user: dict = Depends(require_role("teacher", "hod", "exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.User).where(models.User.role == "student")
    result = await session.execute(stmt)
    students = result.scalars().all()
    attempts_r = await session.execute(
        select(models.QuizAttempt).where(models.QuizAttempt.quiz_id == quiz_id)
    )
    attempts_map = {a.student_id: a for a in attempts_r.scalars().all()}
    results = []
    for s in students:
        attempt = attempts_map.get(s.id)
        if attempt:
            pct = attempt.final_score or 0
            raw = attempt.status
            time_elapsed = 0
            if attempt.start_time and attempt.end_time:
                time_elapsed = max(0, int((attempt.end_time - attempt.start_time).total_seconds() / 60))
            results.append({
                "id": s.id, "name": s.name,
                "rollNo": (s.profile_data or {}).get("college_id", s.id),
                "scoreValue": pct, "score": f"{pct}%",
                "timeTaken": f"{time_elapsed} mins",
                "status": "In Progress" if raw == "in_progress" else ("Pass" if pct >= 40 else "Fail"),
                "raw_status": raw
            })
        else:
            results.append({
                "id": s.id, "name": s.name,
                "rollNo": (s.profile_data or {}).get("college_id", s.id),
                "scoreValue": -1, "score": "-", "timeTaken": "-",
                "status": "Not Attempted", "raw_status": "none"
            })
    results.sort(key=lambda x: str(x["rollNo"]))
    return results

@app.get("/api/leaderboard")
async def get_leaderboard(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    user: dict = Depends(get_current_user), 
    session: AsyncSession = Depends(get_db)):
    from sqlalchemy import func, desc
    stmt = (
        select(
            models.QuizAttempt.student_id,
            func.avg(models.QuizAttempt.final_score).label("avg_score"),
            func.count(models.QuizAttempt.id).label("quizzes_taken"),
            models.User.name,
            models.User.profile_data
        )
        .join(models.User, models.User.id == models.QuizAttempt.student_id)
        .where(models.QuizAttempt.status == "submitted")
        .group_by(models.QuizAttempt.student_id, models.User.id)
        .order_by(desc("avg_score"))
        .offset(offset)
        .limit(limit)
    )
    result = await session.execute(stmt)
    rows = result.all()
    
    leaderboard = []
    for i, row in enumerate(rows):
        leaderboard.append({
            "rank": offset + i + 1, 
            "student_id": row.student_id,
            "name": row.name,
            "college_id": (row.profile_data or {}).get("college_id", ""),
            "avg_score": round(row.avg_score, 1) if row.avg_score else 0, 
            "quizzes_taken": row.quizzes_taken, 
            "cgpa": 0,
        })
    return leaderboard


# ─── Dashboard Stats ───────────────────────────────────────────────────────
@app.get("/api/dashboard/student")
async def student_dashboard(user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    # All submitted attempts
    attempts_r = await session.execute(
        select(models.QuizAttempt)
        .where(models.QuizAttempt.student_id == user["id"], models.QuizAttempt.status == "submitted")
        .order_by(models.QuizAttempt.end_time.desc())
    )
    all_attempts = attempts_r.scalars().all()

    # Active quizzes (Quiz model uses 'type' for subject, 'duration_minutes'; no status/total_marks columns)
    quizzes_r = await session.execute(
        select(models.Quiz).order_by(models.Quiz.created_at.desc())
    )
    active_quizzes_raw = quizzes_r.scalars().all()
    attempted_ids = {a.quiz_id for a in all_attempts}
    active_quizzes = [{
        "id": q.id, "title": q.title, "subject": q.type,
        "duration_mins": q.duration_minutes, "total_marks": 0,
        "already_attempted": q.id in attempted_ids
    } for q in active_quizzes_raw[:10]]

    # In-progress attempts
    prog_r = await session.execute(
        select(models.QuizAttempt)
        .where(models.QuizAttempt.student_id == user["id"], models.QuizAttempt.status == "in_progress")
    )
    in_progress = [{"id": a.id, "quiz_id": a.quiz_id, "status": a.status} for a in prog_r.scalars().all()]

    # Semester data
    sem_r = await session.execute(
        select(models.SemesterGrade)
        .where(models.SemesterGrade.student_id == user["id"])
        .order_by(models.SemesterGrade.semester.desc())
    )
    sem_rows = sem_r.scalars().all()
    latest_sem = sem_rows[0].semester if sem_rows else 0

    total_attempts = len(all_attempts)
    avg = round(sum(a.final_score or 0 for a in all_attempts) / total_attempts, 1) if total_attempts > 0 else 0

    score_trend = [{
        "quiz": f"Quiz {i+1}",
        "score": round(a.final_score or 0, 1),
        "date": a.end_time.strftime("%b %d") if a.end_time else ""
    } for i, a in enumerate(reversed(all_attempts[:15]))]

    # Leaderboard rank via subquery
    all_students_r = await session.execute(
        select(models.QuizAttempt.student_id, models.QuizAttempt.final_score)
        .where(models.QuizAttempt.status == "submitted")
    )
    student_scores: dict = {}
    for row in all_students_r.all():
        sid, sc = row[0], row[1] or 0
        student_scores.setdefault(sid, []).append(sc)
    ranked = sorted(student_scores.items(), key=lambda x: sum(x[1]) / len(x[1]) if x[1] else 0, reverse=True)
    rank = next((i + 1 for i, (sid, _) in enumerate(ranked) if sid == user["id"]), None)
    total_students = len(ranked)

    recent_results = [{
        "id": a.id, "quiz_id": a.quiz_id,
        "final_score": a.final_score,
        "submitted_at": a.end_time.isoformat() if a.end_time else ""
    } for a in all_attempts[:5]]

    activity = [{
        "type": "quiz_result",
        "title": f"Scored {a.final_score or 0:.0f}% on quiz",
        "score": a.final_score or 0,
        "timestamp": a.end_time.isoformat() if a.end_time else ""
    } for a in all_attempts[:10]]

    return {
        "recent_results": recent_results,
        "upcoming_quizzes": active_quizzes,
        "in_progress": in_progress,
        "cgpa": 0, "current_sgpa": 0, "current_semester": latest_sem,
        "total_quizzes": total_attempts, "avg_score": avg,
        "score_trend": score_trend, "rank": rank,
        "total_students": total_students, "weak_topics": [], "activity": activity,
    }

@app.get("/api/dashboard/teacher")
async def teacher_dashboard(user: dict = Depends(require_role("teacher", "admin")), session: AsyncSession = Depends(get_db)):
    quizzes_r = await session.execute(
        select(models.Quiz)
        .where(models.Quiz.faculty_id == user["id"])
        .order_by(models.Quiz.created_at.desc())
    )
    my_quizzes = quizzes_r.scalars().all()
    quiz_list = []
    
    if my_quizzes:
        quiz_ids = [q.id for q in my_quizzes]
        from sqlalchemy import func
        stats_r = await session.execute(
            select(
                models.QuizAttempt.quiz_id,
                func.count(models.QuizAttempt.id).label("attempt_count"),
                func.avg(models.QuizAttempt.final_score).label("avg_score")
            )
            .where(models.QuizAttempt.quiz_id.in_(quiz_ids), models.QuizAttempt.status == "submitted")
            .group_by(models.QuizAttempt.quiz_id)
        )
        stats_map = {row.quiz_id: row for row in stats_r.all()}
        for q in my_quizzes:
            s = stats_map.get(q.id)
            quiz_list.append({
                "id": q.id, "title": q.title, "subject": q.type, "status": "active",
                "attempt_count": s.attempt_count if s else 0,
                "avg_score": round(s.avg_score, 1) if s and s.avg_score else 0
            })
    students_r = await session.execute(select(models.User).where(models.User.role == "student"))
    total_students = len(students_r.scalars().all())
    recent_r = await session.execute(
        select(models.QuizAttempt)
        .where(models.QuizAttempt.status == "submitted")
        .order_by(models.QuizAttempt.end_time.desc())
    )
    recent = recent_r.scalars().all()[:10]
    return {
        "quizzes": quiz_list, "total_students": total_students,
        "recent_submissions": [{"id": r.id, "quiz_id": r.quiz_id, "student_id": r.student_id,
                                "final_score": r.final_score, "submitted_at": r.end_time.isoformat() if r.end_time else ""} for r in recent]
    }

@app.get("/api/dashboard/admin")
async def admin_dashboard(user: dict = Depends(require_role("admin")), session: AsyncSession = Depends(get_db)):
    from sqlalchemy import func
    counts_r = await session.execute(
        select(models.User.role, func.count(models.User.id)).group_by(models.User.role)
    )
    role_counts = {role: cnt for role, cnt in counts_r.all()}
    quizzes_r = await session.execute(select(func.count(models.Quiz.id)))
    total_quizzes = quizzes_r.scalar() or 0
    active_r = await session.execute(select(func.count(models.Quiz.id)).where(models.Quiz.status == "active"))
    active_quizzes = active_r.scalar() or 0
    return {
        "total_students": role_counts.get("student", 0),
        "total_teachers": role_counts.get("teacher", 0),
        "total_hods": role_counts.get("hod", 0),
        "total_exam_cell": role_counts.get("exam_cell", 0),
        "total_quizzes": total_quizzes, "active_quizzes": active_quizzes,
        "departments": [],
    }

# ─── Faculty Assignment Routes (HOD) ───────────────────────────────────────
@app.get("/api/faculty/teachers")
async def list_department_teachers(user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.User).where(
        models.User.college_id == user["college_id"],
        models.User.role.in_(["teacher", "hod"])
    )
    result = await session.execute(stmt)
    teachers = result.scalars().all()
    return [{"id": t.id, "name": t.name, "email": t.email, "role": t.role, **(t.profile_data or {})} for t in teachers]

@app.get("/api/faculty/assignments")
async def list_assignments(user: dict = Depends(require_role("hod", "admin", "teacher")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.FacultyAssignment).where(models.FacultyAssignment.college_id == user["college_id"])
    if user["role"] == "teacher":
        stmt = stmt.where(models.FacultyAssignment.teacher_id == user["id"])
    result = await session.execute(stmt)
    rows = result.scalars().all()
    return [{"id": r.id, "course_id": r.subject_code, "teacher_id": r.teacher_id, "subject_code": r.subject_code, "subject_name": r.subject_name, "department": r.department, "batch": r.batch, "section": r.section, "semester": r.semester} for r in rows]

@app.post("/api/faculty/assignments")
async def create_assignment(req: FacultyAssignment, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    teacher_r = await session.execute(select(models.User).where(models.User.id == req.teacher_id))
    teacher = teacher_r.scalars().first()
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")
    row = models.FacultyAssignment(
        college_id=user["college_id"],
        teacher_id=req.teacher_id,
        subject_code=req.subject_code,
        subject_name=req.subject_name,
        department=req.department,
        batch=req.batch,
        section=req.section,
        semester=req.semester
    )
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return {"id": row.id, "teacher_id": req.teacher_id, "subject_code": req.subject_code,
            "subject_name": req.subject_name, "department": req.department}

@app.delete("/api/faculty/assignments/{assignment_id}")
async def delete_assignment(assignment_id: str, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.FacultyAssignment).where(models.FacultyAssignment.id == assignment_id))
    row = result.scalars().first()
    if not row:
        raise HTTPException(status_code=404, detail="Assignment not found")
    await session.delete(row)
    await session.commit()
    return {"message": "Assignment deleted"}

# ─── Marks Entry Routes (Teacher) ─────────────────────────────────────────
@app.get("/api/marks/my-assignments")
async def my_assignments(user: dict = Depends(require_role("teacher", "hod")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.FacultyAssignment).where(models.FacultyAssignment.teacher_id == user["id"])
    )
    rows = result.scalars().all()
    return [{"id": r.id, "course_id": r.subject_code, "subject_code": r.subject_code, "subject_name": r.subject_name, "department": r.department, "batch": r.batch, "section": r.section, "semester": r.semester} for r in rows]

@app.get("/api/marks/students")
async def get_students_for_marks(department: str, batch: str, section: str, user: dict = Depends(require_role("teacher", "hod", "admin", "exam_cell")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.User).where(
            models.User.college_id == user["college_id"],
            models.User.role == "student",
            models.User.profile_data["department"].astext == department,
            models.User.profile_data["batch"].astext == batch,
            models.User.profile_data["section"].astext == section,
        )
    )
    students = result.scalars().all()
    return [{"id": s.id, "name": s.name, "email": s.email, **(s.profile_data or {})} for s in students]

@app.get("/api/marks/entry/{assignment_id}/{exam_type}")
async def get_mark_entry(assignment_id: str, exam_type: str, component_id: Optional[str] = None, user: dict = Depends(require_role("teacher", "hod")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.MarkEntry).where(
            models.MarkEntry.assignment_id == assignment_id,
            models.MarkEntry.exam_type == exam_type,
            models.MarkEntry.faculty_id == user["id"],
        )
    )
    entries = result.scalars().all()
    entry = None
    if component_id:
        for e in entries:
            if (e.extra_data or {}).get("component_id") == component_id:
                entry = e
                break
    else:
        entry = entries[0] if entries else None
        
    if not entry:
        return None
    return {"id": entry.id, "course_id": entry.course_id, "exam_type": entry.exam_type,
            "max_marks": entry.max_marks, "status": (entry.extra_data or {}).get("status", "draft"),
            "entries": (entry.extra_data or {}).get("entries", [])}

@app.post("/api/marks/entry")
async def save_mark_entry(req: MarkEntrySave, user: dict = Depends(require_permission("marks", "edit")), session: AsyncSession = Depends(get_db)):
    entries_data = [e.dict() for e in req.entries]
    assign_r = await session.execute(
        select(models.FacultyAssignment).where(
            models.FacultyAssignment.id == req.assignment_id
        )
    )
    assignment = assign_r.scalars().first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    existing_r = await session.execute(
        select(models.MarkEntry).where(
            models.MarkEntry.course_id == assignment.subject_code,
            models.MarkEntry.exam_type == req.exam_type,
            models.MarkEntry.faculty_id == user["id"],
        )
    )
    entries = existing_r.scalars().all()
    existing = None
    for e in entries:
        if (e.extra_data or {}).get("component_id") == req.component_id:
            existing = e
            break
    if existing:
        current_status = (existing.extra_data or {}).get("status", "draft")
        if current_status == "approved":
            if not req.revision_reason or not req.revision_reason.strip():
                raise HTTPException(status_code=400, detail="Revision reason is required to edit approved marks")
        if current_status == "submitted":
            raise HTTPException(status_code=400, detail="Cannot edit submitted marks. Wait for approval or rejection.")
        existing.extra_data = {**(existing.extra_data or {}), "entries": entries_data,
                          "status": "draft", "max_marks": req.max_marks, "component_id": req.component_id}
        existing.max_marks = req.max_marks
        await log_audit(session, user["id"], "mark_entry", "update", {"entry_id": existing.id, "course_id": existing.course_id})
        await session.commit()
        return {"id": existing.id, "status": "draft", "entries": entries_data}
    row = models.MarkEntry(
        student_id=user["id"],
        course_id=assignment.subject_code,
        faculty_id=user["id"],
        exam_type=req.exam_type,
        marks_obtained=0,
        max_marks=req.max_marks,
        extra_data={
            "assignment_id": req.assignment_id, "entries": entries_data,
            "status": "draft", "semester": req.semester, "max_marks": req.max_marks,
            "subject_name": assignment.subject_name, "department": assignment.department,
            "batch": assignment.batch, "section": assignment.section,
            "component_id": req.component_id
        }
    )
    session.add(row)
    await log_audit(session, user["id"], "mark_entry", "create", {"exam_type": req.exam_type, "course_id": assignment.subject_code})
    await session.commit()
    await session.refresh(row)
    return {"id": row.id, "status": "draft", "entries": entries_data}

@app.post("/api/marks/submit/{entry_id}")
async def submit_marks(entry_id: str, user: dict = Depends(require_role("teacher", "hod")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.MarkEntry).where(
            models.MarkEntry.id == entry_id,
            models.MarkEntry.faculty_id == user["id"]
        ).with_for_update()
    )
    entry = result.scalars().first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    current_status = (entry.extra_data or {}).get("status", "draft")
    if current_status != "draft":
        raise HTTPException(status_code=400, detail=f"Cannot submit - current status: {current_status}")
    entry.extra_data = {**(entry.extra_data or {}), "status": "submitted",
                   "submitted_at": datetime.now(timezone.utc).isoformat()}
    await log_audit(session, user["id"], "mark_entry", "submit", {"entry_id": entry_id, "course_id": entry.course_id})
    await session.commit()
    return {"message": "Marks submitted for HOD approval"}

@app.get("/api/marks/submissions")
async def list_submissions(status: Optional[str] = None, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    from sqlalchemy import func
    stmt = select(models.MarkEntry)
    
    if status:
        stmt = stmt.where(func.jsonb_extract_path_text(models.MarkEntry.extra_data, 'status') == status)
    else:
        stmt = stmt.where(func.jsonb_extract_path_text(models.MarkEntry.extra_data, 'status').in_(['submitted', 'approved', 'rejected']))
        
    result = await session.execute(stmt)
    submissions = result.scalars().all()
    
    return [{
        "id": r.id, 
        "course_id": r.course_id, 
        "exam_type": r.exam_type,
        "max_marks": r.max_marks,
        "status": (r.extra_data or {}).get("status", "draft"), 
        **(r.extra_data or {})
    } for r in submissions]

@app.post("/api/marks/review/{entry_id}")
async def review_marks(entry_id: str, req: MarkReview, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.MarkEntry).where(models.MarkEntry.id == entry_id))
    entry = result.scalars().first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    current_status = (entry.extra_data or {}).get("status", "draft")
    if current_status != "submitted":
        raise HTTPException(status_code=400, detail="Only submitted marks can be reviewed")
    new_status = "approved" if req.action == "approve" else "rejected"
    entry.extra_data = {**(entry.extra_data or {}), "status": new_status,
                   "reviewed_by": user["id"], "reviewer_name": user["name"],
                   "reviewed_at": datetime.now(timezone.utc).isoformat(), "review_remarks": req.remarks}
    await session.commit()
    return {"message": f"Marks {new_status}"}


# ─── Exam Cell Routes ──────────────────────────────────────────────────────
@app.get("/api/examcell/approved-marks")
async def get_approved_marks(user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.MarkEntry))
    all_rows = result.scalars().all()
    return [{
        "id": r.id, "course_id": r.course_id, "exam_type": r.exam_type,
        "max_marks": r.max_marks, **(r.extra_data or {})
    } for r in all_rows if (r.extra_data or {}).get("status") == "approved"]

@app.post("/api/examcell/endterm")
async def save_endterm(req: EndtermEntry, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    # Store endterm as SemesterGrade rows
    for entry in req.entries:
        sid = entry.get("student_id", "")
        if not sid:
            continue
        row = models.SemesterGrade(
            student_id=sid,
            course_id=req.subject_code,
            semester=req.semester,
            grade=entry.get("grade", "O"),
            credits_earned=int(entry.get("credits", 3)),
        )
        session.add(row)
    await session.commit()
    return {"message": f"Endterm saved for {req.subject_code}"}

@app.get("/api/examcell/endterm")
async def list_endterm(user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.SemesterGrade).order_by(models.SemesterGrade.semester.desc())
    )
    rows = result.scalars().all()
    from collections import defaultdict
    grouped: dict = defaultdict(list)
    for r in rows:
        grouped[(r.course_id, r.semester)].append({
            "student_id": r.student_id, "grade": r.grade, "credits": r.credits_earned
        })
    return [
        {"course_id": cid, "semester": sem, "entries": entries}
        for (cid, sem), entries in grouped.items()
    ]

@app.post("/api/examcell/upload")
async def upload_marks_file(file: UploadFile = File(...), semester: int = Form(...), subject_code: str = Form(...), subject_name: str = Form(...), department: str = Form(...), batch: str = Form(...), section: str = Form(...), user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    # 1. Check for duplicate batch submission
    existing = await session.execute(
        select(models.MarkEntry).where(
            models.MarkEntry.course_id == subject_code,
            models.MarkEntry.exam_type == "endterm_csv",
            models.MarkEntry.extra_data["section"].astext == section,
            models.MarkEntry.extra_data["batch"].astext == batch
        )
    )
    if existing.scalars().first():
        raise HTTPException(status_code=400, detail="CSV for this subject and section already uploaded")

    content = await file.read()
    entries = []
    filename = file.filename.lower()
    try:
        if filename.endswith('.csv'):
            text = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                cid = row.get('college_id', row.get('roll_no', row.get('College ID', ''))).strip().upper()
                if cid:
                    marks_val = float(row.get('marks', row.get('Marks', row.get('score', '0'))) or 0)
                    grade = row.get('grade', row.get('Grade', 'O'))
                    entries.append({"college_id": cid, "marks": marks_val, "grade": grade})
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [str(c.value or '').lower().strip() for c in ws[1]]
            cid_col = next((i for i, h in enumerate(headers) if h in ['college_id', 'roll_no', 'roll number']), 0)
            marks_col = next((i for i, h in enumerate(headers) if h in ['marks', 'score']), 1)
            grade_col = next((i for i, h in enumerate(headers) if h in ['grade']), -1)
            for row in ws.iter_rows(min_row=2, values_only=True):
                cid_val = str(row[cid_col] or '').strip().upper()
                if not cid_val:
                    continue
                entries.append({
                    "college_id": cid_val,
                    "marks": float(row[marks_col] or 0),
                    "grade": str(row[grade_col] or 'O') if grade_col >= 0 else 'O'
                })
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    if not entries:
        raise HTTPException(status_code=400, detail="No valid entries found in file")
        
    # 2. Strict Validation against College Constraints
    validated_entries = []
    for entry in entries:
        student_r = await session.execute(
            select(models.User).where(models.User.profile_data["college_id"].astext == entry["college_id"])
        )
        student = student_r.scalars().first()
        if not student:
            raise HTTPException(status_code=400, detail=f"Student ID {entry['college_id']} not found in database.")
        if student.college_id != user["college_id"] and user["role"] != "super_admin":
            raise HTTPException(status_code=403, detail=f"Student ID {entry['college_id']} belongs to a different college tenant.")
            
        validated_entries.append({
            "student_id": student.id,
            "college_id": entry["college_id"],
            "marks": entry["marks"],
            "grade": entry["grade"]
        })

    # 3. Queue for HOD Approval instead of Silent Mutation
    mark_entry = models.MarkEntry(
        student_id=user["id"], 
        course_id=subject_code,
        faculty_id=user["id"],
        exam_type="endterm_csv",
        marks_obtained=0,
        max_marks=100,
        extra_data={
            "subject_name": subject_name,
            "department": department,
            "batch": batch,
            "section": section,
            "semester": semester,
            "filename": filename,
            "status": "pending_hod_approval",
            "entries": validated_entries
        }
    )
    session.add(mark_entry)
    
    await log_audit(session, user["id"], "exam_cell_csv_upload", "create", {"subject": subject_code, "entries": len(validated_entries)})
    await session.commit()
    
    return {"message": f"Successfully uploaded {len(validated_entries)} entries. Pending HOD approval.", "entry_id": mark_entry.id}

@app.post("/api/examcell/publish/{entry_id}")
async def publish_results(entry_id: str, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.MarkEntry)
        .where(models.MarkEntry.id == entry_id)
        .with_for_update()
    )
    entry = result.scalars().first()
    
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
        
    current_status = (entry.extra_data or {}).get("status", "draft")
    if current_status != "approved":
        raise HTTPException(status_code=400, detail=f"Only approved marks can be published (current: {current_status})")
        
    entries = (entry.extra_data or {}).get("entries", [])
    if not entries:
        raise HTTPException(status_code=400, detail="No student marks found in entry")
        
    metadata = entry.extra_data or {}
    subject_code = entry.course_id
    semester = int(metadata.get("semester", 1))
    max_marks = float(metadata.get("max_marks", 100))
    
    college_r = await session.execute(
        select(models.College).where(models.College.id == user["college_id"])
    )
    college = college_r.scalars().first()
    college_settings = college.settings if college and college.settings else {}
    grade_scale = college_settings.get("grade_scale") or DEFAULT_GRADE_SCALE
    
    semester_grades = []
    for student_entry in entries:
        student_id = student_entry.get("student_id")
        marks = float(student_entry.get("marks", 0.0))
        pct = (marks / max_marks) * 100 if max_marks > 0 else 0
        
        # Calculate grade from percentage dynamically
        grade = "F"
        for boundary in sorted(grade_scale, key=lambda x: x.get("min_pct", 0), reverse=True):
            if pct >= boundary.get("min_pct", 0):
                grade = boundary.get("grade", "F")
                break
        
        if not student_id:
            continue
            
        existing = await session.execute(
            select(models.SemesterGrade).where(
                models.SemesterGrade.student_id == student_id,
                models.SemesterGrade.semester == semester,
                models.SemesterGrade.course_id == subject_code
            )
        )
        if existing.scalars().first():
            continue
            
        credits_assigned = 3
        assignment_id = metadata.get("assignment_id")
        if assignment_id:
            ass_r = await session.execute(select(models.FacultyAssignment).where(models.FacultyAssignment.id == assignment_id))
            ass = ass_r.scalars().first()
            if ass:
                credits_assigned = ass.credits
                
        semester_grades.append(models.SemesterGrade(
            student_id=student_id,
            semester=semester,
            course_id=subject_code,
            grade=grade,
            credits_earned=credits_assigned
        ))
        
    if semester_grades:
        session.add_all(semester_grades)

    entry.extra_data = {
        **(entry.extra_data or {}),
        "status": "published",
        "published_at": datetime.now(timezone.utc).isoformat(),
        "published_by": user["id"]
    }
    
    await log_audit(session, user["id"], "mark_entry", "publish", {
        "entry_id": entry_id, "course_id": subject_code, 
        "student_count": len(semester_grades)
    })
    await session.commit()
    
    return {
        "message": f"Published {len(semester_grades)} student grades",
        "entry_id": entry_id,
        "published_count": len(semester_grades)
    }

# ─── Timetable Routes (HOD) ───────────────────────────────────────────────────────────
@app.get("/api/timetable")
async def get_timetable(section: str, semester: int = 3, user: dict = Depends(require_role("hod", "admin", "teacher", "student")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.Timetable).where(
            models.Timetable.college_id == user["college_id"],
            models.Timetable.semester == semester,
        )
    )
    slots = result.scalars().all()
    return [{
        "id": s.id, "day": s.day, "time_slot": s.time_slot, "room": s.room,
        "semester": s.semester, "course_id": s.course_id, "faculty_id": s.faculty_id,
        "department_id": s.department_id
    } for s in slots]

@app.post("/api/timetable")
async def save_timetable_slot(req: TimetableSlot, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    # 1. Validate against FacultyAssignment (Explicit DB Link resolving point 5)
    assignment_r = await session.execute(
        select(models.FacultyAssignment).where(
            models.FacultyAssignment.teacher_id == req.teacher_id,
            models.FacultyAssignment.subject_code == req.subject_code,
            models.FacultyAssignment.college_id == user["college_id"]
        )
    )
    assignment = assignment_r.scalars().first()
    if not assignment:
        raise HTTPException(status_code=400, detail="Invalid timetable slot: Teacher is not assigned to this subject code.")

    existing_r = await session.execute(
        select(models.Timetable).where(
            models.Timetable.college_id == user["college_id"],
            models.Timetable.day == req.day,
            models.Timetable.time_slot == str(req.period),
            models.Timetable.semester == req.semester,
        )
    )
    existing = existing_r.scalars().first()
    if existing:
        existing.course_id = req.subject_code
        existing.faculty_id = req.teacher_id
        existing.room = ""
        await session.commit()
        return {"id": existing.id, "day": existing.day, "time_slot": existing.time_slot, "course_id": existing.course_id}
        
    slot = models.Timetable(
        college_id=user["college_id"],
        department_id="", 
        course_id=req.subject_code,
        faculty_id=req.teacher_id,
        semester=req.semester,
        day=req.day,
        time_slot=str(req.period),
        room="",
    )
    # Look up real department UUID based on the assignment, fundamentally decoupling it from User session state.
    dept_r = await session.execute(
        select(models.Department).where(
            models.Department.college_id == user["college_id"],
            or_(models.Department.code == assignment.department, models.Department.name == assignment.department)
        )
    )
    dept = dept_r.scalars().first()
    if dept:
        slot.department_id = dept.id
    else:
        raise HTTPException(status_code=400, detail=f"Department '{assignment.department}' not found. Create it first.")

    session.add(slot)
    await session.commit()
    await session.refresh(slot)
    return {"id": slot.id, "day": slot.day, "time_slot": slot.time_slot, "course_id": slot.course_id}

@app.delete("/api/timetable/{slot_id}")
async def delete_timetable_slot(slot_id: str, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.Timetable).where(models.Timetable.id == slot_id))
    slot = result.scalars().first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    await session.delete(slot)
    await session.commit()
    return {"message": "Slot deleted"}

# ─── Announcement Routes ─────────────────────────────────────────────────────
@app.get("/api/announcements")
async def list_announcements(user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    dept = user.get("department", "")
    role = user["role"]
    stmt = select(models.Announcement).where(
        models.Announcement.college_id == user["college_id"]
    ).order_by(models.Announcement.created_at.desc()).limit(50)
    result = await session.execute(stmt)
    rows = result.scalars().all()
    out = []
    for a in rows:
        details = a.details or {}
        vis = details.get("visibility", "all")
        a_dept = details.get("department", "")
        if a_dept and a_dept != dept:
            continue
        if role == "student" and vis not in ("all", "students"):
            continue
        if role == "teacher" and vis not in ("all", "faculty"):
            continue
        out.append({
            "id": a.id, "title": a.title, "message": a.message,
            "priority": a.priority, "visibility": vis,
            "department": a_dept,
            "posted_by": details.get("posted_by", ""),
            "created_at": a.created_at.isoformat() if a.created_at else ""
        })
    return out

@app.post("/api/announcements")
async def create_announcement(req: AnnouncementCreate, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    row = models.Announcement(
        college_id=user["college_id"],
        title=req.title,
        message=req.message,
        priority=req.priority,
        details={
            "visibility": req.visibility,
            "department": user.get("department", ""),
            "posted_by": user["name"],
            "posted_by_id": user["id"]
        }
    )
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return {"id": row.id, "title": row.title, "message": row.message, "priority": row.priority}

@app.delete("/api/announcements/{announcement_id}")
async def delete_announcement(announcement_id: str, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(select(models.Announcement).where(models.Announcement.id == announcement_id))
    row = result.scalars().first()
    if not row:
        raise HTTPException(status_code=404, detail="Announcement not found")
    await session.delete(row)
    await session.commit()
    return {"message": "Announcement deleted"}

# ─── At-Risk Students (HOD) ────────────────────────────────────────────────
@app.get("/api/hod/at-risk-students")
async def get_at_risk_students(
    cgpa_threshold: float = 5.0, 
    backlog_threshold: int = 2,
    user: dict = Depends(require_role("hod", "admin")), 
    session: AsyncSession = Depends(get_db)
):
    students_r = await session.execute(
        select(models.User).where(
            models.User.college_id == user["college_id"],
            models.User.role == "student"
        )
    )
    students = students_r.scalars().all()
    at_risk = []
    
    for student in students:
        grades_r = await session.execute(
            select(models.SemesterGrade).where(models.SemesterGrade.student_id == student.id)
        )
        grades = grades_r.scalars().all()
        
        if not grades:
            continue
            
        total_credits = sum(g.credits_earned for g in grades)
        total_points = sum(grade_to_points(g.grade) * g.credits_earned for g in grades)
        backlogs = sum(1 for g in grades if g.grade == "F")
        cgpa = round(total_points / total_credits, 2) if total_credits > 0 else 0
        
        if cgpa < cgpa_threshold or backlogs >= backlog_threshold:
            if cgpa < (cgpa_threshold - 1.5) or backlogs >= (backlog_threshold + 2):
                severity = "critical"
            else:
                severity = "warning"
                
            at_risk.append({
                "id": student.id, 
                "name": student.name,
                "college_id": (student.profile_data or {}).get("college_id", ""),
                "section": (student.profile_data or {}).get("section", ""),
                "batch": (student.profile_data or {}).get("batch", ""),
                "cgpa": cgpa,
                "cgpa_threshold": cgpa_threshold,
                "backlogs": backlogs,
                "backlog_threshold": backlog_threshold,
                "severity": severity
            })
            
    at_risk.sort(key=lambda x: (x["severity"] != "critical", x["cgpa"]))
    return at_risk


@app.get("/api/dashboard/hod")
async def hod_dashboard(user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    from sqlalchemy import func
    teachers_r = await session.execute(
        select(func.count(models.User.id)).where(
            models.User.college_id == user["college_id"], models.User.role == "teacher"
        )
    )
    students_r = await session.execute(
        select(func.count(models.User.id)).where(
            models.User.college_id == user["college_id"], models.User.role == "student"
        )
    )
    assignments_r = await session.execute(
        select(func.count(models.FacultyAssignment.id))
    )
    # Pending/approved from JSONB status field
    mark_entries_r = await session.execute(
        select(models.MarkEntry)
    )
    all_entries = mark_entries_r.scalars().all()
    pending = sum(1 for e in all_entries if (e.extra_data or {}).get("status") == "submitted")
    approved = sum(1 for e in all_entries if (e.extra_data or {}).get("status") == "approved")
    recent = [{
        "id": e.id, "course_id": e.course_id, "exam_type": e.exam_type,
        "status": (e.extra_data or {}).get("status", "draft"), "activity_type": "marks_review"
    } for e in all_entries[-15:]]
    return {
        "total_teachers": teachers_r.scalar() or 0,
        "total_students": students_r.scalar() or 0,
        "total_assignments": assignments_r.scalar() or 0,
        "pending_reviews": pending, "approved_count": approved,
        "recent_submissions": recent
    }

@app.get("/api/dashboard/exam_cell")
async def exam_cell_dashboard(user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    mark_entries_r = await session.execute(
        select(models.MarkEntry)
    )
    all_entries = mark_entries_r.scalars().all()
    approved = sum(1 for e in all_entries if (e.extra_data or {}).get("status") == "approved")
    sem_r = await session.execute(select(models.SemesterGrade))
    all_grades = sem_r.scalars().all()
    return {
        "total_approved_midterms": approved,
        "total_endterm": len(all_grades),
        "total_published": len(all_grades),
        "total_draft": 0,
        "recent_entries": []
    }

# ─── Code Execution (Sandboxed) ────────────────────────────────────────────
import asyncio as _asyncio
import re as _re
import shutil as _shutil

# Dangerous patterns per language — block common attack vectors
import ast as _ast

_BLOCKED_PATTERNS = {
    "python": [
        # File system
        r"\bimport\s+os\b", r"\bimport\s+subprocess\b", r"\bimport\s+shutil\b", r"\bimport\s+pathlib\b",
        r"\bopen\s*\(", r"\bos\.", r"\bsubprocess\.", r"\bshutil\.", r"\bpathlib\.",
        # Network
        r"\bimport\s+socket\b", r"\bimport\s+http\b", r"\bimport\s+urllib\b", r"\bimport\s+requests\b",
        r"\bsocket\.", r"\burllib\.", r"\brequests\.",
        # Code execution
        r"\b__import__\s*\(", r"\bexec\s*\(", r"\beval\s*\(", r"\bcompile\s*\(", 
        # Introspection/reflection
        r"\bglobals\s*\(", r"\blocals\s*\(", r"\bgetattr\s*\(", r"\bhasattr\s*\(",
        r"\btype\s*\(", r"\bvars\s*\(", r"\bdir\s*\(", r"\binspect\.",
        # Module introspection
        r"\b__dict__\b", r"\b__code__\b", r"\b__class__\b", r"\b__bases__\b",
    ],
    "javascript": [
        r"require\s*\(\s*['\"]child_process", r"require\s*\(\s*['\"]fs",
        r"require\s*\(\s*['\"]net", r"require\s*\(\s*['\"]http",
        r"\bprocess\.exit", r"\bprocess\.env",
        r"\bexecSync\b", r"\bspawnSync\b",
        r"\beval\s*\(", r"\bFunction\s*\(",
    ],
    "java": [
        r"Runtime\.getRuntime", r"ProcessBuilder", r"System\.exit", 
        r"java\.io\.File", r"java\.net\.", r"java\.nio\.file",
    ],
    "c": [
        r"#include\s*<unistd\.h>", r"#include\s*<sys/",
        r"\bsystem\s*\(", r"\bexecl?[vpe]*\s*\(", r"\bfork\s*\(", 
        r"\bpopen\s*\(", r"\bsocket\s*\(",
    ],
}
_BLOCKED_PATTERNS["cpp"] = _BLOCKED_PATTERNS["c"]

def _validate_code_ast(code: str, language: str):
    if language.lower() != "python": return
    try:
        tree = _ast.parse(code)
    except SyntaxError as e:
        raise HTTPException(status_code=400, detail=f"Syntax error: {e}")
        
    BLOCKED_IMPORTS = {"os", "subprocess", "shutil", "socket", "http", "urllib", "requests", "pathlib", "inspect", "__builtin__"}
    BLOCKED_CALLS = {"__import__", "exec", "eval", "compile", "globals", "locals", "getattr", "hasattr", "type", "vars", "dir", "open"}
    
    class CodeValidator(_ast.NodeVisitor):
        def __init__(self):
            self.violations = []
        def visit_Import(self, node):
            for alias in node.names:
                if alias.name.split(".")[0] in BLOCKED_IMPORTS:
                    self.violations.append(f"Blocked import: {alias.name}")
            self.generic_visit(node)
        def visit_ImportFrom(self, node):
            if node.module and node.module.split(".")[0] in BLOCKED_IMPORTS:
                self.violations.append(f"Blocked import: {node.module}")
            self.generic_visit(node)
        def visit_Call(self, node):
            func_name = getattr(node.func, "id", getattr(node.func, "attr", None))
            if func_name in BLOCKED_CALLS:
                self.violations.append(f"Blocked call: {func_name}()")
            self.generic_visit(node)
        def visit_Attribute(self, node):
            obj_name = getattr(node.value, "id", None)
            if obj_name in BLOCKED_IMPORTS:
                self.violations.append(f"Blocked access: {obj_name}.{node.attr}")
            if node.attr.startswith("__") and node.attr.endswith("__"):
                self.violations.append(f"Blocked dunder: {node.attr}")
            self.generic_visit(node)
            
    validator = CodeValidator()
    validator.visit(tree)
    if validator.violations:
        raise HTTPException(status_code=400, detail=f"Code blocked: {'; '.join(validator.violations[:3])}")

def _validate_code(code: str, language: str):
    patterns = _BLOCKED_PATTERNS.get(language, [])
    for pattern in patterns:
        match = _re.search(pattern, code)
        if match:
            raise HTTPException(status_code=400, detail=f"Blocked: '{match.group()}' is not allowed.")
    _validate_code_ast(code, language)

async def _run_process(cmd, stdin_data="", timeout=5, cwd=None):
    """Run a subprocess in a thread executor (Windows-compatible)."""
    import concurrent.futures
    import asyncio as _asyncio
    def _exec():
        try:
            result = subprocess.run(
                cmd,
                input=stdin_data or None,
                capture_output=True, text=True,
                timeout=timeout,
                cwd=cwd,
            )
            return result.stdout, result.stderr, result.returncode
        except subprocess.TimeoutExpired:
            return "", f"Execution timed out ({timeout} second limit)", -1
        except Exception as e:
            return "", str(e), -1

    loop = _asyncio.get_event_loop()
    with concurrent.futures.ThreadPoolExecutor() as pool:
        return await loop.run_in_executor(pool, _exec)

import tenacity

TIMEOUT_CONFIG = {
    "python": 15.0,
    "javascript": 15.0,
    "c": 65.0,
    "cpp": 65.0,
    "java": 50.0,
    "sql": 15.0,
}

@app.post("/api/code/execute")
@limiter.limit("30/minute")
async def execute_code(request: Request, req: CodeExecuteRequest, user: dict = Depends(get_current_user)):
    _validate_code(req.code, req.language.lower())
    lang_timeout = TIMEOUT_CONFIG.get(req.language.lower(), 65.0)

    @tenacity.retry(
        stop=tenacity.stop_after_attempt(3),
        wait=tenacity.wait_exponential(multiplier=1, min=1, max=10),
        retry=tenacity.retry_if_exception_type((httpx.RequestError, httpx.TimeoutException)),
        reraise=True
    )
    async def _do_request():
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                f"{code_runner_url}/run",
                json={"language": req.language, "code": req.code, "test_input": req.test_input},
                timeout=lang_timeout
            )
            if resp.status_code == 400:
                raise HTTPException(status_code=400, detail=resp.json().get("detail", "Error"))
            if resp.status_code in [502, 503, 504]:
                raise httpx.RequestError("Code runner gateway timeout")
            if resp.status_code != 200:
                return {"output": "", "error": "Code runner service unavailable", "exit_code": -1}
            return resp.json()

    try:
        return await _do_request()
    except HTTPException:
        raise
    except Exception as e:
        return {"output": "", "error": str(e)[:500], "exit_code": -1}

# ─── Health ─────────────────────────────────────────────────────────────────
@app.get("/api/health")
async def health():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

@app.get("/api/test-sentry")
async def test_sentry():
    """Intentional crash to verify Sentry connection and PII scrubbing."""
    division_by_zero = 1 / 0
    return {"message": "If you see this, Sentry didn't catch the intentional crash!"}

# NOTE: seed_data() is superseded by the @app.on_event('startup') handler above
# which seeds the admin user via PostgreSQL. This stub is kept for compatibility.
async def seed_data():
    pass


_placements_store: list = []  # legacy compat — routes now use DB exclusively

@app.get("/api/placements/student")
async def student_placements(user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    college_id = user.get("college_id", "")
    dept = user.get("department", "")
    email = user.get("email", "")
    stmt = select(models.Placement).where(models.Placement.college_id == college_id)
    result = await session.execute(stmt)
    rows = result.scalars().all()
    out = []
    for p in rows:
        details = p.details or {}
        candidates = details.get("candidates", [])
        is_shortlisted = any(c.get("college_id") == college_id or c.get("email") == email for c in candidates)
        open_dept = details.get("department", "ALL")
        is_open = details.get("open_to_all") and open_dept in (dept, "ALL", "all", "")
        if is_shortlisted or is_open:
            out.append({"id": p.id, "company": p.company, "role": p.role, "package": p.package, "date": p.date, **{k: v for k, v in details.items() if k != "candidates"}})
    out.sort(key=lambda x: x.get("drive_date", ""))
    return out

@app.post("/api/placements")
async def create_placement(req: dict, user: dict = Depends(require_role("admin", "hod")), session: AsyncSession = Depends(get_db)):
    row = models.Placement(
        college_id=user["college_id"],
        company=req.get("company", ""),
        role=req.get("role", ""),
        package=req.get("package", ""),
        date=req.get("date", ""),
        details={k: v for k, v in req.items() if k not in ("company", "role", "package", "date")}
    )
    row.details = {**(row.details or {}), "created_by": user["id"]}
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return {"id": row.id, "company": row.company, "role": row.role, "date": row.date}

@app.get("/api/placements")
async def list_placements(user: dict = Depends(require_role("admin", "hod", "teacher")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.Placement).where(models.Placement.college_id == user["college_id"])
    result = await session.execute(stmt)
    rows = result.scalars().all()
    return [{"id": p.id, "company": p.company, "role": p.role, "package": p.package, "date": p.date, **(p.details or {})} for p in rows]

# ─── Phase 3: Student T&P Workflows ──────────────────────────────────────────

@app.get("/api/student/drives")
async def get_eligible_student_drives(user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.PlacementDrive).where(
        models.PlacementDrive.college_id == user["college_id"],
        models.PlacementDrive.status.in_(["upcoming", "ongoing"])
    )
    res = await session.execute(stmt)
    drives = res.scalars().all()
    
    # Pre-fetch Student Profile Data to match CGPA and active Backlogs
    pd = user.get("profile_data") or {}
    student_cgpa = float(pd.get("cgpa", 0))
    student_backlogs = int(pd.get("active_backlogs", 0))
    student_dept = pd.get("department", "")
    
    eligible = []
    for d in drives:
        crit = d.eligibility_criteria or {}
        if crit:
            # Check conditions
            if float(crit.get("min_cgpa", 0)) > student_cgpa:
                continue
            if "max_backlogs" in crit and int(crit.get("max_backlogs", 99)) < student_backlogs:
                continue
            if "allowed_departments" in crit and crit["allowed_departments"]:
                if student_dept not in crit["allowed_departments"]:
                    continue
        eligible.append(d)
        
    return eligible

@app.post("/api/student/drives/{drive_id}/apply")
async def apply_for_placement_drive(drive_id: str, user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.PlacementDrive).where(models.PlacementDrive.id == drive_id, models.PlacementDrive.college_id == user["college_id"])
    res = await session.execute(stmt)
    drive = res.scalars().first()
    
    if not drive:
        raise HTTPException(status_code=404, detail="Drive not found")
        
    if drive.status not in ["upcoming", "ongoing"]:
        raise HTTPException(status_code=400, detail="Drive is not accepting applications")
        
    pd = user.get("profile_data") or {}
    student_cgpa = float(pd.get("cgpa", 0))
    student_backlogs = int(pd.get("active_backlogs", 0))
    student_dept = pd.get("department", "")
    
    crit = drive.eligibility_criteria or {}
    if crit:
        if float(crit.get("min_cgpa", 0)) > student_cgpa:
            raise HTTPException(status_code=403, detail="Ineligible: CGPA below requirement")
        if "max_backlogs" in crit and int(crit.get("max_backlogs", 99)) < student_backlogs:
            raise HTTPException(status_code=403, detail="Ineligible: Exceeds active backlog limits")
        if "allowed_departments" in crit and crit["allowed_departments"] and student_dept not in crit["allowed_departments"]:
            raise HTTPException(status_code=403, detail="Ineligible: Department not permitted")
            
    if drive.linked_quiz_id:
        quiz_r = await session.execute(
            select(models.QuizAttempt).where(
                models.QuizAttempt.quiz_id == drive.linked_quiz_id,
                models.QuizAttempt.student_id == user["id"],
                models.QuizAttempt.status == "submitted"
            )
        )
        attempt = quiz_r.scalars().first()
        if not attempt:
            raise HTTPException(status_code=403, detail="Pre-screening test not completed or not passed.")
        if drive.quiz_threshold and float(attempt.final_score or 0) < drive.quiz_threshold:
            raise HTTPException(status_code=403, detail="Pre-screening test not completed or not passed.")

    dup_r = await session.execute(
        select(models.PlacementApplication).where(
            models.PlacementApplication.drive_id == drive_id,
            models.PlacementApplication.student_id == user["id"]
        )
    )
    if dup_r.scalars().first():
        raise HTTPException(status_code=400, detail="Already applied")
        
    appl = models.PlacementApplication(
        college_id=user["college_id"],
        student_id=user["id"],
        drive_id=drive_id,
        status="registered"
    )
    session.add(appl)
    await session.commit()
    return {"message": "Successfully applied to drive", "application_id": appl.id}

@app.delete("/api/student/drives/{drive_id}/withdraw")
async def withdraw_application(drive_id: str, user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.PlacementApplication).where(
        models.PlacementApplication.drive_id == drive_id,
        models.PlacementApplication.student_id == user["id"]
    )
    res = await session.execute(stmt)
    appl = res.scalars().first()
    
    if not appl:
        raise HTTPException(status_code=404, detail="Application not found")
        
    if appl.status != "registered":
        raise HTTPException(status_code=400, detail="Already shortlisted — contact T&P Officer")
        
    await session.delete(appl)
    await session.commit()
    return {"message": "Application withdrawn successfully"}

@app.get("/api/student/applications")
async def get_student_application_history(user: dict = Depends(require_role("student")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.PlacementApplication).where(models.PlacementApplication.student_id == user["id"])
    res = await session.execute(stmt)
    return res.scalars().all()

# ─── Phase 4 & 5: TPO Workflows & Analytics ──────────────────────────────────────

@app.get("/api/tpo/drives/{drive_id}/applicants")
async def get_drive_applicants(drive_id: str, user: dict = Depends(require_role("tp_officer", "admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.PlacementApplication).where(
        models.PlacementApplication.drive_id == drive_id,
        models.PlacementApplication.college_id == user["college_id"]
    )
    res = await session.execute(stmt)
    return res.scalars().all()

class ShortlistRequest(BaseModel):
    student_ids: List[str]

@app.put("/api/tpo/drives/{drive_id}/shortlist")
async def bulk_shortlist(drive_id: str, req: ShortlistRequest, user: dict = Depends(require_role("tp_officer", "admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.PlacementApplication).where(
        models.PlacementApplication.drive_id == drive_id,
        models.PlacementApplication.college_id == user["college_id"],
        models.PlacementApplication.student_id.in_(req.student_ids)
    )
    res = await session.execute(stmt)
    apps = res.scalars().all()
    for app in apps:
        if app.status == "registered":
            app.status = "shortlisted"
            
    await session.commit()
    return {"message": f"Successfully shortlisted {len(apps)} candidates"}

class ResultRequest(BaseModel):
    student_id: str
    round_name: str
    result: str # pass or fail
    remarks: Optional[str] = ""

@app.put("/api/tpo/drives/{drive_id}/results")
async def append_round_result(drive_id: str, req: ResultRequest, user: dict = Depends(require_role("tp_officer", "admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.PlacementApplication).where(
        models.PlacementApplication.drive_id == drive_id,
        models.PlacementApplication.student_id == req.student_id,
        models.PlacementApplication.college_id == user["college_id"]
    )
    res = await session.execute(stmt)
    app = res.scalars().first()
    
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")
        
    res_list = app.round_results or []
    from datetime import datetime
    res_list.append({
        "round": req.round_name,
        "result": req.result,
        "remarks": req.remarks,
        "evaluated_at": datetime.utcnow().isoformat()
    })
    app.round_results = res_list
    
    if req.result == "fail":
        app.status = "rejected"
        
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(app, "round_results")
    await session.commit()
    return {"message": "Result appended"}

class SelectRequest(BaseModel):
    student_id: str
    ctc: float
    role: str
    joining_date: Optional[str] = None
    location: Optional[str] = None
    offer_url: Optional[str] = None

@app.put("/api/tpo/drives/{drive_id}/select")
async def select_candidate(drive_id: str, req: SelectRequest, user: dict = Depends(require_role("tp_officer", "admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.PlacementApplication).where(
        models.PlacementApplication.drive_id == drive_id,
        models.PlacementApplication.student_id == req.student_id,
        models.PlacementApplication.college_id == user["college_id"]
    )
    res = await session.execute(stmt)
    app = res.scalars().first()
    
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")
        
    app.status = "selected"
    app.offer_details = {
        "ctc": req.ctc,
        "role": req.role,
        "joining_date": req.joining_date,
        "location": req.location,
        "offer_url": req.offer_url,
        "is_accepted": False
    }
    await session.commit()
    return {"message": "Candidate selected and generated offer metadata"}

@app.get("/api/tpo/statistics")
async def get_tpo_statistics(user: dict = Depends(require_role("tp_officer", "admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.PlacementApplication).where(
        models.PlacementApplication.college_id == user["college_id"],
        models.PlacementApplication.status == "selected"
    )
    res = await session.execute(stmt)
    selected = res.scalars().all()
    
    total_ctc = 0
    highest = 0
    for s in selected:
        if s.offer_details:
            ctc = float(s.offer_details.get("ctc") or 0)
            total_ctc += ctc
            if ctc > highest:
                highest = ctc
                
    avg = total_ctc / len(selected) if selected else 0
    return {
        "total_selected": len(selected),
        "highest_package": highest,
        "average_package": avg
    }

from fastapi.responses import StreamingResponse
import io
import openpyxl

@app.get("/api/tpo/statistics/export")
async def export_tpo_statistics(user: dict = Depends(require_role("tp_officer", "admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.PlacementApplication).where(
        models.PlacementApplication.college_id == user["college_id"],
        models.PlacementApplication.status == "selected"
    )
    res = await session.execute(stmt)
    selected = res.scalars().all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Placement Record"
    ws.append(["Student ID", "Drive ID", "Role", "CTC (LPA)", "Location"])
    
    for s in selected:
        details = s.offer_details or {}
        ws.append([
            s.student_id,
            s.drive_id,
            details.get("role", ""),
            details.get("ctc", ""),
            details.get("location", "")
        ])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]), 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=placement_statistics.xlsx"}
    )

# ─── Coding Challenges ─────────────────────────────────────────────────────────
@app.get("/api/challenges")
async def get_challenges(page: int = 1, limit: int = 20, difficulty: str = "", topic: str = "", session: AsyncSession = Depends(get_db)):
    stmt = select(models.CodingChallenge)
    if difficulty:
        stmt = stmt.where(models.CodingChallenge.difficulty == difficulty)
    
    result = await session.execute(stmt)
    all_challenges = result.scalars().all()
    if topic:
        all_challenges = [c for c in all_challenges if topic in (c.topics or [])]
        
    start = (page - 1) * limit
    page_data = all_challenges[start : start + limit]
    
    return {
        "data": [{"id": c.id, "title": c.title, "description": c.description, "difficulty": c.difficulty, "topics": c.topics, "language_support": c.language_support} for c in page_data],
        "total": len(all_challenges),
        "page": page,
        "limit": limit
    }

@app.get("/api/challenges/stats")
async def get_challenge_stats(user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    stmt = select(models.CodingChallenge).join(
        models.ChallengeProgress,
        models.CodingChallenge.id == models.ChallengeProgress.challenge_id
    ).where(
        models.ChallengeProgress.student_id == user["id"],
        models.ChallengeProgress.status == "completed"
    )
    cr = await session.execute(stmt)
    solved_challenges = cr.scalars().all()
    
    easy = medium = hard = 0
    topics_count = {}
    
    for c in solved_challenges:
        d = getattr(c, "difficulty", "easy").lower()
        if d == "easy": easy += 1
        elif d == "medium": medium += 1
        elif d == "hard": hard += 1
        for t in (c.topics or []):
            topics_count[t] = topics_count.get(t, 0) + 1
            
    return {"total_solved": len(solved_challenges), "difficulty": {"Easy": easy, "Medium": medium, "Hard": hard}, "topics": topics_count}

@app.post("/api/challenges/submit")
@limiter.limit("30/minute")
async def submit_challenge(request: Request, req: ChallengeSubmit, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    cr = await session.execute(select(models.CodingChallenge).where(models.CodingChallenge.id == req.challenge_id))
    challenge = cr.scalars().first()
    if not challenge:
        raise HTTPException(status_code=404, detail="Challenge not found")
        
    init_sql_script = ""
    # Inject the SQL datasets for DataLemur style tests
    if req.language.lower() == "sql" and challenge.init_code:
        init_sql_script = challenge.init_code.get("sql", "")
        
    lang_timeout = TIMEOUT_CONFIG.get(req.language.lower(), 60.0)

    @tenacity.retry(
        stop=tenacity.stop_after_attempt(3),
        wait=tenacity.wait_exponential(multiplier=1, min=1, max=10),
        retry=tenacity.retry_if_exception_type((httpx.RequestError, httpx.TimeoutException)),
        reraise=True
    )
    async def _do_request():
        async with httpx.AsyncClient() as client:
            return await client.post(
                f"{code_runner_url}/run",
                json={"language": req.language, "code": req.code, "test_input": init_sql_script},
                timeout=lang_timeout
            )

    try:
        resp = await _do_request()
        
        if resp.status_code == 200:
            result = resp.json()
            exit_code = result.get("exit_code", -1)
            
            is_success = (exit_code == 0)
            
            if is_success:
                pr = await session.execute(select(models.ChallengeProgress).where(models.ChallengeProgress.student_id == user["id"], models.ChallengeProgress.challenge_id == req.challenge_id))
                prog = pr.scalars().first()
                if not prog:
                    prog = models.ChallengeProgress(student_id=user["id"], challenge_id=req.challenge_id, status="completed", language_used=req.language)
                    session.add(prog)
                else:
                    prog.status = "completed"
                await session.commit()
                
            return {"output": result.get("output", ""), "error": result.get("error", ""),
                    "exit_code": exit_code, "success": is_success}
        return {"error": "Code runner service error", "exit_code": -1, "success": False}
    except Exception as e:
        return {"error": str(e), "exit_code": -1, "success": False}

# ─── Phase 7: HOD Governance & Mentorship ────────────────────────────────────

@app.get("/api/hod/assignments/class-in-charge")
async def get_class_in_charges(user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.ClassInCharge, models.User).join(
            models.User, models.ClassInCharge.faculty_id == models.User.id
        ).where(
            models.ClassInCharge.college_id == user["college_id"]
        )
    )
    return [{
        "id": c.id, "faculty_id": c.faculty_id, "faculty_name": u.name,
        "department": c.department, "batch": c.batch, "section": c.section, "semester": c.semester, "academic_year": c.academic_year
    } for c, u in result.all()]

@app.post("/api/hod/assignments/class-in-charge")
async def create_class_in_charge(req: ClassInChargeCreate, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    created = []
    academic_year = await get_current_academic_year(session, user["college_id"])
    for fac_id in req.faculty_ids:
        cic = models.ClassInCharge(
            college_id=user["college_id"],
            faculty_id=fac_id,
            department=req.department,
            batch=req.batch,
            section=req.section,
            semester=req.semester,
            academic_year=academic_year
        )
        session.add(cic)
        created.append(cic)
    try:
        await session.commit()
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail="Assignment conflicts with an existing assignment or invalid data.")
    return {"message": f"{len(created)} class in-charges assigned"}

@app.delete("/api/hod/assignments/class-in-charge/{assignment_id}")
async def delete_class_in_charge(assignment_id: str, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    await session.execute(delete(models.ClassInCharge).where(
        models.ClassInCharge.id == assignment_id, models.ClassInCharge.college_id == user["college_id"]
    ))
    await session.commit()
    return {"message": "Assignment deleted"}

@app.get("/api/hod/assignments/mentors")
async def get_mentor_assignments(user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.MentorAssignment, models.User).join(
            models.User, models.MentorAssignment.student_id == models.User.id
        ).where(
            models.MentorAssignment.college_id == user["college_id"],
            models.MentorAssignment.is_active == True
        )
    )
    
    assignments = result.all()
    if not assignments: return []
    
    faculty_ids = {m.faculty_id for m, _ in assignments}
    fac_result = await session.execute(select(models.User.id, models.User.name).where(models.User.id.in_(faculty_ids)))
    faculty_names = {u_id: name for u_id, name in fac_result.all()}
    
    return [{
        "id": m.id, "faculty_id": m.faculty_id, "faculty_name": faculty_names.get(m.faculty_id, "Unknown"),
        "student_id": m.student_id, "student_name": s.name, "academic_year": m.academic_year
    } for m, s in assignments]

@app.post("/api/hod/assignments/mentors")
async def create_mentor_assignments(req: MentorAssignmentCreate, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    academic_year = await get_current_academic_year(session, user["college_id"])
    created = 0
    for stud_id in req.student_ids:
        existing_r = await session.execute(select(models.MentorAssignment).where(
            models.MentorAssignment.college_id == user["college_id"],
            models.MentorAssignment.student_id == stud_id,
            models.MentorAssignment.academic_year == academic_year,
            models.MentorAssignment.is_active == True
        ))
        if existing_r.scalars().first():
            continue 
        
        m = models.MentorAssignment(
            college_id=user["college_id"],
            faculty_id=req.faculty_id,
            student_id=stud_id,
            academic_year=academic_year,
            is_active=True
        )
        session.add(m)
        created += 1
    
    await session.commit()
    return {"message": f"{created} students assigned to mentor"}

@app.delete("/api/hod/assignments/mentors/{assignment_id}")
async def deactivate_mentor_assignment(assignment_id: str, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    r = await session.execute(select(models.MentorAssignment).where(
        models.MentorAssignment.id == assignment_id, models.MentorAssignment.college_id == user["college_id"]
    ))
    m = r.scalars().first()
    if not m:
        raise HTTPException(status_code=404, detail="Assignment not found")
    m.is_active = False
    await session.commit()
    return {"message": "Assignment deactivated"}

@app.get("/api/faculty/students/{student_id}/progression")
async def get_student_progression(student_id: str, user: dict = Depends(get_current_user), session: AsyncSession = Depends(get_db)):
    if user["role"] not in ["hod", "teacher", "faculty", "admin"]:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    academic_year = await get_current_academic_year(session, user["college_id"])
    
    if user["role"] in ["teacher", "faculty"]:
        m_r = await session.execute(select(models.MentorAssignment).where(
            models.MentorAssignment.faculty_id == user["id"],
            models.MentorAssignment.student_id == student_id,
            models.MentorAssignment.academic_year == academic_year,
            models.MentorAssignment.is_active == True
        ))
        is_mentor = m_r.scalars().first() is not None
        
        stud_r = await session.execute(select(models.User).where(models.User.id == student_id))
        stud = stud_r.scalars().first()
        is_cic = False
        if stud and stud.profile_data:
            c_r = await session.execute(select(models.ClassInCharge).where(
                models.ClassInCharge.faculty_id == user["id"],
                models.ClassInCharge.academic_year == academic_year,
                models.ClassInCharge.department == (stud.profile_data or {}).get("department", ""),
                models.ClassInCharge.batch == (stud.profile_data or {}).get("batch", ""),
                models.ClassInCharge.section == (stud.profile_data or {}).get("section", "")
            ))
            is_cic = c_r.scalars().first() is not None
            
        if not (is_mentor or is_cic):
            raise HTTPException(status_code=403, detail="Not authorized to view this student's progression data")
            
    p_r = await session.execute(select(models.StudentProgression).where(
        models.StudentProgression.student_id == student_id,
        models.StudentProgression.college_id == user["college_id"]
    ).order_by(models.StudentProgression.created_at.desc()))
    
    return [{
        "id": p.id,
        "academic_year": p.academic_year,
        "progression_type": p.progression_type,
        "details": p.details,
        "created_at": p.created_at.isoformat()
    } for p in p_r.scalars().all()]
    
@app.post("/api/hod/progression")
async def create_progression(req: StudentProgressionCreate, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    academic_year = await get_current_academic_year(session, user["college_id"])
    prog = models.StudentProgression(
        college_id=user["college_id"],
        student_id=req.student_id,
        academic_year=academic_year,
        progression_type=req.progression_type,
        details=req.details
    )
    session.add(prog)
    await session.commit()
    await session.refresh(prog)
    return {"id": prog.id, "message": "Progression record created"}

@app.delete("/api/hod/progression/{prog_id}")
async def delete_progression(prog_id: str, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    await session.execute(delete(models.StudentProgression).where(
        models.StudentProgression.id == prog_id, models.StudentProgression.college_id == user["college_id"]
    ))
    await session.commit()
    return {"message": "Progression record deleted"}

@app.on_event("startup")
async def startup():
    import asyncio
    max_retries = 5
    for attempt in range(1, max_retries + 1):
        try:
            await _seed_db()
            return
        except Exception as e:
            if attempt == max_retries:
                print(f"[startup] FATAL: Could not connect to database after {max_retries} attempts: {e}")
                raise
            wait = 2 ** attempt  # 2s, 4s, 8s, 16s
            print(f"[startup] DB connection failed (attempt {attempt}/{max_retries}), retrying in {wait}s... ({e})")
            await asyncio.sleep(wait)

async def _seed_db():
    """PostgreSQL startup: schema is managed by Alembic migrations.
    Seed order matters:
      1. Upsert College (GNI) → get college.id UUID
      2. Upsert default Departments linked to college.id
      3. Upsert Admin user with college_id = college.id
    """
    from database import AsyncSessionLocal
    async with AsyncSessionLocal() as session:
        admin_college_id_str = os.environ.get("ADMIN_COLLEGE_ID", "A001")
        admin_pwd = os.environ.get("ADMIN_PASSWORD", "admin123")
        college_name = os.environ.get("COLLEGE_NAME", "Guru Nanak Institutions Technical Campus")

        # 1. Upsert College
        college_r = await session.execute(
            select(models.College).where(models.College.name == college_name)
        )
        college = college_r.scalars().first()
        if not college:
            college = models.College(name=college_name, domain="gnitc.ac.in")
            session.add(college)
            await session.commit()
            await session.refresh(college)
            print(f"[startup] College '{college_name}' seeded with id={college.id}")
        else:
            print(f"[startup] College '{college_name}' already exists (id={college.id})")

        # 2. Upsert default Departments
        default_depts = [
            {"name": "Electronics & Telematics", "code": "ET"},
            {"name": "Computer Science & Engineering", "code": "CSE"},
            {"name": "Electrical & Electronics Engineering", "code": "EEE"},
            {"name": "Mechanical Engineering", "code": "ME"},
            {"name": "Civil Engineering", "code": "CE"},
            {"name": "Information Technology", "code": "IT"},
        ]
        for dept in default_depts:
            dept_r = await session.execute(
                select(models.Department).where(
                    models.Department.college_id == college.id,
                    models.Department.code == dept["code"]
                )
            )
            if not dept_r.scalars().first():
                session.add(models.Department(college_id=college.id, name=dept["name"], code=dept["code"]))
        await session.commit()
        print(f"[startup] Default departments ensured for {college_name}")

        # 3. Upsert Admin user with real college FK
        result = await session.execute(
            select(models.User).where(
                models.User.profile_data["college_id"].astext == admin_college_id_str
            )
        )
        existing = result.scalars().first()
        if not existing:
            admin = models.User(
                name="GNI Admin",
                email="admin@gni.edu",
                password_hash=hash_password(admin_pwd),
                role="admin",
                college_id=college.id,
                profile_data={"college_id": admin_college_id_str, "college": college_name, "department": "Administration"},
            )
            session.add(admin)
            await session.commit()
            print(f"[startup] Admin user '{admin_college_id_str}' seeded with college_id={college.id}")
        else:
            # Fix existing admin if college_id is None
            if existing.college_id is None:
                existing.college_id = college.id
                await session.commit()
                print(f"[startup] Fixed admin college_id: None → {college.id}")
            else:
                print(f"[startup] Admin '{admin_college_id_str}' already exists. Skipping seed.")




class AttendanceOverride(BaseModel):
    date: str
    period_no: int
    status: str
    reason: str

@app.put("/api/faculty/attendance/override/{subject_code}/{student_id}")
async def override_student_attendance(
    subject_code: str,
    student_id: str,
    req: AttendanceOverride,
    user: dict = Depends(require_role("teacher", "hod")),
    session: AsyncSession = Depends(get_db)
):
    """Selective attendance override for a single student."""
    from sqlalchemy import cast, Date
    stmt = select(models.AttendanceRecord).where(
        models.AttendanceRecord.student_id == student_id,
        models.AttendanceRecord.subject_code == subject_code,
        models.AttendanceRecord.date == cast(req.date, Date),
        models.AttendanceRecord.period_no == req.period_no,
        models.AttendanceRecord.is_deleted == False
    )
    result = await session.execute(stmt)
    record = result.scalars().first()
    
    if not record:
        raise HTTPException(status_code=404, detail="Attendance slot not found")
        
    record.status = req.status
    # We can track the override metadata in extra_data manually for an audit trail
    record.is_late_entry = True # Force late entry flag for HOD review since it's an override 
    await session.commit()
    return {"message": "Override applied successfully."}

class ActivityPermissionCreate(BaseModel):
    activity_type: str
    title: str
    description: Optional[str] = None
    date: str
    venue: Optional[str] = None

@app.post("/api/faculty/activities")
async def request_activity_permission(req: ActivityPermissionCreate, user: dict = Depends(require_role("teacher", "hod")), session: AsyncSession = Depends(get_db)):
    """Faculty submits a request for extra-curricular activity."""
    row = models.ActivityPermission(
        college_id=user["college_id"],
        faculty_id=user["id"],
        activity_type=req.activity_type,
        title=req.title,
        description=req.description,
        date=req.date,
        venue=req.venue,
        phase="pre_event",
        status="pending"
    )
    session.add(row)
    await session.commit()
    return {"message": "Activity permission requested."}

class OutOfCampusCreate(BaseModel):
    destination: str
    purpose: str
    departure_time: str
    return_time: str

@app.post("/api/faculty/out-of-campus")
async def request_out_of_campus(req: OutOfCampusCreate, user: dict = Depends(require_role("teacher", "hod")), session: AsyncSession = Depends(get_db)):
    row = models.OutOfCampusPermission(
        college_id=user["college_id"],
        faculty_id=user["id"],
        destination=req.destination,
        purpose=req.purpose,
        departure_time=req.departure_time,
        return_time=req.return_time,
        status="pending"
    )
    session.add(row)
    await session.commit()
    return {"message": "Out of campus permission requested."}

class FreePeriodRequestCreate(BaseModel):
    period_slot_id: str
    date: str
    reason: str

@app.post("/api/faculty/free-periods")
async def request_free_period(req: FreePeriodRequestCreate, user: dict = Depends(require_role("teacher", "hod")), session: AsyncSession = Depends(get_db)):
    row = models.FreePeriodRequest(
        college_id=user["college_id"],
        faculty_id=user["id"],
        period_slot_id=req.period_slot_id,
        date=req.date,
        reason=req.reason,
        status="pending"
    )
    session.add(row)
    await session.commit()
    return {"message": "Free period requested."}


class ActivityReview(BaseModel):
    action: str  # approve or reject

@app.put("/api/hod/activity-permissions/{permission_id}/review")
async def review_activity_permission(permission_id: str, req: ActivityReview, user: dict = Depends(require_role("hod", "admin")), session: AsyncSession = Depends(get_db)):
    result = await session.execute(
        select(models.ActivityPermission).where(
            models.ActivityPermission.id == permission_id,
            models.ActivityPermission.college_id == user["college_id"]
        )
    )
    perm = result.scalars().first()
    if not perm:
        raise HTTPException(status_code=404, detail="Activity permission request not found")
        
    perm.status = "approved" if req.action == "approve" else "rejected"
    perm.hod_approved_by = user["id"]
    await session.commit()
    return {"message": f"Activity permission {perm.status}"}


class ManualRegistrationCreate(BaseModel):
    student_id: str
    semester: int
    academic_year: str
    subject_code: str
    is_arrear: bool = False

@app.post("/api/examcell/registrations/manual-add")
async def manual_add_registration(req: ManualRegistrationCreate, user: dict = Depends(require_role("exam_cell", "admin")), session: AsyncSession = Depends(get_db)):
    """Allow Exam Cell to explicitly inject a student registration"""
    row = models.CourseRegistration(
        student_id=req.student_id,
        semester=req.semester,
        academic_year=req.academic_year,
        subject_code=req.subject_code,
        is_arrear=req.is_arrear,
        status="approved" # Automatically approved
    )
    session.add(row)
    await log_audit(session, user["id"], "course_registration", "manual_add", {"student_id": req.student_id, "subject_code": req.subject_code})
    await session.commit()
    return {"message": "Registration added manually."}

@app.get("/api/admin/reports/faculty-research")
async def get_faculty_research_report(user: dict = Depends(require_role("admin", "principal")), session: AsyncSession = Depends(get_db)):
    """NAAC Criterion 3.3 Research Aggregation"""
    stmt = select(models.User).where(
        models.User.college_id == user["college_id"],
        models.User.role.in_(["teacher", "faculty"])
    )
    result = await session.execute(stmt)
    faculty = result.scalars().all()
    
    report = []
    for f in faculty:
        profile = f.profile_data or {}
        research = profile.get("research", [])
        publications = profile.get("publications", [])
        
        report.append({
            "faculty_id": f.id,
            "name": f.name,
            "department": profile.get("department", "N/A"),
            "research_projects": research,
            "publications": publications,
            "total_publications": len(publications),
            "total_research": len(research)
        })
    return report
import csv
import io
import openpyxl

@app.post("/api/admin/users/bulk-import")
async def bulk_import_users(role: str, file: UploadFile = File(...), user: dict = Depends(require_role("admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    if role not in ["student", "faculty"]:
        raise HTTPException(status_code=400, detail="Role must be student or faculty")
        
    contents = await file.read()
    data = []
    
    if file.filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(contents.decode('utf-8')))
        for row in reader:
            data.append(row)
    elif file.filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
        sheet = wb.active
        headers = [str(cell.value) if cell.value else f"col_{i}" for i, cell in enumerate(sheet[1])]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                data.append(dict(zip(headers, row)))
    else:
        raise HTTPException(status_code=400, detail="Only CSV and XLSX files are supported")
        
    created = 0
    errors = []
    
    for row in data:
        student_id = row.get("id") or row.get("student_id") or row.get("roll_number")
        if not student_id:
            continue
            
        student_id = str(student_id).strip()
        # Test Verification isolated trigger
        if not student_id.isalnum() and '-' not in student_id:
            raise HTTPException(status_code=400, detail=f"Invalid format for Student ID: {student_id}. Import aborted.")

        existing_r = await session.execute(select(models.User).where(models.User.id == student_id))
        if existing_r.scalars().first():
            continue 
            
        new_user = models.User(
            id=student_id,
            name=row.get("name", "Unknown"),
            email=row.get("email", f"{student_id}@student.college.edu"),
            role=role,
            college_id=user["college_id"],
            password_hash=get_password_hash("password123"),
            profile_data={
                "department": row.get("department", ""),
                "batch": row.get("batch", ""),
                "force_password_change": True
            }
        )
        session.add(new_user)
        created += 1
        
    await session.commit()
    return {"message": f"Successfully imported {created} {role}s"}

@app.post("/api/admin/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, admin: dict = Depends(require_role("admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    user_req = await session.execute(select(models.User).where(models.User.id == user_id, models.User.college_id == admin["college_id"]))
    target = user_req.scalars().first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
        
    new_temp = secrets.token_urlsafe(8)
    target.password_hash = get_password_hash(new_temp)
    
    pd = target.profile_data or {}
    pd["force_password_change"] = True
    target.profile_data = pd
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(target, "profile_data")
    
    await log_audit(session, admin["id"], "user", "reset_password", {"target_id": user_id})
    await session.commit()
    
    return {"message": "Password reset successfully", "temporary_password": new_temp}

from fastapi.responses import StreamingResponse

@app.get("/api/admin/users/export")
async def export_users(role: str = "student", batch: Optional[str] = None, user: dict = Depends(require_role("admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.User).where(models.User.college_id == user["college_id"], models.User.role == role)
    res = await session.execute(stmt)
    users = res.scalars().all()
    
    if batch:
        users = [u for u in users if (u.profile_data or {}).get("batch") == batch]
        
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Name", "Email", "Department", "Batch", "Role"])
    for u in users:
        pd = u.profile_data or {}
        writer.writerow([u.id, u.name, u.email, pd.get("department", ""), pd.get("batch", ""), u.role])
        
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=users_export_{role}.csv"})
from collections import defaultdict
from sqlalchemy.orm import selectinload

@app.get("/api/admin/academic-calendars/{calendar_id}/year-view")
async def get_calendar_year_view(calendar_id: str, admin: dict = Depends(require_role("admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    cal_r = await session.execute(select(models.AcademicCalendar).where(models.AcademicCalendar.id == calendar_id, models.AcademicCalendar.college_id == admin["college_id"]))
    cal = cal_r.scalars().first()
    if not cal:
        raise HTTPException(status_code=404, detail="Calendar not found")
        
    year_map = {}
    events = cal.events or []
    for evt in events:
        d = evt.get("date")
        if d:
            year_map[d] = {
                "title": evt.get("title", ""),
                "type": evt.get("type", "holiday")
            }
    return year_map

@app.get("/api/admin/staff-profiles/pending")
async def get_pending_staff_profiles(admin: dict = Depends(require_role("admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.User).where(
        models.User.college_id == admin["college_id"],
        models.User.role.in_(["faculty", "teacher"])
    )
    res = await session.execute(stmt)
    faculty = res.scalars().all()
    
    pending = []
    for f in faculty:
        pd = f.profile_data or {}
        has_pending = False
        for section in ["education", "experience", "research"]:
            for record in pd.get(section, []):
                if record.get("status") == "submitted":
                    has_pending = True
                    break
        if has_pending:
            pending.append({"id": f.id, "name": f.name, "department": pd.get("department", "")})
            
    return pending

class ProfileReview(BaseModel):
    section: str
    record_index: int
    action: str
    remarks: str = ""

@app.put("/api/admin/staff-profiles/{user_id}/review")
async def review_staff_profile(user_id: str, req: ProfileReview, admin: dict = Depends(require_role("admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    user_req = await session.execute(select(models.User).where(models.User.id == user_id, models.User.college_id == admin["college_id"]))
    target = user_req.scalars().first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
        
    pd = target.profile_data or {}
    records = pd.get(req.section, [])
    if req.record_index < 0 or req.record_index >= len(records):
        raise HTTPException(status_code=400, detail="Invalid record index")
        
    records[req.record_index]["status"] = "approved" if req.action == "approve" else "rejected"
    records[req.record_index]["remarks"] = req.remarks
    
    pd[req.section] = records
    target.profile_data = pd
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(target, "profile_data")
    await session.commit()
    return {"message": "Profile record reviewed"}

@app.get("/api/admin/registration-windows/{window_id}/unregistered")
async def get_unregistered_students(window_id: str, admin: dict = Depends(require_role("admin", "super_admin", "exam_cell")), session: AsyncSession = Depends(get_db)):
    win_r = await session.execute(select(models.RegistrationWindow).where(models.RegistrationWindow.id == window_id, models.RegistrationWindow.college_id == admin["college_id"]))
    window = win_r.scalars().first()
    if not window:
        raise HTTPException(status_code=404, detail="Window not found")
        
    # Get all students
    all_studs_r = await session.execute(select(models.User).where(models.User.role == "student", models.User.college_id == admin["college_id"]))
    all_studs = {u.id: u for u in all_studs_r.scalars().all()}
    
    # Get registrations
    regs_r = await session.execute(select(models.CourseRegistration).where(models.CourseRegistration.semester == window.semester, models.CourseRegistration.academic_year == window.academic_year))
    reg_student_ids = {r.student_id for r in regs_r.scalars().all()}
    
    unregistered = []
    for sid, user in all_studs.items():
        if sid not in reg_student_ids:
            pd = user.profile_data or {}
            unregistered.append({
                "id": user.id,
                "name": user.name,
                "department": pd.get("department", ""),
                "batch": pd.get("batch", "")
            })
            
    return unregistered

@app.get("/api/admin/activity-reports")
async def get_activity_reports(admin: dict = Depends(require_role("admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    # Blocked resolved: ActivityPermission exists natively now
    reports_r = await session.execute(
        select(models.ActivityPermission).where(
            models.ActivityPermission.college_id == admin["college_id"],
            models.ActivityPermission.phase == "post_event",
            models.ActivityPermission.hod_report_decision == "approved"
        )
    )
    return reports_r.scalars().all()
@app.get("/api/admin/timetable/conflicts")
async def get_timetable_conflicts(academic_year: str, session: AsyncSession = Depends(get_db)):
    # Simple conflict check returning duplicates based on day/period/faculty across all depts
    stmt = select(models.Timetable).where(models.Timetable.academic_year == academic_year)
    res = await session.execute(stmt)
    slots = res.scalars().all()
    
    seen = {}
    conflicts = []
    for s in slots:
        key = (s.faculty_id, s.day_of_week, s.period_no)
        if key in seen:
            conflicts.append({
                "faculty_id": s.faculty_id,
                "day": s.day_of_week,
                "period": s.period_no,
                "dept_1": seen[key],
                "dept_2": s.department_id
            })
        else:
            seen[key] = s.department_id
    return conflicts

@app.put("/api/admin/timetable/{department_id}/approve")
async def approve_timetable(department_id: str, academic_year: str, semester: int, admin: dict = Depends(require_role("admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    stmt = select(models.TimetableApproval).where(
        models.TimetableApproval.department_id == department_id,
        models.TimetableApproval.academic_year == academic_year,
        models.TimetableApproval.semester == semester
    )
    res = await session.execute(stmt)
    approval = res.scalars().first()
    
    from datetime import datetime
    if approval:
        approval.is_approved = True
        approval.approved_by = admin["id"]
        approval.approved_at = datetime.utcnow()
    else:
        approval = models.TimetableApproval(
            college_id=admin["college_id"],
            department_id=department_id,
            academic_year=academic_year,
            semester=semester,
            is_approved=True,
            approved_by=admin["id"],
            approved_at=datetime.utcnow()
        )
        session.add(approval)
        
    await session.commit()
    return {"message": "Timetable approved"}
@app.get("/api/admin/cia-config/coverage")
async def get_cia_config_coverage(semester: int, academic_year: str, admin: dict = Depends(require_role("admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    # Get all subjects assigned in this semester
    assign_r = await session.execute(
        select(models.FacultyAssignment).where(
            models.FacultyAssignment.college_id == admin["college_id"],
            models.FacultyAssignment.semester == semester,
            models.FacultyAssignment.academic_year == academic_year
        )
    )
    allocations = assign_r.scalars().all()
    assigned_subjects = list({a.subject_code for a in allocations})
    
    # Get CIA configurations for these subjects
    config_r = await session.execute(
        select(models.SubjectCIAConfig).where(
            models.SubjectCIAConfig.college_id == admin["college_id"],
            models.SubjectCIAConfig.semester == semester,
            models.SubjectCIAConfig.academic_year == academic_year
        )
    )
    configured_subjects = {c.subject_code for c in config_r.scalars().all()}
    
    unconfigured = [s for s in assigned_subjects if s not in configured_subjects]
    
    return {
        "total_subjects": len(assigned_subjects),
        "configured_subjects": len(configured_subjects),
        "unconfigured_subjects": len(unconfigured),
        "missing_codes": unconfigured
    }

@app.get("/api/admin/dashboard-stats")
async def get_admin_dashboard_stats(admin: dict = Depends(require_role("admin", "super_admin")), session: AsyncSession = Depends(get_db)):
    # Basic counts
    studs_r = await session.execute(select(func.count(models.User.id)).where(models.User.college_id == admin["college_id"], models.User.role == "student"))
    student_count = studs_r.scalar() or 0
    
    facs_r = await session.execute(select(func.count(models.User.id)).where(models.User.college_id == admin["college_id"], models.User.role.in_(["faculty", "teacher"])))
    faculty_count = facs_r.scalar() or 0
    
    depts_r = await session.execute(select(func.count(models.Department.id)).where(models.Department.college_id == admin["college_id"]))
    dept_count = depts_r.scalar() or 0
    
    return {
        "total_students": student_count,
        "total_faculty": faculty_count,
        "total_departments": dept_count,
        "system_health": "Optimal",
        "pending_approvals_estimate": 0 # Real-time aggregation logic placeholder
    }
